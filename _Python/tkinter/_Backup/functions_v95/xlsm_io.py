"""
xlsm_io.py — write live Tk app state into the .xlsm workbook.

The popup .exe files in dist/ all read parameters from the workbook via
openpyxl.  The app holds the user's edits in StringVars, so we must push
those StringVar values into the workbook BEFORE launching any popup EXE.
After the EXE finishes the existing refresh_from_xlsm() call pulls
results back into the UI.

Cell map mirrors XLSM_CELL_MAP in main.py.  Source-term (rows 8..18) and
field-data (rows 34..40) lists are added programmatically.

Public API:
    push_state_to_xlsm(app, xlsm_path, sheet_name) -> bool
"""
from __future__ import annotations
import os
from typing import Any


CELL_MAP = {
    # Section 1 (Site / Date)
    "v_site":         "B4",
    "v_date":         "E4",
    # Section 2 (Model Configuration)
    "v_x_size":       "E11",
    "v_y_size":       "E12",
    "v_z_size":       "E13",
    "v_sw_width":     "E15",
    "v_sw_thick":     "E16",
    "v_yr_start":     "E18",
    "v_yr_end":       "E19",
    "v_run_time":     "M16",
    # Section 3
    "v_darcy":        "C22",
    "v_porf":         "G22",
    # Section 4
    "v_lowk_media":   "K26",
    "v_lowk_por":     "K27",
    "v_lowk_tort":    "K28",
    # Section 5
    "v_pfaa1":        "E38",
    "v_pfaa2":        "G38",
    "v_pfaa3":        "K38",   # Detailed only
    "v_pfaa4":        "M38",   # Detailed only
    "v_ret_trans1":   "E39",
    "v_ret_lowk1":    "E40",
    "v_ret_trans2":   "G39",
    "v_ret_lowk2":    "G40",
    "v_ret_trans3":   "K39",
    "v_ret_lowk3":    "K40",
    "v_ret_trans4":   "M39",
    "v_ret_lowk4":    "M40",
    "v_mol_diff":     "E44",
    # Section 6
    "v_alpha_l":      "V4",
    "v_alpha_t":      "X4",
    "v_alpha_v":      "Z4",
    # Section 8
    "v_src_rem_yr":   "D27",
    "v_src_conc_red": "D28",
    # Section 9 — full mapping (was missing the Detailed-only PFAA-3/4
    # columns, the Kf unit dropdown, the converted Kf for PFAA-2..4, the
    # PFAS molecular-weight row, and the # of cells field — which is why
    # §9 inputs weren't reaching the input.inp generator on Run).
    "v_model_psb":    "R22",
    # Freundlich "a"  (row 23)
    "v_psb_a_1":      "V23",
    "v_psb_a_2":      "X23",
    "v_psb_a_3":      "Z23",     # Detailed only
    "v_psb_a_4":      "AB23",    # Detailed only
    # Freundlich Kf  (row 24) + Kf unit dropdown (U24)
    "v_psb_kf_unit":  "U24",
    "v_psb_kf_1":     "V24",
    "v_psb_kf_2":     "X24",
    "v_psb_kf_3":     "Z24",     # Detailed only
    "v_psb_kf_4":     "AB24",    # Detailed only
    # PFAS molecular weight (g/mol)  (row 25 — only used for mol-based Kf units)
    "v_psb_mw_1":     "V25",
    "v_psb_mw_2":     "X25",
    "v_psb_mw_3":     "Z25",     # Detailed only
    "v_psb_mw_4":     "AB25",    # Detailed only
    # Converted Kf (row 26)
    "v_psb_kf_conv":  "V26",
    "v_psb_kf_conv2": "X26",
    "v_psb_kf_conv3": "Z26",     # Detailed only
    "v_psb_kf_conv4": "AB26",    # Detailed only
    # PSB geometry / install
    "v_psb_yr":       "AB28",
    "v_psb_dist":     "X74",
    "v_psb_width":    "Y82",
    "v_psb_load":     "AA82",
    "v_psb_cells":    "AC82",
    # Section 10 / 11
    "v_sample_yr":    "Y74",
    "v_see_every":    "V47",
}

# Section 7 source-term decade rows (1977..2077)
for _i in range(11):
    CELL_MAP[f"v_src_years_{_i}"]  = f"U{8 + _i}"
    CELL_MAP[f"v_src_pfaa1_{_i}"]  = f"V{8 + _i}"
    CELL_MAP[f"v_src_pfaa2_{_i}"]  = f"X{8 + _i}"
del _i

# Section 10 monitoring well rows (7 entries)
for _j in range(7):
    CELL_MAP[f"v_mw_names_{_j}"] = f"U{34 + _j}"
    CELL_MAP[f"v_mw_conc_{_j}"]  = f"V{34 + _j}"
    CELL_MAP[f"v_mw_conc2_{_j}"] = f"X{34 + _j}"
    CELL_MAP[f"v_mw_dist_{_j}"]  = f"AF{34 + _j}"
del _j


def _coerce(value: str) -> Any:
    """Try int → float → fall back to original string."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() == "none":
        return None
    if s.lower() == "true":
        return True
    if s.lower() == "false":
        return False
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def _read_var(app, name: str):
    """Read a StringVar/list-StringVar attribute by dotted name."""
    import tkinter as tk
    if "_" in name and name.rsplit("_", 1)[1].isdigit():
        base, idx = name.rsplit("_", 1)
        lst = getattr(app, base, None)
        if isinstance(lst, list):
            try:
                v = lst[int(idx)]
            except (IndexError, ValueError):
                return None
            if isinstance(v, tk.Variable):
                return v.get()
            return v
        return None
    var = getattr(app, name, None)
    if isinstance(var, tk.Variable):
        return var.get()
    return None


def push_state_to_xlsm(app, xlsm_path: str, sheet_name: str) -> bool:
    """Write current app state into the xlsm workbook.

    Returns True on success, False if openpyxl missing / file locked /
    sheet absent.  Errors are swallowed so that a missing dependency
    doesn't break the popup launch — the EXE will just operate on the
    last-saved values.
    """
    if not os.path.exists(xlsm_path):
        return False
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    try:
        wb = load_workbook(xlsm_path, keep_vba=True)
    except Exception as e:
        print(f"[xlsm_io] open failed: {e}")
        return False

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Push the cell-mapped StringVars
    for name, addr in CELL_MAP.items():
        raw = _read_var(app, name)
        if raw is None:
            continue
        try:
            ws[addr] = _coerce(raw)
        except Exception as e:
            print(f"[xlsm_io] write {addr} failed: {e}")

    # Push version flag (A8: 1=Simple, 2=Detailed)
    try:
        ver = getattr(app, "v_model_version", None)
        if ver is not None:
            v = ver.get() if hasattr(ver, "get") else str(ver)
            ws["A8"] = 1 if "Simple" in v else 2
    except Exception:
        pass

    # Push units flag (AD1: 1=feet, 2=meters)
    try:
        u = getattr(app, "v_units", None)
        if u is not None:
            uv = u.get() if hasattr(u, "get") else str(u)
            ws["AD1"] = 1 if uv == "feet" else 2
    except Exception:
        pass

    # Push heterogeneity flag (A1: 1=High, 2=Medium, 3=Weak)
    try:
        h = getattr(app, "v_het", None)
        if h is not None:
            hv = h.get() if hasattr(h, "get") else str(h)
            ws["A1"] = {"High": 1, "Medium": 2, "Weak": 3}.get(hv, 2)
    except Exception:
        pass

    try:
        wb.save(xlsm_path)
        return True
    except Exception as e:
        print(f"[xlsm_io] save failed: {e}")
        return False
