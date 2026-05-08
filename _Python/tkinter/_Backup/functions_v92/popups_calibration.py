"""
popups_calibration.py — full-fidelity pure-Python port.

Adapted from Source_Py/popups_calibration.py.  Lets the user pick a
calibration template Excel file, previews the chosen sheet ("model
location" or "model data") in a Treeview, and persists the chosen path
to calibration_inputs.txt — the format export_calibration_data /
cali_1 already consume.

Called from main.run_script() when CalibrationDataLoader is dispatched.
"""
from __future__ import annotations
import os
import platform
import subprocess
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from .state import get_state


FONT_TITLE  = ("Arial", 16, "bold")
FONT_LABEL  = ("Arial", 11)
FONT_BOLD   = ("Arial", 11, "bold")
FONT_BTN    = ("Arial", 11)
FONT_SMALL  = ("Arial", 9)


def _docs_root():
    here = os.path.dirname(os.path.abspath(__file__))
    project = os.path.abspath(os.path.join(here, "..", "..", ".."))
    return os.path.join(project, "docs", "_site")


def _open_help():
    f = os.path.join(_docs_root(), "appendix", "appendix_10_1.html")
    if not os.path.exists(f):
        f = os.path.join(_docs_root(), "data_chicklets",
                         "Step10_FieldDataToCalibrate.html")
    if not os.path.exists(f):
        messagebox.showerror("Help Not Found", f"Help file not found:\n{f}")
        return
    abs_p = os.path.abspath(f).replace("\\", "/")
    url = (f"file:///{abs_p}" if os.name == "nt" and abs_p[1] == ":"
           else f"file://{abs_p}")
    try:
        if platform.system() == "Windows":
            for exe in (r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"):
                if os.path.exists(exe):
                    subprocess.Popen([exe, url]); return
        webbrowser.open(url)
    except Exception:
        webbrowser.open(url)


def _load_existing(path):
    if not os.path.exists(path):
        return ""
    try:
        with open(path) as f:
            for ln in f:
                if ln.startswith("Excel File Path:"):
                    return ln.split(":", 1)[1].strip()
    except Exception:
        pass
    return ""


def _save_path(p, dest):
    """v92: preserve existing calibration_inputs.txt content when adding
    or updating the Excel File Path line.  The §calibration "1. Save
    Calibration Data" button writes ranges/weights to this same file —
    we used to clobber it; now we merge."""
    try:
        existing_lines = []
        if os.path.exists(dest):
            try:
                with open(dest, "r", encoding="utf-8") as f:
                    existing_lines = f.readlines()
            except Exception:
                existing_lines = []
        # Strip any prior Excel File Path line(s)
        kept = [ln for ln in existing_lines
                if not ln.lstrip().startswith("Excel File Path:")]
        # If there's no existing content at all, use the legacy header
        if not kept:
            kept = []
        try:
            os.chmod(dest, 0o666)
        except Exception:
            pass
        with open(dest, "w", encoding="utf-8") as f:
            f.write(f"Excel File Path: {p}\n")
            if kept:
                # Ensure a separating newline if needed
                if kept[0].strip() != "":
                    f.write("\n")
                f.writelines(kept)
        return True
    except Exception:
        return False


def _import_xlsx_into_app(app, xlsx_path):
    """v92: parse the user's calibration .xlsx and push monitoring well
    metadata into §10 StringVars.

    - Sheet "Model Location" → fills v_mw_names + v_mw_dist by matching
      the column whose header contains "Location Name" / "Distance"
      (case-insensitive).
    - Sheet "Model Data"     → for each (well, analyte), keeps the
      latest concentration and writes it into v_mw_conc (PFAA-1) or
      v_mw_conc2 (PFAA-2), matching by analyte name against
      app.v_pfaa1 / app.v_pfaa2.

    Best-effort: any parse error is logged and skipped — the path is
    still saved.  Returns a short status string for the messagebox.
    """
    if not _HAS_OPENPYXL:
        return "openpyxl not installed; only path was saved"
    if not os.path.isfile(xlsx_path):
        return "file not found; only path was saved"
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        return f"could not open .xlsx ({exc}); only path was saved"

    n_names = n_concs = 0
    try:
        # ── Model Location → names + distances ────────────────────
        loc_sheet = None
        for nm in wb.sheetnames:
            if nm.strip().lower() == "model location":
                loc_sheet = wb[nm]
                break
        if loc_sheet is not None:
            # find header row (first row where any cell contains
            # "location name" or "distance")
            header_row = 1
            name_col = dist_col = None
            for r in range(1, 20):
                for c in range(1, 27):
                    val = loc_sheet.cell(row=r, column=c).value
                    if val is None:
                        continue
                    s = str(val).strip().lower()
                    if "location" in s and "name" in s:
                        name_col = c; header_row = r
                    elif "distance" in s:
                        dist_col = c
                if name_col is not None or dist_col is not None:
                    break
            mw_names = getattr(app, "v_mw_names", []) or []
            mw_dists = getattr(app, "v_mw_dist",  []) or []
            seen_wells = []
            for r in range(header_row + 1, header_row + 50):
                nm = (loc_sheet.cell(row=r, column=name_col).value
                      if name_col else None)
                ds = (loc_sheet.cell(row=r, column=dist_col).value
                      if dist_col else None)
                if nm is None or str(nm).strip() == "":
                    if seen_wells:
                        break  # blank row after data → stop
                    continue
                seen_wells.append(str(nm).strip())
                idx = len(seen_wells) - 1
                if idx < len(mw_names):
                    try: mw_names[idx].set(str(nm).strip())
                    except Exception: pass
                    n_names += 1
                if ds is not None and idx < len(mw_dists):
                    try: mw_dists[idx].set(str(ds).strip())
                    except Exception: pass
                if idx >= 6:
                    break

        # ── Model Data → latest concentration per (well, analyte) ─
        data_sheet = None
        for nm in wb.sheetnames:
            if nm.strip().lower() == "model data":
                data_sheet = wb[nm]
                break
        if data_sheet is not None:
            # latest[(well_lower, analyte_lower)] = (date_sortkey, conc)
            latest = {}
            header_row = 1
            for r in range(1, 20):
                a = data_sheet.cell(row=r, column=1).value
                c_ = data_sheet.cell(row=r, column=3).value
                if a is None or c_ is None:
                    continue
                if ("date" in str(a).lower() or "time" in str(a).lower()) \
                        and ("analyte" in str(c_).lower() or "compound" in str(c_).lower()):
                    header_row = r
                    break
            max_row = data_sheet.max_row or 1000
            for r in range(header_row + 1, min(header_row + 5000, max_row + 1)):
                d = data_sheet.cell(row=r, column=1).value
                w = data_sheet.cell(row=r, column=2).value
                a = data_sheet.cell(row=r, column=3).value
                c_ = data_sheet.cell(row=r, column=4).value
                if w is None or a is None or c_ is None:
                    continue
                try:
                    cval = float(c_)
                except (TypeError, ValueError):
                    continue
                wkey = str(w).strip().lower()
                akey = str(a).strip().lower()
                # Sort key: timestamp if datetime, else string
                try:
                    sk = d.timestamp() if hasattr(d, "timestamp") else float(d)
                except Exception:
                    sk = 0.0
                cur = latest.get((wkey, akey))
                if cur is None or sk > cur[0]:
                    latest[(wkey, akey)] = (sk, cval)

            # Resolve PFAS-1 / PFAS-2 names from app
            pfaa1_name = ""
            pfaa2_name = ""
            try:
                pfaa1_name = str(app.v_pfaa1.get()).strip().lower()
            except Exception:
                pass
            try:
                pfaa2_name = str(app.v_pfaa2.get()).strip().lower()
            except Exception:
                pass

            # Map each well in v_mw_names → latest concs
            mw_names = getattr(app, "v_mw_names", []) or []
            mw_conc  = getattr(app, "v_mw_conc",  []) or []
            mw_conc2 = getattr(app, "v_mw_conc2", []) or []
            for i, nv in enumerate(mw_names):
                try:
                    well_str = str(nv.get()).strip().lower()
                except Exception:
                    continue
                if not well_str:
                    continue
                if pfaa1_name and (well_str, pfaa1_name) in latest:
                    if i < len(mw_conc):
                        try: mw_conc[i].set(f"{latest[(well_str, pfaa1_name)][1]:g}")
                        except Exception: pass
                        n_concs += 1
                if pfaa2_name and pfaa2_name != "none" \
                        and (well_str, pfaa2_name) in latest:
                    if i < len(mw_conc2):
                        try: mw_conc2[i].set(f"{latest[(well_str, pfaa2_name)][1]:g}")
                        except Exception: pass
                        n_concs += 1
    finally:
        try: wb.close()
        except Exception: pass

    return (f"Imported {n_names} well names + "
            f"{n_concs} concentration value(s) into §10")


def run(app, parent=None):
    state = get_state()
    work_dir = state.work_dir or os.getcwd()
    txt_path = os.path.join(work_dir, "calibration_inputs.txt")

    here = os.path.dirname(os.path.abspath(__file__))
    project = os.path.abspath(os.path.join(here, "..", "..", ".."))
    is_detailed = (getattr(app, "active_sheet", "Simple") == "Detailed_2")
    template = ("CalibrationTemplate_Detailed.xlsx" if is_detailed
                else "CalibrationTemplate_Simple.xlsx")
    default_path = os.path.join(project, template)
    existing = _load_existing(txt_path)
    initial = existing or (default_path if os.path.exists(default_path) else "")

    root = tk.Toplevel(parent or app)
    root.title("Calibration Data Upload")
    root.configure(bg="#F0F0F0")
    try: root.withdraw()
    except Exception: pass
    try:
        root.grab_set()
    except Exception:
        pass

    # Treeview rowheight fix — Windows default themes (vista/xpnative)
    # ignore rowheight, so force "clam" which respects all style options.
    try:
        tv_style = ttk.Style(root)
        try:
            tv_style.theme_use("clam")
        except Exception:
            pass
        tv_style.configure("Treeview",
                           rowheight=36, font=FONT_LABEL,
                           background="#FFFFFF",
                           fieldbackground="#FFFFFF")
        tv_style.configure("Treeview.Heading",
                           font=FONT_BOLD, padding=(4, 6))
    except Exception:
        pass

    title = tk.Frame(root, bg="#F0F0F0"); title.pack(pady=10)
    tk.Label(title, text="Calibration Data Upload",
             font=FONT_TITLE, bg="#F0F0F0").pack()

    main = tk.Frame(root, bg="#F0F0F0", padx=20, pady=10)
    main.pack(expand=True, fill="both")

    tk.Label(main, text="Select an Excel file containing calibration data:",
             font=FONT_LABEL, bg="#F0F0F0").pack(pady=10, anchor="w")

    file_frame = tk.Frame(main, bg="#F0F0F0"); file_frame.pack(fill="x", pady=10)
    tk.Label(file_frame, text="Excel File:", font=FONT_BOLD, bg="#F0F0F0",
             width=12, anchor="w").pack(side="left")
    path_var = tk.StringVar(value=initial)
    file_entry = tk.Entry(file_frame, textvariable=path_var, font=FONT_LABEL,
                           width=50)
    file_entry.pack(side="left", padx=5, fill="x", expand=True)

    sheet_frame = tk.Frame(main, bg="#F0F0F0"); sheet_frame.pack(fill="x", pady=10)
    tk.Label(sheet_frame, text="Preview Sheet:", font=FONT_BOLD,
             bg="#F0F0F0", width=12, anchor="w").pack(side="left")
    sheet_var = tk.StringVar(value="model location")

    preview_frame = tk.Frame(main, bg="#F0F0F0")
    preview_frame.pack(fill="both", expand=True, pady=10)
    info_label = tk.Label(preview_frame, text="File Preview:",
                           font=FONT_BOLD, bg="#F0F0F0")
    info_label.pack(anchor="w", pad