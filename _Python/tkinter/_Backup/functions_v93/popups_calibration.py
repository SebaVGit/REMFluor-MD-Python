"""
popups_calibration.py — full-fidelity pure-Python port.

Adapted from Source_Py/popups_calibration.py.  Lets the user pick a
calibration template Excel file, previews the chosen sheet ("model
location" or "model data") in a Treeview, and persists the chosen path
to calibration_inputs.txt — the format export_calibration_data /
cali_1 already consume.

Called from main.run_script() when CalibrationDataLoader is dispatched.
"""
from __future__ import annotations
import os
import platform
import subprocess
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from .state import get_state


FONT_TITLE  = ("Arial", 16, "bold")
FONT_LABEL  = ("Arial", 11)
FONT_BOLD   = ("Arial", 11, "bold")
FONT_BTN    = ("Arial", 11)
FONT_SMALL  = ("Arial", 9)


def _docs_root():
    here = os.path.dirname(os.path.abspath(__file__))
    project = os.path.abspath(os.path.join(here, "..", "..", ".."))
    return os.path.join(project, "docs", "_site")


def _open_help():
    f = os.path.join(_docs_root(), "appendix", "appendix_10_1.html")
    if not os.path.exists(f):
        f = os.path.join(_docs_root(), "data_chicklets",
                         "Step10_FieldDataToCalibrate.html")
    if not os.path.exists(f):
        messagebox.showerror("Help Not Found", f"Help file not found:\n{f}")
        return
    abs_p = os.path.abspath(f).replace("\\", "/")
    url = (f"file:///{abs_p}" if os.name == "nt" and abs_p[1] == ":"
           else f"file://{abs_p}")
    try:
        if platform.system() == "Windows":
            for exe in (r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"):
                if os.path.exists(exe):
                    subprocess.Popen([exe, url]); return
        webbrowser.open(url)
    except Exception:
        webbrowser.open(url)


def _load_existing(path):
    if not os.path.exists(path):
        return ""
    try:
        with open(path) as f:
            for ln in f:
                if ln.startswith("Excel File Path:"):
                    return ln.split(":", 1)[1].strip()
    except Exception:
        pass
    return ""


def _save_path(p, dest):
    """v92: preserve existing calibration_inputs.txt content when adding
    or updating the Excel File Path line.  The §calibration "1. Save
    Calibration Data" button writes ranges/weights to this same file —
    we used to clobber it; now we merge."""
    try:
        existing_lines = []
        if os.path.exists(dest):
            try:
                with open(dest, "r", encoding="utf-8") as f:
                    existing_lines = f.readlines()
            except Exception:
                existing_lines = []
        # Strip any prior Excel File Path line(s)
        kept = [ln for ln in existing_lines
                if not ln.lstrip().startswith("Excel File Path:")]
        # If there's no existing content at all, use the legacy header
        if not kept:
            kept = []
        try:
            os.chmod(dest, 0o666)
        except Exception:
            pass
        with open(dest, "w", encoding="utf-8") as f:
            f.write(f"Excel File Path: {p}\n")
            if kept:
                # Ensure a separating newline if needed
                if kept[0].strip() != "":
                    f.write("\n")
                f.writelines(kept)
        return True
    except Exception:
        return False


def _import_xlsx_into_app(app, xlsx_path):
    """v93: parse the user's calibration .xlsx and push monitoring well
    metadata into §10 StringVars.

    Permissive parser — tries multiple strategies:
      (1) Sheet name match: "model location" / "modellocation" /
          "location" / "locations" / "wells" (case-insensitive).
      (2) Header detection: look for rows containing "name"/"location"
          and "distance"/"x" — but if none found, fall back to:
          column A = names, column B (or any numeric column) = dist.
      (3) If no Location sheet exists at all, derive unique well
          names from the Data sheet's well column.

    Writes a debug log to <project>/calibration_import.log so the
    user can see exactly what happened.

    Returns a short status string for the messagebox.
    """
    proj_dir = os.path.dirname(os.path.abspath(xlsx_path))
    log_path = os.path.join(proj_dir, "calibration_import.log")
    log_lines = [f"Source .xlsx: {xlsx_path}"]

    def _log(msg):
        log_lines.append(str(msg))

    if not _HAS_OPENPYXL:
        _log("openpyxl not installed; nothing imported")
        _flush_log(log_path, log_lines)
        return "openpyxl not installed; only path was saved"
    if not os.path.isfile(xlsx_path):
        _log(f"file not found: {xlsx_path}")
        _flush_log(log_path, log_lines)
        return "file not found; only path was saved"
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        _log(f"could not open .xlsx: {exc}")
        _flush_log(log_path, log_lines)
        return f"could not open .xlsx ({exc}); only path was saved"

    _log(f"Sheets in workbook: {wb.sheetnames}")

    n_names = n_concs = 0
    try:
        # ── Pick a "location" sheet ──────────────────────────────
        loc_sheet = None
        loc_keywords = ("model location", "modellocation",
                        "location", "locations", "wells", "well")
        for nm in wb.sheetnames:
            low = nm.strip().lower()
            if low in loc_keywords or "location" in low:
                loc_sheet = wb[nm]
                _log(f"Picked location sheet: {nm!r}")
                break
        if loc_sheet is None:
            _log("No 'Model Location' / 'Locations' / 'Wells' sheet found")

        if loc_sheet is not None:
            # Try header detection
            header_row = None
            name_col = dist_col = None
            for r in range(1, 30):
                for c in range(1, 27):
                    val = loc_sheet.cell(row=r, column=c).value
                    if val is None:
                        continue
                    s = str(val).strip().lower()
                    if name_col is None and (
                        ("location" in s and "name" in s)
                        or s in ("name", "well", "well name", "well id",
                                 "id", "site", "site name")
                    ):
                        name_col = c; header_row = r
                    elif dist_col is None and (
                        "distance" in s or s == "x"
                        or "from source" in s or s == "x (m)"
                        or s == "x (ft)" or s.endswith(" (m)") and "x" in s
                    ):
                        dist_col = c
                if header_row is not None:
                    break

            # Fallback: no header → assume column A=names, B=distances
            if header_row is None:
                header_row = 0  # data starts at row 1
                name_col = 1
                dist_col = 2
                _log("No header found; defaulting to col A=names, col B=distances")
            else:
                _log(f"Header row {header_row}: name_col={name_col}, "
                     f"dist_col={dist_col}")

            if name_col is None:
                # Final fallback: column A is names
                name_col = 1
                _log("No name column matched; defaulting to col A")

            mw_names = getattr(app, "v_mw_names", []) or []
            mw_dists = getattr(app, "v_mw_dist",  []) or []
            seen_wells = []
            blank_streak = 0
            for r in range(header_row + 1, header_row + 100):
                nm_v = loc_sheet.cell(row=r, column=name_col).value
                ds_v = (loc_sheet.cell(row=r, column=dist_col).value
                        if dist_col else None)
                # Reject rows that look like header repeats / units
                if nm_v is not None and isinstance(nm_v, str):
                    s = nm_v.strip().lower()
                    if s in ("name", "well", "location", "x", "id", ""):
                        continue
                if nm_v is None or str(nm_v).strip() == "":
                    blank_streak += 1
                    if blank_streak >= 2 and seen_wells:
                        break
                    continue
                blank_streak = 0
                well_name = str(nm_v).strip()
                seen_wells.append(well_name)
                idx = len(seen_wells) - 1
                if idx < len(mw_names):
                    try:
                        mw_names[idx].set(well_name)
                        n_names += 1
                    except Exception as exc:
                        _log(f"failed to set v_mw_names[{idx}]: {exc}")
                if ds_v is not None and idx < len(mw_dists):
                    try:
                        # numeric coerce if possible
                        try: ds_clean = f"{float(ds_v):g}"
                        except (TypeError, ValueError): ds_clean = str(ds_v).strip()
                        mw_dists[idx].set(ds_clean)
                    except Exception as exc:
                        _log(f"failed to set v_mw_dist[{idx}]: {exc}")
                if idx >= 6:
                    break
            _log(f"Wells imported from Location sheet: "
                 f"{n_names} of {len(seen_wells)} found "
                 f"(slots: {len(mw_names)})")

        # ── Pick a "data" sheet ───────────────────────────────────
        data_sheet = None
        data_keywords = ("model data", "modeldata", "data",
                         "concentrations", "measurements", "field data",
                         "fielddata")
        for nm in wb.sheetnames:
            low = nm.strip().lower()
            if low in data_keywords or "data" in low:
                # Don't re-use the Location sheet
                if loc_sheet is not None and wb[nm] is loc_sheet:
                    continue
                data_sheet = wb[nm]
                _log(f"Picked data sheet: {nm!r}")
                break
        if data_sheet is None:
            _log("No 'Model Data' / 'Data' sheet found; "
                 "skipping concentration import")

        # If still no Location sheet but we have Data, derive names
        if loc_sheet is None and data_sheet is not None:
            _log("Falling back to deriving well names from Data sheet")
            wells_in_data = []
            for r in range(2, min(2000, (data_sheet.max_row or 0) + 1)):
                w = data_sheet.cell(row=r, column=2).value
                if w is None:
                    continue
                ws = str(w).strip()
                if ws and ws.lower() not in (
                    "well", "location", "name", "site"
                ) and ws not in wells_in_data:
                    wells_in_data.append(ws)
                if len(wells_in_data) >= 7:
                    break
            mw_names = getattr(app, "v_mw_names", []) or []
            for i, name in enumerate(wells_in_data[:len(mw_names)]):
                try:
                    mw_names[i].set(name)
                    n_names += 1
                except Exception:
                    pass
            _log(f"Derived {n_names} well names from Data sheet")

        if data_sheet is not None:
            latest = {}
            # Try to find header row by scanning column A/C
            header_row = 1
            for r in range(1, 30):
                a = data_sheet.cell(row=r, column=1).value
                c_ = data_sheet.cell(row=r, column=3).value
                if a is None or c_ is None:
                    continue
                a_s = str(a).strip().lower()
                c_s = str(c_).strip().lower()
                if ("date" in a_s or "time" in a_s) \
                        and ("analyte" in c_s or "compound" in c_s):
                    header_row = r
                    break
            _log(f"Data sheet header row: {header_row}")
            max_row = data_sheet.max_row or 1000
            for r in range(header_row + 1, min(header_row + 10000, max_row + 1)):
                d = data_sheet.cell(row=r, column=1).value
                w = data_sheet.cell(row=r, column=2).value
                a = data_sheet.cell(row=r, column=3).value
                c_ = data_sheet.cell(row=r, column=4).value
                if w is None or a is None or c_ is None:
                    continue
                try:
                    cval = float(c_)
                except (TypeError, ValueError):
                    continue
                wkey = str(w).strip().lower()
                akey = str(a).strip().lower()
                try:
                    sk = d.timestamp() if hasattr(d, "timestamp") else float(d)
                except Exception:
                    sk = 0.0
                cur = latest.get((wkey, akey))
                if cur is None or sk > cur[0]:
             