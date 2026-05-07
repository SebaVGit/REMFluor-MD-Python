"""
popups_calibration.py — full-fidelity pure-Python port.

Adapted from Source_Py/popups_calibration.py.  Lets the user pick a
calibration template Excel file, previews the chosen sheet ("model
location" or "model data") in a Treeview, and persists the chosen path
to calibration_inputs.txt — the format export_calibration_data /
cali_1 already consume.

Called from main.run_script() when CalibrationDataLoader is dispatched.
"""
from __future__ import annotations
import os
import platform
import subprocess
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from .state import get_state


FONT_TITLE  = ("Arial", 16, "bold")
FONT_LABEL  = ("Arial", 11)
FONT_BOLD   = ("Arial", 11, "bold")
FONT_BTN    = ("Arial", 11)
FONT_SMALL  = ("Arial", 9)


def _docs_root():
    here = os.path.dirname(os.path.abspath(__file__))
    project = os.path.abspath(os.path.join(here, "..", "..", ".."))
    return os.path.join(project, "docs", "_site")


def _open_help():
    f = os.path.join(_docs_root(), "appendix", "appendix_10_1.html")
    if not os.path.exists(f):
        f = os.path.join(_docs_root(), "data_chicklets",
                         "Step10_FieldDataToCalibrate.html")
    if not os.path.exists(f):
        messagebox.showerror("Help Not Found", f"Help file not found:\n{f}")
        return
    abs_p = os.path.abspath(f).replace("\\", "/")
    url = (f"file:///{abs_p}" if os.name == "nt" and abs_p[1] == ":"
           else f"file://{abs_p}")
    try:
        if platform.system() == "Windows":
            for exe in (r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"):
                if os.path.exists(exe):
                    subprocess.Popen([exe, url]); return
        webbrowser.open(url)
    except Exception:
        webbrowser.open(url)


def _load_existing(path):
    if not os.path.exists(path):
        return ""
    try:
        with open(path) as f:
            for ln in f:
                if ln.startswith("Excel File Path:"):
                    return ln.split(":", 1)[1].strip()
    except Exception:
        pass
    return ""


def _save_path(p, dest):
    try:
        if os.path.exists(dest):
            try:
                os.chmod(dest, 0o666); os.remove(dest)
            except Exception: pass
        with open(dest, "w") as f:
            f.write(f"Excel File Path: {p}\n")
        return True
    except Exception:
        return False


def run(app, parent=None):
    state = get_state()
    work_dir = state.work_dir or os.getcwd()
    txt_path = os.path.join(work_dir, "calibration_inputs.txt")

    here = os.path.dirname(os.path.abspath(__file__))
    project = os.path.abspath(os.path.join(here, "..", "..", ".."))
    is_detailed = (getattr(app, "active_sheet", "Simple") == "Detailed_2")
    template = ("CalibrationTemplate_Detailed.xlsx" if is_detailed
                else "CalibrationTemplate_Simple.xlsx")
    default_path = os.path.join(project, template)
    existing = _load_existing(txt_path)
    initial = existing or (default_path if os.path.exists(default_path) else "")

    root = tk.Toplevel(parent or app)
    root.title("Calibration Data Upload")
    root.configure(bg="#F0F0F0")
    try: root.withdraw()
    except Exception: pass
    try:
        root.grab_set()
    except Exception:
        pass

    # Treeview rowheight fix — Windows default themes (vista/xpnative)
    # ignore rowheight, so force "clam" which respects all style options.
    try:
        tv_style = ttk.Style(root)
        try:
            tv_style.theme_use("clam")
        except Exception:
            pass
        tv_style.configure("Treeview",
                           rowheight=36, font=FONT_LABEL,
                           background="#FFFFFF",
                           fieldbackground="#FFFFFF")
        tv_style.configure("Treeview.Heading",
                           font=FONT_BOLD, padding=(4, 6))
    except Exception:
        pass

    title = tk.Frame(root, bg="#F0F0F0"); title.pack(pady=10)
    tk.Label(title, text="Calibration Data Upload",
             font=FONT_TITLE, bg="#F0F0F0").pack()

    main = tk.Frame(root, bg="#F0F0F0", padx=20, pady=10)
    main.pack(expand=True, fill="both")

    tk.Label(main, text="Select an Excel file containing calibration data:",
             font=FONT_LABEL, bg="#F0F0F0").pack(pady=10, anchor="w")

    file_frame = tk.Frame(main, bg="#F0F0F0"); file_frame.pack(fill="x", pady=10)
    tk.Label(file_frame, text="Excel File:", font=FONT_BOLD, bg="#F0F0F0",
             width=12, anchor="w").pack(side="left")
    path_var = tk.StringVar(value=initial)
    file_entry = tk.Entry(file_frame, textvariable=path_var, font=FONT_LABEL,
                           width=50)
    file_entry.pack(side="left", padx=5, fill="x", expand=True)

    sheet_frame = tk.Frame(main, bg="#F0F0F0"); sheet_frame.pack(fill="x", pady=10)
    tk.Label(sheet_frame, text="Preview Sheet:", font=FONT_BOLD,
             bg="#F0F0F0", width=12, anchor="w").pack(side="left")
    sheet_var = tk.StringVar(value="model location")

    preview_frame = tk.Frame(main, bg="#F0F0F0")
    preview_frame.pack(fill="both", expand=True, pady=10)
    info_label = tk.Label(preview_frame, text="File Preview:",
                           font=FONT_BOLD, bg="#F0F0F0")
    info_label.pack(anchor="w", pady=(0, 5))

    table_frame = tk.Frame(preview_frame, bg="#F0F0F0")
    table_frame.pack(fill="both", expand=True)
    h_sb = ttk.Scrollbar(table_frame, orient="horizontal")
    h_sb.pack(side="bottom", fill="x")
    v_sb = ttk.Scrollbar(table_frame, orient="vertical")
    v_sb.pack(side="right", fill="y")
    preview_tree = ttk.Treeview(table_frame, xscrollcommand=h_sb.set,
                                 yscrollcommand=v_sb.set, show="headings")
    preview_tree.pack(side="left", fill="both", expand=True)
    h_sb.config(command=preview_tree.xview)
    v_sb.config(command=preview_tree.yview)

    def _update_preview():
        for it in preview_tree.get_children():
            preview_tree.delete(it)
        p = path_var.get().strip()
        if not p or not os.path.exists(p):
            info_label.config(text="File Preview: No file selected.")
            return
        if not _HAS_OPENPYXL:
            info_label.config(
                text="File Preview: openpyxl not installed (preview disabled).")
            return
        try:
            wb = load_workbook(p, data_only=True)
            target = sheet_var.get()
            sheet_name = None
            for s in wb.sheetnames:
                if s.lower() == target.lower():
                    sheet_name = s; break
            if sheet_name is None:
                sheet_name = wb.sheetnames[0] if wb.sheetnames else None
            if sheet_name is None:
                wb.close()
                info_label.config(text="File Preview: No sheets found.")
                return
            ws = wb[sheet_name]
            info_label.config(
                text=(f"File: {os.path.basename(p)} | Sheet: {ws.title} | "
                      f"Dimensions: {ws.max_row} rows x {ws.max_column} cols"))

            max_rows = min(30, ws.max_row)
            max_cols = min(15, ws.max_column)
            col_names = []
            for ci in range(1, max_cols + 1):
                v = ws.cell(row=1, column=ci).value
                col_names.append(str(v) if v is not None
                                 else get_column_letter(ci))
            cols = ["#"] + col_names
            preview_tree["columns"] = cols
            preview_tree.column("#", width=50, anchor="center",
                                 stretch=False, minwidth=40)
            preview_tree.heading("#", text="#", anchor="center")
            for cn in col_names:
                preview_tree.column(cn, width=110, anchor="w", stretch=True,
                                     minwidth=60)
                preview_tree.heading(cn, text=cn, anchor="center")

            data_start = 2
            data_n = min(max_rows - 1, ws.max_row - 1)
            for di in range(data_start, data_start + data_n):
                if di > ws.max_row:
                    break
                row_values = [str(di - 1)]
                for ci in range(1, max_cols + 1):
                    val = ws.cell(row=di, column=ci).value
                    sval = "" if val is None else str(val)
                    if len(sval) > 30:
                        sval = sval[:27] + "..."
                    row_values.append(sval)
                tag = "even" if (di - data_start) % 2 == 0 else "odd"
                preview_tree.insert("", "end", values=row_values, tags=(tag,))
            preview_tree.tag_configure("odd", background="#FFFFFF")
            preview_tree.tag_configure("even", background="#F0F0F0")
            if ws.max_row > max_rows or ws.max_column > max_cols:
                preview_tree.insert("", "end", tags=("note",),
                                     values=[f"... showing first {data_n} rows "
                                             f"x {max_cols} cols"]
                                            + [""] * max_cols)
                preview_tree.tag_configure("note", background="#E8E8E8",
                                            foreground="#666666")
            wb.close()
        except Exception as e:
            info_label.config(text=f"File Preview: error - {e}")

    for opt in ("model location", "model data"):
        tk.Radiobutton(sheet_frame, text=opt, variable=sheet_var, value=opt,
                       command=_update_preview,
                       font=FONT_LABEL, bg="#F0F0F0").pack(side="left", padx=10)

    def _browse():
        p = filedialog.askopenfilename(
            title="Select Excel File for Calibration",
            initialdir=(os.path.dirname(path_var.get()) if path_var.get()
                        else project),
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"),
                       ("All files", "*.*")],
            parent=root)
        if p:
            path_var.set(p); _update_preview()

    tk.Button(file_frame, text="Browse...", command=_browse,
              font=FONT_LABEL, width=10).pack(side="left", padx=5)

    if initial:
        _update_preview()

    def _save_and_exit():
        p = path_var.get().strip()
        if not p:
            messagebox.showerror("Error",
                                 "Please select an Excel file.",
                                 parent=root); return
        if not os.path.isfile(p):
            messagebox.showerror("Error", f"File not found:\n{p}",
                                 parent=root); return
        if _HAS_OPENPYXL:
            try:
                wb = load_workbook(p, data_only=True)
                names = [s.lower() for s in wb.sheetnames]
                missing = [s for s in ("model location", "model data")
                           if s not in names]
                wb.close()
                if missing:
                    messagebox.showwarning(
                        "Warning",
                        ("Excel file is missing sheets: "
                         f"{', '.join(missing)}"),
                        parent=root)
            except Exception as e:
                messagebox.showerror("Error", f"Invalid Excel file:\n{e}",
                                     parent=root); return
        if _save_path(p, txt_path):
            messagebox.showinfo("Saved",
                                f"Calibration template path saved.\n"
                                f"{os.path.basename(txt_path)}",
                                parent=root)
            try: root.grab_release()
            except Exception: pass
            root.destroy()
        else:
            messagebox.showerror("Error",
                                 f"Could not write {txt_path}",
                                 parent=root)

    def _cancel():
        try: root.grab_release()
        except Exception: pass
        root.destroy()

    bar = tk.Frame(root, bg="#F0F0F0"); bar.pack(pady=12)
    tk.Button(bar, text="OK",     width=10, font=FONT_BTN,
              command=_save_and_exit).pack(side="left", padx=6)
    tk.Button(bar, text="Cancel", width=10, font=FONT_BTN,
              command=_cancel).pack(side="left", padx=6)
    tk.Button(bar, text="Help",   width=10, font=FONT_BTN,
              command=_open_help).pack(side="left", padx=6)

    root.protocol("WM_DELETE_WINDOW", _cancel)
    root.update_idletasks()
    w = max(root.winfo_reqwidth() + 32, 980)
    h = max(root.winfo_reqheight() + 24, 640)
    try:
        sw = root.winfo_screenwidth(); sh = root.winfo_screenheight()
        w = min(w, int(sw * 0.95)); h = min(h, int(sh * 0.92))
        x = max(0, (sw - w) // 2); y = max(0, (sh - h) // 2 - 30)
        root.geometry(f"{w}x{h}+{x}+{y}")
    except Exception:
        root.geometry(f"{w}x{h}")
    root.minsize(min(w, 800), min(h, 540))
    root.resizable(True, True)
    try:
        root.deiconify(); root.lift(); root.focus_force()
    except Exception:
        pass
    root.wait_window()
