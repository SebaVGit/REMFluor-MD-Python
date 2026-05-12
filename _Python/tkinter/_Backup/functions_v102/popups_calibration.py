"""
popups_calibration.py — full-fidelity pure-Python port.

Adapted from Source_Py/popups_calibration.py.  Lets the user pick a
calibration template Excel file, previews the chosen sheet ("model
location" or "model data") in a Treeview, and persists the chosen path
to calibration_inputs.txt — the format export_calibration_data /
cali_1 already consume.

Called from main.run_script() when CalibrationDataLoader is dispatched.
"""
from __future__ import annotations
import os
import platform
import subprocess
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from .state import get_state


FONT_TITLE  = ("Arial", 16, "bold")
FONT_LABEL  = ("Arial", 11)
FONT_BOLD   = ("Arial", 11, "bold")
FONT_BTN    = ("Arial", 11)
FONT_SMALL  = ("Arial", 9)


def _docs_root():
    """v100: in a frozen --onefile build, walking up from __file__
    lands inside sys._MEIPASS/functions which is wrong.  Use the
    state singleton's bundle_dir / work_dir set by main.py at
    startup; fall back to the dev-tree walk if neither is set."""
    try:
        from .state import get_state
        st = get_state()
        for base in (getattr(st, "bundle_dir", ""),
                     getattr(st, "work_dir", "")):
            if base:
                cand = os.path.join(base, "docs", "_site")
                if os.path.isdir(cand):
                    return cand
    except Exception:
        pass
    here = os.path.dirname(os.path.abspath(__file__))
    project = os.path.abspath(os.path.join(here, "..", "..", ".."))
    return os.path.join(project, "docs", "_site")


def _open_help():
    f = os.path.join(_docs_root(), "appendix", "appendix_10_1.html")
    if not os.path.exists(f):
        f = os.path.join(_docs_root(), "data_chicklets",
                         "Step10_FieldDataToCalibrate.html")
    if not os.path.exists(f):
        messagebox.showerror("Help Not Found", f"Help file not found:\n{f}")
        return
    abs_p = os.path.abspath(f).replace("\\", "/")
    url = (f"file:///{abs_p}" if os.name == "nt" and abs_p[1] == ":"
           else f"file://{abs_p}")
    try:
        if platform.system() == "Windows":
            for exe in (r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"):
                if os.path.exists(exe):
                    subprocess.Popen([exe, url]); return
        webbrowser.open(url)
    except Exception:
        webbrowser.open(url)


def _load_existing(path):
    if not os.path.exists(path):
        return ""
    try:
        with open(path) as fh:
            for ln in fh:
                if ln.startswith("Excel File Path:"):
                    return ln.split(":", 1)[1].strip()
    except Exception:
        pass
    return ""


def _save_path(p, dest):
    """v92: preserve existing calibration_inputs.txt content when adding
    or updating the Excel File Path line."""
    try:
        existing_lines = []
        if os.path.exists(dest):
            try:
                with open(dest, "r", encoding="utf-8") as fh:
                    existing_lines = fh.readlines()
            except Exception:
                existing_lines = []
        kept = [ln for ln in existing_lines
                if not ln.lstrip().startswith("Excel File Path:")]
        try:
            os.chmod(dest, 0o666)
        except Exception:
            pass
        with open(dest, "w", encoding="utf-8") as fh:
            fh.write(f"Excel File Path: {p}\n")
            if kept:
                if kept[0].strip() != "":
                    fh.write("\n")
                fh.writelines(kept)
        return True
    except Exception:
        return False


def _flush_log(path, lines):
    try:
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("REMFluor-MD calibration import log\n")
            fh.write("=" * 60 + "\n")
            for ln in lines:
                fh.write(str(ln) + "\n")
    except Exception:
        pass


def _import_xlsx_into_app(app, xlsx_path):
    """Parse the user's calibration .xlsx and push monitoring well
    metadata into §10 StringVars.  Also writes mw_observations.json
    sidecar so generate_dashboard can render observed concentration
    markers without re-reading the .xlsx at dashboard time."""
    proj_dir = os.path.dirname(os.path.abspath(xlsx_path))
    log_path = os.path.join(proj_dir, "calibration_import.log")
    log_lines = [f"Source .xlsx: {xlsx_path}"]

    def _log(msg):
        log_lines.append(str(msg))

    if not _HAS_OPENPYXL:
        _log("openpyxl not installed; nothing imported")
        _flush_log(log_path, log_lines)
        return "openpyxl not installed; only path was saved"
    if not os.path.isfile(xlsx_path):
        _log(f"file not found: {xlsx_path}")
        _flush_log(log_path, log_lines)
        return "file not found; only path was saved"
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        _log(f"could not open .xlsx: {exc}")
        _flush_log(log_path, log_lines)
        return f"could not open .xlsx ({exc}); only path was saved"

    _log(f"Sheets in workbook: {wb.sheetnames}")

    n_names = n_concs = 0
    n_total_obs = 0
    try:
        loc_sheet = None
        loc_keywords = ("model location", "modellocation",
                        "location", "locations", "wells", "well")
        for nm in wb.sheetnames:
            low = nm.strip().lower()
            if low in loc_keywords or "location" in low:
                loc_sheet = wb[nm]
                _log(f"Picked location sheet: {nm!r}")
                break
        if loc_sheet is None:
            _log("No 'Model Location' / 'Locations' / 'Wells' sheet found")

        BAD_NAMES = {
            "well", "location", "name", "site", "id", "x", "distance",
            "monitoring well", "monitoring well name", "monitoring",
            "well name", "well id", "site name", "location name",
            "wells", "locations",
        }

        # v102: per-well screen depths from .xlsx Model Location
        # (top_screen / bot_screen columns).  Empty when not in the
        # sheet — generate_input_file falls back to constant Z, Z/2.
        well_depths = {}
        if loc_sheet is not None:
            header_row = None
            name_col = dist_col = top_col = bot_col = None
            for r in range(1, 30):
                for c in range(1, 27):
                    val = loc_sheet.cell(row=r, column=c).value
                    if val is None:
                        continue
                    s = str(val).strip().lower()
                    if name_col is None and ("name" in s or s in (
                            "well", "id", "site", "location")):
                        name_col = c; header_row = r
                    elif dist_col is None and ("distance" in s or s == "x"
                                               or "from source" in s):
                        dist_col = c
                    elif top_col is None and ("top" in s and "screen" in s):
                        top_col = c
                    elif bot_col is None and (
                            ("bottom" in s or "bot" in s) and "screen" in s):
                        bot_col = c
                if header_row is not None:
                    break

            if header_row is None:
                header_row = 0; name_col = 1; dist_col = 2
                _log("No header found; defaulting to col A=names, col B=distances")
            else:
                _log(f"Location sheet header row={header_row}: "
                     f"name_col={name_col}, dist_col={dist_col}, "
                     f"top_col={top_col}, bot_col={bot_col}")
            if name_col is None:
                name_col = 1
            # v102: if no header was found for top/bot, fall back to
            # columns 4 and 5 (matches the legacy Source_Py reader).
            if top_col is None:
                top_col = 4
            if bot_col is None:
                bot_col = 5

            mw_names = getattr(app, "v_mw_names", []) or []
            mw_dists = getattr(app, "v_mw_dist", []) or []
            seen_wells = []; blank_streak = 0
            for r in range(header_row + 1, header_row + 100):
                nm_v = loc_sheet.cell(row=r, column=name_col).value
                ds_v = (loc_sheet.cell(row=r, column=dist_col).value
                        if dist_col else None)
                if nm_v is not None and isinstance(nm_v, str):
                    s = nm_v.strip().lower()
                    if s in BAD_NAMES or s == "":
                        continue
                if nm_v is None or str(nm_v).strip() == "":
                    blank_streak += 1
                    if blank_streak >= 2 and seen_wells:
                        break
                    continue
                blank_streak = 0
                well_name = str(nm_v).strip()
                seen_wells.append(well_name)
                idx = len(seen_wells) - 1
                if idx < len(mw_names):
                    try:
                        mw_names[idx].set(well_name); n_names += 1
                    except Exception as exc:
                        _log(f"failed to set v_mw_names[{idx}]: {exc}")
                if ds_v is not None and idx < len(mw_dists):
                    try:
                        try: ds_clean = f"{float(ds_v):g}"
                        except (TypeError, ValueError): ds_clean = str(ds_v).strip()
                        mw_dists[idx].set(ds_clean)
                    except Exception as exc:
                        _log(f"failed to set v_mw_dist[{idx}]: {exc}")
                # v102: capture top/bot screen depths if present
                try:
                    top_v = (loc_sheet.cell(row=r, column=top_col).value
                             if top_col else None)
                    bot_v = (loc_sheet.cell(row=r, column=bot_col).value
                             if bot_col else None)
                    if top_v is not None or bot_v is not None:
                        try: top_f = float(top_v) if top_v is not None else None
                        except (TypeError, ValueError): top_f = None
                        try: bot_f = float(bot_v) if bot_v is not None else None
                        except (TypeError, ValueError): bot_f = None
                        if top_f is not None or bot_f is not None:
                            well_depths[well_name] = (top_f, bot_f)
                except Exception as exc:
                    _log(f"failed to read top/bot for {well_name}: {exc}")
                if idx >= 6:
                    break
            _log(f"Wells imported from Location sheet: {seen_wells}")
            _log(f"  total: {n_names} written into v_mw_names slots")

        data_sheet = None
        data_keywords = ("model data", "modeldata", "data",
                         "concentrations", "measurements", "field data",
                         "fielddata")
        for nm in wb.sheetnames:
            low = nm.strip().lower()
            if low in data_keywords or "data" in low:
                if loc_sheet is not None and wb[nm] is loc_sheet:
                    continue
                data_sheet = wb[nm]
                _log(f"Picked data sheet: {nm!r}")
                break
        if data_sheet is None:
            _log("No 'Model Data' / 'Data' sheet found")

        data_header_row = 1
        if data_sheet is not None:
            for r in range(1, 30):
                a = data_sheet.cell(row=r, column=1).value
                c_ = data_sheet.cell(row=r, column=3).value
                if a is None or c_ is None:
                    continue
                a_s = str(a).strip().lower()
                c_s = str(c_).strip().lower()
                if ("date" in a_s or "time" in a_s) and \
                   ("analyte" in c_s or "compound" in c_s):
                    data_header_row = r
                    break
            _log(f"Data sheet header_row detected at row {data_header_row}")

        if loc_sheet is None and data_sheet is not None:
            _log("Falling back to deriving well names from Data sheet")
            wells_in_data = []
            for r in range(data_header_row + 1,
                           min(data_header_row + 2000,
                               (data_sheet.max_row or 0) + 1)):
                w = data_sheet.cell(row=r, column=2).value
                if w is None:
                    continue
                ws = str(w).strip()
                if not ws or ws.lower() in BAD_NAMES:
                    continue
                if ws not in wells_in_data:
                    wells_in_data.append(ws)
                if len(wells_in_data) >= 7:
                    break
            _log(f"Wells extracted from Data sheet: {wells_in_data}")
            mw_names = getattr(app, "v_mw_names", []) or []
            for i, name in enumerate(wells_in_data[:len(mw_names)]):
                try:
                    mw_names[i].set(name); n_names += 1
                except Exception:
                    pass
            _log(f"Derived {n_names} well names from Data sheet")

        # ── Parse observations from the Data sheet ───────────────
        latest = {}
        unique_analytes = []
        if data_sheet is not None:
            _log(f"Data sheet using header row: {data_header_row}")
            for r in range(data_header_row + 1,
                           (data_sheet.max_row or 0) + 1):
                dval = data_sheet.cell(row=r, column=1).value
                well = data_sheet.cell(row=r, column=2).value
                anal = data_sheet.cell(row=r, column=3).value
                cval = data_sheet.cell(row=r, column=4).value
                if well is None or anal is None or cval is None:
                    continue
                well_str = str(well).strip()
                anal_str = str(anal).strip()
                if (not well_str or not anal_str
                        or well_str.lower() in BAD_NAMES):
                    continue
                try:
                    conc = float(cval)
                except (TypeError, ValueError):
                    try:
                        conc = float(str(cval).replace(",", "").strip())
                    except Exception:
                        continue
                sort_key = r
                try:
                    if hasattr(dval, "timestamp"):
                        sort_key = dval.timestamp()
                    elif isinstance(dval, (int, float)):
                        sort_key = float(dval)
                except Exception:
                    sort_key = r
                date_str = ""
                try:
                    date_str = (dval.isoformat() if hasattr(dval, "isoformat")
                                else str(dval) if dval is not None else "")
                except Exception:
                    date_str = ""
                wkey = well_str.lower(); akey = anal_str.lower()
                cur = latest.get((wkey, akey))
                if cur is None or sort_key > cur[0]:
                    latest[(wkey, akey)] = (sort_key, conc, well_str,
                                            anal_str, date_str)
                if anal_str not in unique_analytes:
                    unique_analytes.append(anal_str)
            _log(f"Latest-conc dict size: {len(latest)} (well,analyte) keys")
            _log(f"Unique analytes found in Data sheet "
                 f"({len(unique_analytes)}): {unique_analytes}")

        def _app_name(attr):
            try:
                v = getattr(app, attr, None)
                if v is None:
                    return ""
                if hasattr(v, "get"):
                    return str(v.get()).strip()
                return str(v).strip()
            except Exception:
                return ""

        pfaa1_name = _app_name("v_pfaa1")
        pfaa2_name = _app_name("v_pfaa2")
        pre1_name  = _app_name("v_pre1")
        pre2_name  = _app_name("v_pre2")
        _log(f"App species: PFAA-1={pfaa1_name!r}, PFAA-2={pfaa2_name!r}, "
             f"Pre-1={pre1_name!r}, Pre-2={pre2_name!r}")

        def _match_analyte(stripped, species):
            if not species or species.lower() in ("none", ""):
                return False
            a_norm = species.lower().replace(" ", "").replace("-", "")
            a_norm = a_norm.replace("perfluoro", "")
            if not a_norm:
                return False
            return (a_norm in stripped) or (stripped in a_norm)

        analyte_set = {a for a in unique_analytes}
        pfaa1_match = pfaa2_match = pre1_match = pre2_match = None
        for a in analyte_set:
            s = a.lower().replace(" ", "").replace("-", "")
            s = s.replace("perfluoro", "")
            if pfaa1_match is None and _match_analyte(s, pfaa1_name):
                pfaa1_match = a
            if pfaa2_match is None and _match_analyte(s, pfaa2_name):
                pfaa2_match = a
            if pre1_match is None and _match_analyte(s, pre1_name):
                pre1_match = a
            if pre2_match is None and _match_analyte(s, pre2_name):
                pre2_match = a
        _log(f"Resolved matches: PFAA-1={pfaa1_match!r}, "
             f"PFAA-2={pfaa2_match!r}, Pre-1={pre1_match!r}, "
             f"Pre-2={pre2_match!r}")

        mw_conc  = getattr(app, "v_mw_conc",  []) or []
        mw_conc2 = getattr(app, "v_mw_conc2", []) or []
        mw_names = getattr(app, "v_mw_names", []) or []

        well_to_idx = {}
        for i, sv in enumerate(mw_names):
            try:
                nm = str(sv.get()).strip()
            except Exception:
                nm = ""
            if nm:
                well_to_idx[nm.lower()] = i

        obs_records = []
        for (wkey, akey), (skey, conc, well_str, anal_str,
                           date_str) in latest.items():
            species_label = None
            if pfaa1_match and akey == pfaa1_match.lower():
                species_label = "PFAA-1"
            elif pfaa2_match and akey == pfaa2_match.lower():
                species_label = "PFAA-2"
            elif pre1_match and akey == pre1_match.lower():
                species_label = "Precursor-1"
            elif pre2_match and akey == pre2_match.lower():
                species_label = "Precursor-2"
            idx = well_to_idx.get(wkey)
            if idx is None or species_label is None:
                _log(f"  miss: well={well_str!r}, species={anal_str!r}")
                continue
            _log(f"  HIT:  well={well_str!r}, species={species_label} "
                 f"({anal_str!r}), conc={conc}")
            obs_records.append({
                "well": well_str,
                "species": species_label,
                "analyte": anal_str,
                "concentration": conc,
                "date": date_str,
                "sort_key": skey,
            })
            try:
                if species_label == "PFAA-1" and idx < len(mw_conc):
                    mw_conc[idx].set(f"{conc:g}"); n_concs += 1
                elif species_label == "PFAA-2" and idx < len(mw_conc2):
                    mw_conc2[idx].set(f"{conc:g}"); n_concs += 1
            except Exception as exc:
                _log(f"failed to set v_mw_conc[{idx}]: {exc}")

        _log(f"Total observation records collected: {len(obs_records)}")
        _log(f"§10 cells filled: {n_concs}")

        try:
            import json as _json
            obs_path = os.path.join(proj_dir, "mw_observations.json")
            payload = {
                "source_xlsx": xlsx_path,
                "analytes_found": sorted(unique_analytes),
                "records": obs_records,
                # v102: per-well screen depths in user's unit (m or ft)
                # — generate_input_file converts to meters as needed.
                "well_depths": {
                    w: {"top": d[0], "bot": d[1]}
                    for w, d in well_depths.items()
                },
            }
            with open(obs_path, "w", encoding="utf-8") as fh:
                _json.dump(payload, fh, indent=2, default=str)
            n_total_obs = len(obs_records)
            _log(f"Wrote {obs_path} with {n_total_obs} records")
        except Exception as exc:
            _log(f"failed to write mw_observations.json: {exc}")

    except Exception as exc:
        _log(f"import failed: {exc}")
        _flush_log(log_path, log_lines)
        return f"import failed: {exc}"

    _flush_log(log_path, log_lines)

    if n_names == 0 and n_concs == 0:
        return ("0 wells loaded — see calibration_import.log next "
                "to your .xlsx for diagnostics")
    return (f"Imported {n_names} well names, {n_concs} value(s) into "
            f"§10, {n_total_obs} total observation(s) → "
            f"mw_observations.json")


def run(app, parent=None):
    state = get_state()
    work_dir = state.work_dir or os.getcwd()
    txt_path = os.path.join(work_dir, "calibration_inputs.txt")

    here = os.path.dirname(os.path.abspath(__file__))
    project = os.path.abspath(os.path.join(here, "..", "..", ".."))
    is_detailed = (getattr(app, "active_sheet", "Simple") == "Detailed_2")
    template = ("CalibrationTemplate_Detailed.xlsx" if is_detailed
                else "CalibrationTemplate_Simple.xlsx")
    default_path = os.path.join(project, template)
    existing = _load_existing(txt_path)
    initial = existing or (default_path if os.path.exists(default_path) else "")

    root = tk.Toplevel(parent or app)
    root.title("Calibration Data Upload")
    root.configure(bg="#F0F0F0")
    try: root.withdraw()
    except Exception: pass
    try:
        root.grab_set()
    except Exception:
        pass

    try:
        tv_style = ttk.Style(root)
        try:
            tv_style.theme_use("clam")
        except Exception:
            pass
        tv_style.configure("Treeview", rowheight=36, font=FONT_LABEL,
                           background="#FFFFFF", fieldbackground="#FFFFFF")
        tv_style.configure("Treeview.Heading", font=FONT_BOLD,
                           padding=(4, 6))
    except Exception:
        pass

    title = tk.Frame(root, bg="#F0F0F0"); title.pack(pady=10)
    tk.Label(title, text="Calibration Data Upload",
             font=FONT_TITLE, bg="#F0F0F0").pack()

    main = tk.Frame(root, bg="#F0F0F0", padx=20, pady=10)
    main.pack(expand=True, fill="both")

    tk.Label(main, text="Select an Excel file containing calibration data:",
             font=FONT_LABEL, bg="#F0F0F0").pack(pady=10, anchor="w")

    file_frame = tk.Frame(main, bg="#F0F0F0"); file_frame.pack(fill="x", pady=10)
    tk.Label(file_frame, text="Excel File:", font=FONT_BOLD, bg="#F0F0F0",
             width=12, anchor="w").pack(side="left")
    path_var = tk.StringVar(value=initial)
    file_entry = tk.Entry(file_frame, textvariable=path_var, font=FONT_LABEL,
                           width=50)
    file_entry.pack(side="left", padx=5, fill="x", expand=True)

    sheet_frame = tk.Frame(main, bg="#F0F0F0"); sheet_frame.pack(fill="x", pady=10)
    tk.Label(sheet_frame, text="Preview Sheet:", font=FONT_BOLD,
             bg="#F0F0F0", width=12, anchor="w").pack(side="left")
    sheet_var = tk.StringVar(value="model location")

    preview_frame = tk.Frame(main, bg="#F0F0F0")
    preview_frame.pack(fill="both", expand=True, pady=10)
    info_label = tk.Label(preview_frame, text="File Preview:",
                           font=FONT_BOLD, bg="#F0F0F0")
    info_label.pack(anchor="w", pady=(0, 5))

    table_frame = tk.Frame(preview_frame, bg="#F0F0F0")
    table_frame.pack(fill="both", expand=True)
    h_sb = ttk.Scrollbar(table_frame, orient="horizontal")
    h_sb.pack(side="bottom", fill="x")
    v_sb = ttk.Scrollbar(table_frame, orient="vertical")
    v_sb.pack(side="right", fill="y")
    preview_tree = ttk.Treeview(table_frame, xscrollcommand=h_sb.set,
                                 yscrollcommand=v_sb.set, show="headings")
    preview_tree.pack(side="left", fill="both", expand=True)
    h_sb.config(command=preview_tree.xview)
    v_sb.config(command=preview_tree.yview)

    def _update_preview():
        for it in preview_tree.get_children():
            preview_tree.delete(it)
        p = path_var.get().strip()
        if not p or not os.path.exists(p):
            info_label.config(text="File Preview: No file selected.")
            return
        if not _HAS_OPENPYXL:
            info_label.config(text="File Preview: openpyxl not installed.")
            return
        try:
            wb = load_workbook(p, data_only=True)
            target = sheet_var.get()
            sheet_name = None
            for s in wb.sheetnames:
                if s.lower() == target.lower():
                    sheet_name = s; break
            if sheet_name is None:
                sheet_name = wb.sheetnames[0] if wb.sheetnames else None
            if sheet_name is None:
                wb.close()
                info_label.config(text="File Preview: No sheets found.")
                return
            ws = wb[sheet_name]
            info_label.config(
                text=(f"File: {os.path.basename(p)} | Sheet: {ws.title} | "
                      f"Dimensions: {ws.max_row} rows x {ws.max_column} cols"))
            max_rows = min(30, ws.max_row)
            max_cols = min(15, ws.max_column)
            col_names = []
            for ci in range(1, max_cols + 1):
                v = ws.cell(row=1, column=ci).value
                col_names.append(str(v) if v is not None
                                 else get_column_letter(ci))
            cols = ["#"] + col_names
            preview_tree["columns"] = cols
            preview_tree.column("#", width=50, anchor="center",
                                 stretch=False, minwidth=40)
            preview_tree.heading("#", text="#", anchor="center")
            for cn in col_names:
                preview_tree.column(cn, width=110, anchor="w", stretch=True,
                                     minwidth=60)
                preview_tree.heading(cn, text=cn, anchor="center")
            data_start = 2
            data_n = min(max_rows - 1, ws.max_row - 1)
            for di in range(data_start, data_start + data_n):
                if di > ws.max_row:
                    break
                row_values = [str(di - 1)]
                for ci in range(1, max_cols + 1):
                    val = ws.cell(row=di, column=ci).value
                    sval = "" if val is None else str(val)
                    if len(sval) > 30:
                        sval = sval[:27] + "..."
                    row_values.append(sval)
                tag = "even" if (di - data_start) % 2 == 0 else "odd"
                preview_tree.insert("", "end", values=row_values, tags=(tag,))
            preview_tree.tag_configure("odd", background="#FFFFFF")
            preview_tree.tag_configure("even", background="#F0F0F0")
            if ws.max_row > max_rows or ws.max_column > max_cols:
                preview_tree.insert("", "end", tags=("note",),
                                     values=[f"... showing first {data_n} rows "
                                             f"x {max_cols} cols"]
                                            + [""] * max_cols)
                preview_tree.tag_configure("note", background="#E8E8E8",
                                            foreground="#666666")
            wb.close()
        except Exception as e:
            info_label.config(text=f"File Preview: error - {e}")

    for opt in ("model location", "model data"):
        tk.Radiobutton(sheet_frame, text=opt, variable=sheet_var, value=opt,
                       command=_update_preview,
                       font=FONT_LABEL, bg="#F0F0F0").pack(side="left", padx=10)

    def _browse():
        p = filedialog.askopenfilename(
            title="Select Excel File for Calibration",
            initialdir=(os.path.dirname(path_var.get()) if path_var.get()
                        else project),
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"),
                       ("All files", "*.*")],
            parent=root)
        if p:
            path_var.set(p); _update_preview()

    tk.Button(file_frame, text="Browse...", command=_browse,
              font=FONT_LABEL, width=10).pack(side="left", padx=5)

    if initial:
        _update_preview()

    def _save_and_exit():
        p = path_var.get().strip()
        if not p:
            messagebox.showerror("Error", "Please select an Excel file.",
                                 parent=root); return
        if not os.path.isfile(p):
            messagebox.showerror("Error", f"File not found:\n{p}",
                                 parent=root); return
        if _HAS_OPENPYXL:
            try:
                wb = load_workbook(p, data_only=True)
                names = [s.lower() for s in wb.sheetnames]
                missing = [s for s in ("model location", "model data")
                           if s not in names]
                wb.close()
                if missing:
                    messagebox.showwarning(
                        "Warning",
                        f"Excel file is missing sheets: {', '.join(missing)}",
                        parent=root)
            except Exception as e:
                messagebox.showerror("Error", f"Invalid Excel file:\n{e}",
                                     parent=root); return
        import_status = ""
        try:
            import_status = _import_xlsx_into_app(app, p)
        except Exception as exc:
            import_status = f"import failed: {exc}"
        if _save_path(p, txt_path):
            messagebox.showinfo("Saved",
                                f"Calibration template path saved.\n"
                                f"{os.path.basename(txt_path)}\n\n"
                                f"{import_status}",
                                parent=root)
            try: root.grab_release()
            except Exception: pass
            root.destroy()
        else:
            messagebox.showerror("Error",
                                 f"Could not write {txt_path}",
                                 parent=root)

    def _cancel():
        try: root.grab_release()
        except Exception: pass
        root.destroy()

    bar = tk.Frame(root, bg="#F0F0F0"); bar.pack(pady=12)
    tk.Button(bar, text="OK",     width=10, font=FONT_BTN,
              command=_save_and_exit).pack(side="left", padx=6)
    tk.Button(bar, text="Cancel", width=10, font=FONT_BTN,
              command=_cancel).pack(side="left", padx=6)
    tk.Button(bar, text="Help",   width=10, font=FONT_BTN,
              command=_open_help).pack(side="left", padx=6)

    root.protocol("WM_DELETE_WINDOW", _cancel)
    root.update_idletasks()
    w = max(root.winfo_reqwidth() + 32, 980)
    h = max(root.winfo_reqheight() + 24, 640)
    try:
        sw = root.winfo_screenwidth(); sh = root.winfo_screenheight()
        w = min(w, int(sw * 0.95)); h = min(h, int(sh * 0.92))
        x = max(0, (sw - w) // 2); y = max(0, (sh - h) // 2 - 30)
        root.geometry(f"{w}x{h}+{x}+{y}")
    except Exception:
        root.geometry(f"{w}x{h}")
    root.minsize(min(w, 800), min(h, 540))
    root.resizable(True, True)
    try:
        root.deiconify(); root.lift(); root.focus_force()
    except Exception:
        pass
    root.wait_window()
