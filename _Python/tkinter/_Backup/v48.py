"""
REMFluor-MD Model Input Dashboard — TKINTER variant.

Sister build to the PySide6 version at  _Python/main.py .  Same Excel/PDF
storyboard ground truth, same Figures/ folder, same VBA macro dispatch
table.  Lives at  REMFluorMD_v2.6/_Python/tkinter/main.py  so BASE_DIR
walks up two levels to reach the project root.

Visual / behavioral spec:
    - Top header bar:  dark teal  #074F69 with white italic title
    - Bottom bar:      pure black #000000
    - Body:            light gray #F2F2F2
    - ESTCP logo:      loaded from BASE_DIR/Figures/ (PNG/GIF/JPG via PIL)
    - Body font:       Calibri (Excel default)
    - "?" help boxes:  squared, sit flush-adjacent to their entry cells
    - Layout:          fixed proportions inside scrollable Canvas, both
                       horizontal and vertical scrollbars when window is
                       smaller than the canvas

Buttons call Python scripts / open Quarto files using the same paths as
the VBA macros.  Run with:
    cd _Python/tkinter
    python main.py
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import subprocess
import os
import sys
import glob
from datetime import datetime

# --- Standalone functions (replace exe subprocess calls) ---
try:
    from functions.state import get_state
    from functions import (
        clear_for_restore,
        restore_from_example,
        restore_from_saved,
        generate_input_file,
        popups_retardation,
        popups_source_remediation,
        popups_transformation,
        popups_numerical,
        popups_cellsize,
        popups_GWvelocity,
        popups_longevity,
        popups_mass_discharge_import,
        popups_calibration,
        popups_heterogeneity,
        run_model,
        xlsm_io,
    )
    _FUNCS_LOADED = True
except Exception as _e:
    print(f"Warning: could not load functions package: {_e}")
    _FUNCS_LOADED = False

# ─────────────────────────────────────────────────────────────────────────────
# HIGH-DPI / RETINA SHARPNESS FIX  (must run BEFORE creating any Tk widgets)
# ─────────────────────────────────────────────────────────────────────────────
# Without this, Windows treats the app as a 96-DPI legacy app and bitmap-
# stretches the entire window on scaled displays (125% / 150% / 175%), making
# text and images look blurry. With it, Tk renders at the display's native
# pixel resolution.
def _enable_high_dpi_awareness():
    if sys.platform != "win32":
        return
    try:
        import ctypes
        # Windows 8.1+: per-monitor DPI awareness (best)
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
            return
        except (AttributeError, OSError):
            pass
        # Windows Vista/7/8: system DPI awareness (good enough)
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except (AttributeError, OSError):
            pass
    except Exception:
        pass

_enable_high_dpi_awareness()

# ─────────────────────────────────────────────────────────────────────────────
# COLOR CONSTANTS  (sampled directly from the PDF storyboard pixels)
# ─────────────────────────────────────────────────────────────────────────────
BG_HEADER_BAR  = "#074F69"   # top bar – DARK TEAL (sampled from PDF)
BG_BOTTOM_BAR  = "#000000"   # bottom bar – PURE BLACK (sampled)
BG_MAIN        = "#F2F2F2"   # body background – LIGHT GRAY (sampled)
BG_WHITE       = "#FFFFFF"
BG_LEGEND_BAR  = "#FFFFFF"
BG_INPUT_BLUE  = "#FFFFFF"   # legend "Enter value directly"           – WHITE
BG_FORMULA     = "#DAE9F8"   # legend "Cell with formula or default..." – light blue
BG_PULLDOWN    = "#FBE2D5"   # legend "Pull Down Menu"                  – peach
BG_LOCKED      = "#000000"   # legend "Calculated value..."             – BLACK
FG_LOCKED      = "#FFFFFF"   # text color used inside black locked cells
BG_SECTION_HDR = "#DAEEF3"
BTN_FILL       = "#D9D9D9"   # standard gray button
BTN_FILL_BLUE  = "#BDD7EE"   # Run Model buttons (light blue)
BTN_FILL_GREEN = "#E6FFDC"   # Authors / Save Data button (light green)
BTN_FILL_RED   = "#FFD7D7"   # red-tinted button background

FG_TITLE       = "#FFFFFF"   # white title on dark teal
FG_SECTION     = "#0070C0"   # blue section headers (sampled)
FG_BTN_NAVY    = "#002060"
FG_BTN_RED     = "#FF0000"
FG_BTN_BLUE    = "#0070C0"
FG_BTN_GREEN   = "#385723"
FG_INPUT       = "#000000"
FG_GREY        = "#7F7F7F"
FG_YELLOW      = "#FFFF00"
FG_HELP        = "#FF0000"   # red "?" help link text

# ─────────────────────────────────────────────────────────────────────────────
# FONTS  (Calibri = Excel default; Tk falls back gracefully if unavailable)
# ─────────────────────────────────────────────────────────────────────────────
FG_TITLE_GREEN = "#DAF2D0"   # "REMFluor" portion of the title

# ─── FONTS as Tk NAMED FONTS ────────────────────────────────────────────────
# Storing fonts as named fonts (instead of tuples) lets us resize every widget
# at runtime by just calling .config(size=...) on the named font.  The string
# constants below are the names; the actual sizes are configured by
# REMFluorApp._init_named_fonts() after the Tk root is created.
FONT_TITLE     = "AppTitle"
FONT_VERSION   = "AppVersion"
FONT_SECTION   = "AppSection"
FONT_LABEL     = "AppLabel"
FONT_LABEL_B   = "AppLabelB"     # bold body label  (PFAA 1, etc.)
FONT_LABEL_I   = "AppLabelI"     # italic body label
FONT_LABEL_BI  = "AppLabelBI"    # bold italic
FONT_LABEL_SM  = "AppLabelSm"
FONT_LABEL_SMI = "AppLabelSmI"   # small italic
FONT_LABEL_XS  = "AppLabelXs"
FONT_INPUT     = "AppInput"
FONT_BTN       = "AppBtn"
FONT_BTN_SM    = "AppBtnSm"
FONT_BTN_LG    = "AppBtnLg"
FONT_BTN_CALIB = "AppBtnCalib"
FONT_HELP      = "AppHelp"
FONT_RADIO     = "AppRadio"

# Baseline (zoom = 1.0) sizes.  Title/version stay big as the user requested;
# every other size is dialled down so the dashboard fits on small displays.
# The user can press Ctrl++ / Ctrl+- / Ctrl+0 / Ctrl+wheel to zoom.
_FONT_DEFS = {
    # name              family            size  weight    slant
    FONT_TITLE:        ("Arial Narrow",   28,   "bold",   "roman"),
    FONT_VERSION:      ("Arial Narrow",   26,   "bold",   "roman"),
    FONT_SECTION:      ("Calibri",        11,   "bold",   "roman"),
    FONT_LABEL:        ("Calibri",         9,   "normal", "roman"),
    FONT_LABEL_B:      ("Calibri",         9,   "bold",   "roman"),
    FONT_LABEL_I:      ("Calibri",         9,   "normal", "italic"),
    FONT_LABEL_BI:     ("Calibri",         9,   "bold",   "italic"),
    FONT_LABEL_SM:     ("Calibri",         8,   "normal", "roman"),
    FONT_LABEL_SMI:    ("Calibri",         8,   "normal", "italic"),
    FONT_LABEL_XS:     ("Calibri",         8,   "normal", "roman"),
    FONT_INPUT:        ("Calibri",         9,   "normal", "roman"),
    FONT_BTN:          ("Calibri",        10,   "normal", "italic"),
    FONT_BTN_SM:       ("Calibri",         9,   "normal", "italic"),
    FONT_BTN_LG:       ("Calibri",        13,   "bold",   "italic"),
    FONT_BTN_CALIB:    ("Calibri",        10,   "bold",   "roman"),
    FONT_HELP:         ("Arial",           9,   "bold",   "italic"),
    FONT_RADIO:        ("Calibri",        12,   "normal", "roman"),
}


# ─────────────────────────────────────────────────────────────────────────────
# SCRIPT / FILE PATHS  (mirrored exactly from VBA macro targets)
# ─────────────────────────────────────────────────────────────────────────────
# This file lives at  REMFluorMD_v2.6/_Python/tkinter/main.py
# BASE_DIR must point at REMFluorMD_v2.6 so Figures/, dist/, docs/, and
# the .xlsm workbook resolve correctly.
_HERE    = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.abspath(os.path.join(_HERE, "..", ".."))

# When running as a PyInstaller --onefile build, sys._MEIPASS is the
# temp folder that holds bundled data (e.g. the Figures/ folder added
# via --add-data). We look there first, then fall back to BASE_DIR.
def _resolve_figures_dir():
    bundled = getattr(sys, "_MEIPASS", None)
    if bundled and os.path.isdir(os.path.join(bundled, "Figures")):
        return os.path.join(bundled, "Figures")
    return os.path.join(BASE_DIR, "Figures")

FIGURES_DIR = _resolve_figures_dir()

def _exe(*parts):
    return os.path.join(BASE_DIR, "dist", *parts)

def _html(*parts):
    return os.path.join(BASE_DIR, "docs", "_site", *parts)

EXES = {
    "GWVelocityCalculator": (
        _exe("popups_GWvelocity", "popups_GWvelocity.exe"),
        ["{workbook}", "{sheet}"]),
    "HeterogeneityCalculator_Unconsolidated_Media": (
        _exe("popups_heterogeneity", "popups_heterogeneity.exe"),
        ["{workbook}", "{unitflag}", "Unconsolidated Media"]),
    "HeterogeneityCalculator_Fractured_Rock": (
        _exe("popups_heterogeneity", "popups_heterogeneity.exe"),
        ["{workbook}", "{unitflag}", "Fractured Rock"]),
    "CalculrateRetardationFactors": (
        _exe("popups_retardation", "popups_retardation.exe"),
        ["{workbook}", "{sheet}"]),
    "ModelingTransformationLowK": (
        _exe("popups_transformation", "popups_transformation.exe"),
        ["{workbook}"]),
    "SourceOption2": (
        _exe("popups_mass_discharge_import", "popups_mass_discharge_import.exe"),
        ["{workbook}", "{sheet}"]),
    "SourceRemediation": (
        _exe("popups_source_remediation", "popups_source_remediation.exe"),
        ["{workbook}", "{sheet}"]),
    "LongevityTool": (
        _exe("popups_longevity", "popups_longevity.exe"),
        ["{workbook}", "{sheet}"]),
    "CalibrationDataLoader": (
        _exe("popups_calibration", "popups_calibration.exe"),
        ["{workbook}", "{sheet}"]),
    "ChangeNumericalParameters": (
        _exe("popups_numerical", "popups_numerical.exe"),
        ["{workbook}", "{sheet}"]),
    "OpenAppendix_2_1_Relative_EXE": (
        _exe("popups_cellsize", "popups_cellsize.exe"),
        ["{workbook}", "{sheet}"]),
    "RunPythonScript": (
        _exe("input_variables", "input_variables.exe"),
        ["{workbook}", "{sheet}"]),
    "Save_Data": (
        _exe("generate_input_file", "generate_input_file.exe"),
        ["{workbook}", "{sheet}", "ask"]),
    "Load_Data_Step1": (
        _exe("clear_for_restore", "clear_for_restore.exe"),
        ["{workbook}", "{sheet}"]),
    "Load_Data_Step2": (
        _exe("restore_from_saved_folder", "restore_from_saved_folder.exe"),
        ["{workbook}", "{sheet}"]),
    "Paste_Example_Step1": (
        _exe("clear_for_restore", "clear_for_restore.exe"),
        ["{workbook}", "{sheet}"]),
    "Paste_Example_Step2": (
        _exe("restore_from_example_folder", "restore_from_example_folder.exe"),
        ["{workbook}", "{sheet}"]),
    "Clear_Data": (
        _exe("clear_for_restore", "clear_for_restore.exe"),
        ["{workbook}", "{sheet}"]),
    "Save_Data_Calibration_Step1": (
        _exe("generate_input_file", "generate_input_file.exe"),
        ["{workbook}", "{sheet}", "noask"]),
    "Save_Data_Calibration_Step2": (
        _exe("export_calibration_data", "export_calibration_data.exe"),
        ["{workbook}", "{sheet}"]),
    "Run_Machine_Based_Calibration": (
        _exe("cali_1", "cali_1.exe"),
        ["{workbook}", "{sheet}"]),
    "Load_Optimal_Data": (
        _exe("restore_from_optimal", "restore_from_optimal.exe"),
        ["{workbook}", "{sheet}"]),
}

HTML_APPENDIX = {
    "OpenAppendix_2_1_Relative": _html("appendix", "appendix_2_1.html"),
    "OpenAppendix_2_2_Relative": _html("appendix", "appendix_2_2.html"),
    "OpenAppendix_4_2_Relative": _html("appendix", "appendix_4_2.html"),
    "OpenAppendix_6_1_Relative": _html("appendix", "appendix_6_1.html"),
    "OpenAppendix_7_1_Relative": _html("appendix", "appendix_7_1.html"),
    "OpenAppendix_8_1_Relative": _html("appendix", "appendix_8_1.html"),
    "OpenAppendix_9_1_Relative": _html("appendix", "appendix_9_1.html"),
}

HTML_CHICKLETS = {
    "OpenTable1":               ("Step1_SystemUnits.html", ""),
    "OpenTable2_1_XDirection":  ("Step2_ModelConfiguration.html", "model-size-in-direction-of-groundwater-flow-x-direction"),
    "OpenTable2_1_YDirection":  ("Step2_ModelConfiguration.html", "model-size-in-direction-of-groundwater-flow-y-direction"),
    "OpenTable2_1_ZDirection":  ("Step2_ModelConfiguration.html", "model-size-in-direction-of-groundwater-flow-z-direction"),
    "OpenTable2_2":             ("Step2_ModelConfiguration.html", "finite-difference-cell-size"),
    "OpenTable2_3":             ("Step2_ModelConfiguration.html", "source-width"),
    "OpenTable2_4":             ("Step2_ModelConfiguration.html", "thickness-of-source-below-water-table"),
    "OpenTable2_5":             ("Step2_ModelConfiguration.html", "starting-year-of-simulation"),
    "OpenTable2_6":             ("Step2_ModelConfiguration.html", "ending-year-of-simulation"),
    "OpenTable3_1":             ("Step3_GroundwaterDarcyVelocity.html", "groundwater-darcy-velocity-vd"),
    "OpenTable3_2":             ("Step3_GroundwaterDarcyVelocity.html", "transmissive-zone-effective-porosity"),
    "OpenTable4_1":             ("Step4_HydrogeologicSettingAndMatrixDiffusion.html", "unconsolidated-aquifers---low-k-media-details"),
    "OpenTable4_2":             ("Step4_HydrogeologicSettingAndMatrixDiffusion.html", "low-k-zone-total-porosity"),
    "OpenTable4_3":             ("Step4_HydrogeologicSettingAndMatrixDiffusion.html", "low-k-zone-tortuosity"),
    "OpenTable5_1":             ("Step5_PFASTransportProperties.html", "constituent"),
    "OpenTable5_2":             ("Step5_PFASTransportProperties.html", "retardation-factor-calculations"),
    "OpenTable5_3":             ("Step5_PFASTransportProperties.html", "retardation-factor-in-t-zone-transmissive-zone"),
    "OpenTable5_4":             ("Step5_PFASTransportProperties.html", "retardation-factor-in-low-k-media"),
    "OpenTable5_5":             ("Step5_PFASTransportProperties.html", "detailed-model-only-precursor-transformation-to-pfaas"),
    "OpenTable5_6":             ("Step5_PFASTransportProperties.html", "microbial-yield-factor"),
    "OpenTable5_7":             ("Step5_PFASTransportProperties.html", "molecular-diffusion-coefficient-in-free-water"),
    "OpenTable6_1":             ("Step6_PlumeTransport.html", "longitudinal-dispersivity"),
    "OpenTable6_2":             ("Step6_PlumeTransport.html", "transverse-dispersivity"),
    "OpenTable6_3":             ("Step6_PlumeTransport.html", "vertical-dispersivity"),
    "OpenTable7_1":             ("Step7_PFASSourceTerm.html", "initial-source-concentration"),
    "OpenTable8_1":             ("Step8_SourceRemediation.html", "percent-source-mass-removed-by-remediation"),
    "OpenTable8_2":             ("Step8_SourceRemediation.html", "remediation-started-in-year"),
    "OpenTable8_3":             ("Step8_SourceRemediation.html", "remediation-ended-in-year"),
    "OpenTable9_1":             ("Step9_PlumeRemediationPSB.html", "psb-freundlich-exponent-a"),
    "OpenTable9_2":             ("Step9_PlumeRemediationPSB.html", "psb-freundlich-kf"),
    "OpenTable9_3":             ("Step9_PlumeRemediationPSB.html", "year-psb-barrier-installed"),
    "OpenTable9_4":             ("Step9_PlumeRemediationPSB.html", "total-width-of-psb-in-x-direction"),
    "OpenTable9_5":             ("Step9_PlumeRemediationPSB.html", "psb-loading-fcac"),
    "OpenTable10_1":            ("Step10_FieldDataToCalibrate.html", "sample-year"),
    "OpenTable10_2":            ("Step10_FieldDataToCalibrate.html", "monitoring-well-name"),
    "OpenTable10_3":            ("Step10_FieldDataToCalibrate.html", "concentration-measured"),
    "OpenTable10_4":            ("Step10_FieldDataToCalibrate.html", "distance-from-source"),
    "OpenTable11_1":            ("Step11_ModelingParameters.html", "see-results-every"),
}

# ─────────────────────────────────────────────────────────────────────────────
# ACTION DISPATCHER
# ─────────────────────────────────────────────────────────────────────────────
XLSM_PATH = os.path.join(BASE_DIR,
                         "REMFluor-MD Interface Storyboard v2.6.xlsm")

def _open_html(path_or_url: str, wait: bool = False):
    if sys.platform == "win32":
        url = "file:///" + path_or_url.replace("\\", "/")
        subprocess.Popen(["cmd", "/c", "start", "", url],
                         shell=False, creationflags=subprocess.CREATE_NO_WINDOW
                         if hasattr(subprocess, "CREATE_NO_WINDOW") else 0)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", path_or_url])
    else:
        subprocess.Popen(["xdg-open", path_or_url])


def _launch_exe(exe_path: str, args: list, wait: bool = True):
    if not os.path.exists(exe_path):
        messagebox.showwarning(
            "Executable Not Found",
            f"Expected executable not found:\n{exe_path}\n\n"
            "Ensure the dist\\ folder is present next to this script."
        )
        return
    # Standalone-pipeline step 1 — push current app state into the .xlsm
    # so the popup EXE (openpyxl-based) sees the latest user edits.
    if _FUNCS_LOADED and _app_ref is not None:
        try:
            sheet_name = _current_sheet()
            xlsm_io.push_state_to_xlsm(_app_ref, XLSM_PATH, sheet_name)
        except Exception as exc:
            print(f"[main] xlsm push failed (continuing): {exc}")
    cmd = [exe_path] + args
    if wait:
        subprocess.run(cmd)
    else:
        subprocess.Popen(cmd)
    # Standalone-pipeline step 3 — pull the updated cell values back from
    # the .xlsm into the Tk UI so the user sees the result of the popup.
    if wait and _app_ref is not None:
        _app_ref.refresh_from_xlsm()


# ─── XLSM read-back (for cells the dist EXEs mutate) ────────────────
# Maps a Tk StringVar attribute name to a cell address on the active
# sheet (Detailed_2 / Simple).  Subset that the EXEs are known to
# modify — anything not listed is left alone.
XLSM_CELL_MAP = {
    # ── Section 1 (Site / Date) ─────────────────────────────────────
    "v_site":         "B4",
    "v_date":         "E4",
    # ── Section 2 (Model Configuration) ─────────────────────────────
    "v_x_size":       "E11",
    "v_y_size":       "E12",
    "v_z_size":       "E13",
    "v_sw_width":     "E15",
    "v_sw_thick":     "E16",
    "v_yr_start":     "E18",
    "v_yr_end":       "E19",
    "v_run_time":     "M16",   # approx run time (locked black)
    # ── Section 3 (Groundwater Darcy Velocity) ──────────────────────
    "v_darcy":        "C22",
    "v_porf":         "G22",
    # ── Section 4 (Hydrogeologic Setting / Matrix Diffusion) ────────
    "v_lowk_media":   "K26",
    "v_lowk_por":     "K27",
    "v_lowk_tort":    "K28",
    # ── Section 5 (PFAS Transport Properties) ───────────────────────
    "v_pfaa1":        "E38",
    "v_pfaa2":        "G38",
    "v_ret_trans1":   "E39",
    "v_ret_lowk1":    "E40",
    "v_ret_trans2":   "G39",
    "v_ret_lowk2":    "G40",
    "v_mol_diff":     "E44",
    # ── Section 6 (Plume Transport — Dispersivity) ──────────────────
    "v_alpha_l":      "V4",
    "v_alpha_t":      "X4",
    "v_alpha_v":      "Z4",
    # ── Section 7 (PFAS Source Term — 11 decade rows) ──────────────
    # Source years U8:U18, PFAA-1 V8:V18, PFAA-2 X8:X18.  The list
    # entries are filled programmatically below to keep the table tidy.
    # ── Section 8 (Source Remediation) ──────────────────────────────
    "v_src_rem_yr":   "D27",
    "v_src_conc_red": "D28",
    # ── Section 9 (Plume Remediation : PSB) ─────────────────────────
    "v_model_psb":    "R22",
    # Freundlich "a"
    "v_psb_a_1":      "V23",
    "v_psb_a_2":      "X23",
    "v_psb_a_3":      "Z23",
    "v_psb_a_4":      "AB23",
    # Freundlich Kf + unit dropdown
    "v_psb_kf_unit":  "U24",
    "v_psb_kf_1":     "V24",
    "v_psb_kf_2":     "X24",
    "v_psb_kf_3":     "Z24",
    "v_psb_kf_4":     "AB24",
    # PFAS molecular weight (g/mol) — only used when Kf unit is mol-based
    "v_psb_mw_1":     "V25",
    "v_psb_mw_2":     "X25",
    "v_psb_mw_3":     "Z25",
    "v_psb_mw_4":     "AB25",
    # Converted Kf in (ug/kg)(ug/L)^(-a) — auto-computed by the app
    "v_psb_kf_conv":  "V26",
    "v_psb_kf_conv2": "X26",
    "v_psb_kf_conv3": "Z26",
    "v_psb_kf_conv4": "AB26",
    # Detailed-only Section 5 transformation row (kept here for proximity)
    "v_trans_rate_3":   "K41",
    "v_trans_rate_4":   "M41",
    "v_yield_factor_3": "K42",
    "v_yield_factor_4": "M42",
    # PSB geometry / install
    "v_psb_yr":       "AB28",
    "v_psb_dist":     "X74",
    "v_psb_width":    "Y82",
    "v_psb_load":     "AA82",
    "v_psb_cells":    "AC82",
    # ── Section 10 (Field Data to Calibrate) — sample year only.
    #    Per-MW rows are populated by the EXEs but addresses depend on
    #    sheet layout; covered in a later refinement pass.
    "v_sample_yr":    "Y74",
    # ── Section 11 (Output / Numerical Parameters) ──────────────────
    "v_see_every":    "V47",
}

# Section 7 source rows — extend the map programmatically
for _i in range(11):
    XLSM_CELL_MAP[f"v_src_years_{_i}"] = f"U{8 + _i}"
    XLSM_CELL_MAP[f"v_src_pfaa1_{_i}"] = f"V{8 + _i}"
    XLSM_CELL_MAP[f"v_src_pfaa2_{_i}"] = f"X{8 + _i}"
    # Detailed-only Precursor 1 / Precursor 2 source columns (XLSM Z / AB)
    XLSM_CELL_MAP[f"v_src_pre1_{_i}"]  = f"Z{8 + _i}"
    XLSM_CELL_MAP[f"v_src_pre2_{_i}"]  = f"AB{8 + _i}"
del _i


def _read_xlsm_cells(addresses, sheet_name="Detailed_2"):
    """Read a list of A1 addresses from the workbook. Returns dict of
    addr → string-formatted value.  Silently returns {} if openpyxl is
    not installed or the workbook is unreachable."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    if not os.path.exists(XLSM_PATH):
        return {}
    try:
        wb = load_workbook(XLSM_PATH, read_only=True,
                           data_only=True, keep_vba=False)
        sh_name = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sh_name]
        out = {}
        for addr in addresses:
            v = ws[addr].value
            out[addr] = "" if v is None else str(v)
        wb.close()
        return out
    except Exception:
        return {}


def _resolve_args(template: list, sheet: str = "", unitflag: str = "2") -> list:
    return [
        a.replace("{workbook}", XLSM_PATH)
         .replace("{sheet}", sheet)
         .replace("{unitflag}", unitflag)
        for a in template
    ]


def _current_sheet() -> str:
    return getattr(_app_ref, "active_sheet", "Simple")


def _unitflag() -> str:
    return "1" if getattr(_app_ref, "v_units", None) and \
           _app_ref.v_units.get() == "feet" else "2"


_app_ref = None


def _ensure_state_work_dir():
    """Set state work_dir to BASE_DIR if not already set."""
    if _FUNCS_LOADED and _app_ref is not None:
        st = get_state()
        if not st.work_dir:
            st.work_dir = BASE_DIR


def run_script(macro_name, extra_args=None):
    sheet     = _current_sheet()
    unitflag  = _unitflag()
    _ensure_state_work_dir()

    # ── Standalone Python replacements ──────────────────────────────────
    if _FUNCS_LOADED and _app_ref is not None:

        if macro_name == "Clear_Data":
            deleted = clear_for_restore.run(_app_ref)
            msg = (f"Cleared {len(deleted)} file(s)." if deleted
                   else "Cells cleared. No .txt files found.")
            messagebox.showinfo("Clear Data", msg)
            return

        if macro_name == "Paste_Example":
            # Step 1: clear, Step 2: restore from example
            clear_for_restore.run(_app_ref)
            restore_from_example.run(_app_ref)
            return

        if macro_name == "Load_Data":
            # Step 1: clear, Step 2: restore from saved folder
            clear_for_restore.run(_app_ref)
            restore_from_saved.run(_app_ref)
            return

        if macro_name == "CalculrateRetardationFactors":
            # In-app port of Source_Py/popups_retardation.py — no .exe, no
            # xlwings/openpyxl.  Opens a Toplevel popup, persists Koc/foc/
            # bulk density to retardation_inputs.txt, then writes computed
            # R values back into the section-5 StringVars.
            try:
                popups_retardation.run(_app_ref)
            except Exception as exc:
                messagebox.showerror(
                    "Calculate Retardation Factors",
                    f"Could not open retardation popup:\n{exc}")
                return
            # Trace recompute as a safety net (e.g. if user only edited foc)
            try:
                _app_ref._on_pfaa_change()
            except Exception:
                pass
            return

        if macro_name == "Save_Data":
            ok = generate_input_file.run(_app_ref)
            if ok:
                messagebox.showinfo("Save Data", "input.inp generated successfully.")
            return

        if macro_name == "RunPythonScript":
            # Full pipeline: build input.inp, run remfluor_v8a.exe with
            # shell redirection (< input.inp > output.out), show a runtime
            # clock, and launch the Plotly-Dash dashboard in the user's
            # default browser when the model finishes.
            try:
                run_model.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("Run Model",
                                     f"Run Model failed:\n{exc}")
            return

    # ── Legacy exe path (fallback when functions not loaded) ────────────
    if macro_name == "Load_Data":
        exe1, t1 = EXES["Load_Data_Step1"]
        exe2, t2 = EXES["Load_Data_Step2"]
        _launch_exe(exe1, _resolve_args(t1, sheet, unitflag), wait=True)
        _launch_exe(exe2, _resolve_args(t2, sheet, unitflag), wait=True)
        return

    if macro_name == "Paste_Example":
        exe1, t1 = EXES["Paste_Example_Step1"]
        exe2, t2 = EXES["Paste_Example_Step2"]
        _launch_exe(exe1, _resolve_args(t1, sheet, unitflag), wait=True)
        _launch_exe(exe2, _resolve_args(t2, sheet, unitflag), wait=True)
        return

    if macro_name == "Save_Data_Calibration":
        exe1, t1 = EXES["Save_Data_Calibration_Step1"]
        exe2, t2 = EXES["Save_Data_Calibration_Step2"]
        _launch_exe(exe1, _resolve_args(t1, sheet, unitflag), wait=True)
        _launch_exe(exe2, _resolve_args(t2, sheet, unitflag), wait=True)
        return

    if macro_name == "OpenAppendix_2_1_Relative":
        html_path = HTML_APPENDIX[macro_name]
        _open_html(html_path)
        exe, tmpl = EXES["OpenAppendix_2_1_Relative_EXE"]
        _launch_exe(exe, _resolve_args(tmpl, sheet, unitflag), wait=True)
        return

    if macro_name in HTML_APPENDIX:
        _open_html(HTML_APPENDIX[macro_name])
        return

    if macro_name in HTML_CHICKLETS:
        html_file, anchor = HTML_CHICKLETS[macro_name]
        url = _html("data_chicklets", html_file)
        if anchor:
            url = url + "#" + anchor
        _open_html(url)
        return

    if macro_name == "Show_Visualization":
        if _app_ref:
            _app_ref.show_calibration_panel()
        return

    if macro_name == "Show_MainInterface":
        if _app_ref:
            _app_ref.show_main_panel()
        return

    if macro_name == "SourceOption1":
        if _app_ref:
            first1 = _app_ref.v_src_pfaa1[0].get()
            first2 = _app_ref.v_src_pfaa2[0].get()
            for v in _app_ref.v_src_pfaa1[1:]:
                v.set(first1)
            for v in _app_ref.v_src_pfaa2[1:]:
                v.set(first2)
        return

    if macro_name == "See_Calibration_Data":
        csv_path = os.path.join(BASE_DIR, "run_history.csv")
        if os.path.exists(csv_path):
            _open_html(csv_path)
        else:
            messagebox.showinfo("Not Found",
                                f"run_history.csv not found in:\n{BASE_DIR}")
        return

    # ── Calibration helper buttons (DefaultRanges, Run_Optimal_Model,
    # Save_Optimal_Model, Help_Calibration) — pure-Python stubs.  They
    # don't have dist/popups_*.exe equivalents in the legacy codebase.
    if macro_name == "DefaultRanges":
        if _app_ref is None:
            return
        # Default literature low/high = mid * 0.5 / mid * 2 (per the
        # button caption: "Experience (default +/- x2)")
        try:
            for cv, lo, mid, hi in zip(_app_ref.v_calib_chk,
                                       _app_ref.v_calib_low,
                                       _app_ref.v_calib_mid,
                                       _app_ref.v_calib_high):
                try:
                    m = float(str(mid.get()).replace(",", ""))
                except (ValueError, TypeError):
                    continue
                lo.set(f"{m * 0.5:g}")
                hi.set(f"{m * 2.0:g}")
        except Exception as exc:
            messagebox.showerror("Default Ranges",
                                 f"Could not fill defaults:\n{exc}")
        return

    if macro_name == "Run_Optimal_Model":
        # Run the model with the current (presumed-optimal) inputs —
        # equivalent to clicking the main "Run Model" button.
        if _FUNCS_LOADED and _app_ref is not None:
            try:
                run_model.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("Run Optimal Model",
                                     f"Run failed:\n{exc}")
        return

    if macro_name == "Save_Optimal_Model":
        # Stub: save the current calibration ranges + checkbox state to
        # a sidecar text file in the project dir.  Once a real optimal-
        # model export exists, route here instead.
        if _app_ref is None:
            return
        try:
            path = os.path.join(BASE_DIR, "optimal_model.txt")
            with open(path, "w", encoding="utf-8") as fp:
                fp.write("REMFluor-MD Optimal Model snapshot\n")
                fp.write(f"Iterations: {_app_ref.v_n_iter.get()}\n\n")
                for cv, lo, mid, hi in zip(_app_ref.v_calib_chk,
                                           _app_ref.v_calib_low,
                                           _app_ref.v_calib_mid,
                                           _app_ref.v_calib_high):
                    fp.write(f"  use={cv.get()} "
                             f"lo={lo.get()} "
                             f"mid={mid.get()} "
                             f"hi={hi.get()}\n")
            messagebox.showinfo("Save Optimal Model",
                                f"Saved to:\n{path}")
        except Exception as exc:
            messagebox.showerror("Save Optimal Model",
                                 f"Save failed:\n{exc}")
        return

    if macro_name == "Help_Calibration":
        messagebox.showinfo(
            "Calibration Panel — Help",
            "REMFluor-MD MACHINE-CALIBRATION (Singh et al., 2025)\n\n"
            "Step 1) Enter monitoring well data in Section 9.\n"
            "Step 2) Tick which parameters to calibrate against.\n"
            "Step 3) Adjust per-well weighting factors (downgradient\n"
            "        wells are typically weighted higher).\n"
            "Step 4) Tick the parameters to vary, set Lowest /\n"
            "        Highest Likely Values.  Mid-Range pulls from\n"
            "        the input sheet automatically.\n"
            "Step 5) Click '2. Run Machine Based Calibration'.\n\n"
            "After the run:\n"
            "  '3. See All Calibration Data' — opens run_history.csv\n"
            "  '4. Load Optimal Data'         — pulls best-fit values\n"
            "                                  back into the inputs\n"
            "  '5. Run Optimal Model'         — re-runs the solver\n"
            "                                  with the optimal set\n"
            "  '6. Save Optimal Model'        — exports the snapshot")
        return

    if macro_name == "Authors":
        messagebox.showinfo(
            "REMFluor-MD Authors",
            "REMFluor-MD v2.6\n\n"
            "Singh et al. (2025)\n"
            "Falta et al. (2025)\n\n"
            "ESTCP – Environmental Security Technology Certification Program"
        )
        return

    # ── Pure-Python ports of the dist/ EXEs (no Excel install needed) ──
    if _FUNCS_LOADED and _app_ref is not None:
        if macro_name == "SourceRemediation":
            try:
                popups_source_remediation.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("Apply Remediation",
                                     f"Could not apply remediation:\n{exc}")
            return
        if macro_name == "ModelingTransformationLowK":
            try:
                popups_transformation.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("Modeling Transformation Low-K",
                                     f"Popup failed:\n{exc}")
            return
        if macro_name == "ChangeNumericalParameters":
            try:
                popups_numerical.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("Change Numerical Parameters",
                                     f"Popup failed:\n{exc}")
            return
        if macro_name == "OpenAppendix_2_1_Relative_EXE":
            try:
                popups_cellsize.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("Cell Size",
                                     f"Popup failed:\n{exc}")
            return
        if macro_name == "GWVelocityCalculator":
            try:
                popups_GWvelocity.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("GW Velocity Calculator",
                                     f"Popup failed:\n{exc}")
            return
        if macro_name == "LongevityTool":
            try:
                popups_longevity.run(_app_ref)
            except Exception as exc:
                import traceback
                tb = traceback.format_exc()
                print("[LongevityTool] popup failed:\n" + tb)
                messagebox.showerror("Longevity Tool",
                                     f"Popup failed:\n{exc}\n\n"
                                     f"Traceback (last lines):\n"
                                     f"{tb.strip().splitlines()[-1] if tb else ''}")
            return
        if macro_name == "SourceOption2":
            try:
                popups_mass_discharge_import.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("Source Concentrations",
                                     f"Popup failed:\n{exc}")
            return
        if macro_name == "CalibrationDataLoader":
            try:
                popups_calibration.run(_app_ref)
            except Exception as exc:
                messagebox.showerror("Calibration Data Loader",
                                     f"Popup failed:\n{exc}")
            return
        if macro_name == "HeterogeneityCalculator_Unconsolidated_Media":
            try:
                popups_heterogeneity.run(_app_ref,
                                        media_type="Unconsolidated Media")
            except Exception as exc:
                messagebox.showerror("Heterogeneity Calculator",
                                     f"Popup failed:\n{exc}")
            return
        if macro_name == "HeterogeneityCalculator_Fractured_Rock":
            try:
                popups_heterogeneity.run(_app_ref,
                                        media_type="Fractured Rock")
            except Exception as exc:
                messagebox.showerror("Heterogeneity Calculator",
                                     f"Popup failed:\n{exc}")
            return

    if macro_name in EXES:
        exe_path, tmpl = EXES[macro_name]
        wait = macro_name != "RunPythonScript"
        _launch_exe(exe_path, _resolve_args(tmpl, sheet, unitflag), wait=wait)
        return

    messagebox.showinfo(
        "Not Mapped",
        f"No action mapped for macro '{macro_name}'.\n"
        "Add it to EXES or HTML_CHICKLETS."
    )


def open_quarto(macro_name):
    run_script(macro_name)

# ─────────────────────────────────────────────────────────────────────────────
# HELPER WIDGETS
# ─────────────────────────────────────────────────────────────────────────────
class _RoundButton(tk.Canvas):
    """
    Canvas-based rounded-corner button — Excel storyboard buttons all
    use rounded rectangles with a 1-px gray outline.  tk.Button cannot
    draw rounded corners; this class does, while exposing the same
    `command` / event semantics.
    """
    def __init__(self, parent, text, command, *, fg=FG_BTN_NAVY,
                 font=None, bg=BTN_FILL, padx=10, pady=4, width=None,
                 radius=8, anchor="center"):
        import tkinter.font as _tkFont
        f = _tkFont.nametofont(font) if font else _tkFont.nametofont("TkDefaultFont")
        lines = text.split("\n")
        text_w = max((f.measure(l) for l in lines), default=20)
        text_h = f.metrics("linespace") * len(lines) + 2 * (len(lines) - 1)
        if width:
            text_w = max(text_w, int(f.measure("0") * width))
        w = text_w + padx * 2
        h = text_h + pady * 2
        super().__init__(parent, width=w, height=h, bg=BG_MAIN,
                         highlightthickness=0, bd=0, cursor="hand2")
        # NOTE: avoid overwriting tk.Misc internal attributes (_w / _h
        # are used by Tk for the widget path).  Use _cw / _ch instead.
        self._cmd  = command
        self._fg   = fg
        self._bg   = bg
        self._cw   = w
        self._ch   = h
        self._r    = radius
        self._bg_id  = self._draw_rrect(self._bg)
        self._txt_id = self.create_text(w/2, h/2, text=text, fill=fg,
                                        font=font, justify="center",
                                        anchor=anchor)
        # Click + hover bindings
        for ev, h_fn in [("<Button-1>",        self._on_press),
                         ("<ButtonRelease-1>", self._on_release),
                         ("<Enter>",           self._on_enter),
                         ("<Leave>",           self._on_leave)]:
            self.bind(ev, h_fn)

    # rounded rectangle as smoothed polygon
    def _draw_rrect(self, fill):
        x1, y1, x2, y2, r = 1, 1, self._cw-1, self._ch-1, self._r
        pts = [
            x1+r, y1, x2-r, y1, x2, y1, x2, y1+r,
            x2,   y2-r, x2,   y2, x2-r, y2, x1+r, y2,
            x1,   y2, x1,   y2-r, x1,   y1+r, x1, y1,
        ]
        return self.create_polygon(pts, smooth=True, fill=fill,
                                   outline="#777")

    def _on_enter(self, _):
        self.itemconfig(self._bg_id, fill="#C0C0C0")
    def _on_leave(self, _):
        self.itemconfig(self._bg_id, fill=self._bg)
    def _on_press(self, _):
        self.itemconfig(self._bg_id, fill="#A8A8A8")
    def _on_release(self, _):
        self.itemconfig(self._bg_id, fill=self._bg)
        if self._cmd:
            self._cmd()


def make_btn(parent, text, macro, quarto=False, fg=FG_BTN_NAVY,
             font=FONT_BTN, width=None, padx=4, pady=2, bg=BTN_FILL,
             anchor="center"):
    cmd = (lambda m=macro: open_quarto(m)) if quarto else (lambda m=macro: run_script(m))
    return _RoundButton(parent, text, cmd, fg=fg, font=font, bg=bg,
                        padx=max(padx, 8), pady=max(pady, 3),
                        width=width, radius=8, anchor=anchor)


def make_entry(parent, var, width=10, bg=BG_INPUT_BLUE, justify="right"):
    # Auto-pick readable text color: white on dark backgrounds, black on light.
    is_locked = bg.lower() in ("#000000", "#000", "black")
    fg = "#FFFFFF" if is_locked else FG_INPUT
    e = tk.Entry(parent, textvariable=var, width=width, font=FONT_INPUT,
                 bg=bg, fg=fg, relief="solid", bd=1, justify=justify,
                 insertbackground=fg,
                 disabledbackground=bg, disabledforeground=fg,
                 readonlybackground=bg)
    if is_locked:
        # Black cells = calculated values (per Excel storyboard).
        # Make them read-only so users can't accidentally overwrite.
        e.configure(state="readonly")
    return e


def make_label(parent, text, font=FONT_LABEL, fg=FG_INPUT, bg=BG_MAIN,
               anchor="w", wraplength=0):
    kw = dict(text=text, font=font, fg=fg, bg=bg, anchor=anchor)
    if wraplength:
        kw["wraplength"] = wraplength
    return tk.Label(parent, **kw)


def help_link(parent, macro, bg=BG_MAIN):
    """
    Red '?' help link inside a TRUE square box, sized to match the
    height of an adjacent tk.Entry so the box appears flush-adjacent
    to its input cell (the tkinter equivalent of the PySide6
    HelpLink button).  Win9x raised look: white top/left, gray
    bottom/right.

    Tk Labels are sized in characters x lines, which gives a tall
    rectangle for a single character.  To force a square we wrap the
    '?' Label inside a Frame with explicit pixel width=height and
    pack_propagate / grid_propagate disabled so the children can't
    push the Frame out of square.
    """
    # Compute the square side from FONT_HELP linespace so the box
    # scales with the chosen font size (and zoom). Use a slightly
    # smaller multiplier so the box hugs the entry instead of
    # towering above it.
    try:
        import tkinter.font as _tkFont
        ls = _tkFont.nametofont(FONT_HELP).metrics("linespace")
        side = max(13, int(ls * 1.05))
    except Exception:
        side = 14

    # "raised" relief on tk.Frame draws a Win9x-style bevel — white
    # top/left, dark bottom/right — which mirrors the PySide6 HelpLink
    # styling.
    box = tk.Frame(parent, bg=BTN_FILL, width=side, height=side,
                   bd=1, relief="raised", highlightthickness=0)
    box.pack_propagate(False)
    box.grid_propagate(False)

    # Prefer the Excel "?" icon (Figures/QuestionMark.png) when present.
    # Fall back to a red text glyph if Pillow / image not available.
    img = _help_qmark_image(side - 4)
    if img is not None:
        lbl = tk.Label(box, image=img, bg=BTN_FILL,
                       cursor="hand2",
                       bd=0, padx=0, pady=0,
                       takefocus=0, highlightthickness=0)
        lbl.image = img        # keep ref so Tk doesn't GC
    else:
        lbl = tk.Label(
            box, text="?",
            font=FONT_HELP,
            fg=FG_HELP,
            bg=BTN_FILL,
            cursor="hand2",
            bd=0, padx=0, pady=0,
            relief="flat",
            takefocus=0,
            highlightthickness=0,
        )
    lbl.place(relx=0.5, rely=0.5, anchor="center")

    handler = lambda e, m=macro: run_script(m)
    box.bind("<Button-1>", handler)
    lbl.bind("<Button-1>", handler)
    return box


# Cache the QuestionMark icon at multiple sizes (Tk PhotoImage refs
# must stay alive — store on a module-level dict keyed by pixel side).
_QMARK_CACHE: dict = {}

def _help_qmark_image(side: int):
    """Return Tk PhotoImage of Figures/QuestionMark.png scaled to
    `side x side`, or None if the file/Pillow is unavailable."""
    side = max(8, int(side))
    if side in _QMARK_CACHE:
        return _QMARK_CACHE[side]
    path = os.path.join(FIGURES_DIR, "QuestionMark.png")
    if not os.path.isfile(path):
        return None
    try:
        from PIL import Image, ImageTk
        img = Image.open(path).convert("RGBA")
        try:
            resample = Image.Resampling.LANCZOS
        except AttributeError:
            resample = Image.LANCZOS
        img = img.resize((side, side), resample)
        photo = ImageTk.PhotoImage(img)
    except Exception:
        try:
            photo = tk.PhotoImage(file=path)
            f = max(1, photo.height() // side)
            if f > 1:
                photo = photo.subsample(f, f)
        except Exception:
            return None
    _QMARK_CACHE[side] = photo
    return photo


def big_radio(parent, text, variable, value, *,
              text_font=None, bullet_font=None, bg=BG_MAIN):
    """Radio with a Canvas-drawn circle so the size is OS-independent.

    Windows locks tk.Radiobutton's indicator to the OS-theme size regardless
    of font, so we draw our own circle on a Canvas and wire up click/trace
    events to replicate full radio behaviour.

    RADIO_R controls the circle radius in pixels (default 9 → 18 px diameter).
    """
    RADIO_R = 9  # radius in pixels — change here to resize all circles
    if text_font is None:
        text_font = FONT_LABEL

    fr  = tk.Frame(parent, bg=bg)
    dia = RADIO_R * 2 + 4          # canvas size with a small margin
    cv  = tk.Canvas(fr, width=dia, height=dia, bg=bg,
                    highlightthickness=0, bd=0)
    cv.pack(side="left")

    lbl = tk.Label(fr, text=text, font=text_font, bg=bg, anchor="w")
    lbl.pack(side="left", padx=(3, 0))

    def _draw(*_):
        cv.delete("all")
        # Outer ring
        cv.create_oval(2, 2, RADIO_R * 2 + 2, RADIO_R * 2 + 2,
                       outline="#555555", width=1.5)
        # Inner filled dot when selected
        if variable.get() == value:
            inner = max(RADIO_R // 2, 3)
            cx = RADIO_R + 2
            cv.create_oval(cx - inner, cx - inner, cx + inner, cx + inner,
                           fill="#333333", outline="#333333")

    def _select(_event=None):
        variable.set(value)

    cv.bind("<Button-1>", _select)
    lbl.bind("<Button-1>", _select)
    variable.trace_add("write", _draw)
    _draw()  # initial render
    return fr


def section_header(parent, num, text, bg=BG_MAIN):
    return tk.Label(parent, text=f"{num}.  {text}",
                    font=FONT_SECTION, fg=FG_SECTION, bg=bg, anchor="w")


def dropdown(parent, var, choices, width=10, bg=None):
    cb = ttk.Combobox(parent, textvariable=var, values=choices, width=width,
                      font=FONT_INPUT, state="readonly")
    cb.set(var.get())
    if bg:
        # ttk widgets ignore plain `bg=`; build a per-color style instead.
        # A readonly Combobox displays its current value using the
        # SELECTION colors, so we have to override those too — otherwise
        # the field stays the OS default.
        style_name = f"C{bg.replace('#','').upper()}.TCombobox"
        s = ttk.Style()
        s.configure(style_name,
                    fieldbackground=bg, background=bg,
                    foreground=FG_INPUT, arrowcolor=FG_INPUT,
                    selectbackground=bg, selectforeground=FG_INPUT)
        s.map(style_name,
              fieldbackground=[("readonly", bg), ("active", bg),
                               ("focus", bg), ("!focus", bg)],
              background=[("readonly", bg), ("active", bg)],
              foreground=[("readonly", FG_INPUT)],
              selectbackground=[("readonly", bg), ("focus", bg),
                                ("!focus", bg)],
              selectforeground=[("readonly", FG_INPUT),
                                ("focus", FG_INPUT),
                                ("!focus", FG_INPUT)])
        cb.configure(style=style_name)
    return cb


# ─────────────────────────────────────────────────────────────────────────────
# LOGO LOADER  (looks in BASE_DIR/Figures/ for the ESTCP image)
# ─────────────────────────────────────────────────────────────────────────────
def _load_figure(filename: str, target_height: int = None, target_width: int = None):
    """
    Load BASE_DIR/Figures/<filename> and return a Tk PhotoImage (or None
    if the file isn't found / can't be decoded).

    If only one of target_height/target_width is given, the other is
    computed to preserve aspect ratio.  Prefers Pillow for resizing
    quality; falls back to native tk.PhotoImage if Pillow is missing.
    """
    if not filename:
        return None
    path = os.path.join(FIGURES_DIR, filename)
    if not os.path.isfile(path):
        return None
    # Pillow path (preferred)
    try:
        from PIL import Image, ImageTk
        img = Image.open(path)
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGBA")
        w, h = img.size
        if target_height and not target_width:
            target_width = max(1, int(w * target_height / h))
        elif target_width and not target_height:
            target_height = max(1, int(h * target_width / w))
        if target_height and target_width:
            try:
                resample = Image.Resampling.LANCZOS
            except AttributeError:
                resample = Image.LANCZOS
            img = img.resize((target_width, target_height), resample)
        return ImageTk.PhotoImage(img)
    except Exception:
        pass
    # Fallback: native Tk PhotoImage (PNG / GIF only, no smooth resize)
    try:
        return tk.PhotoImage(file=path)
    except Exception:
        return None


def _load_logo_image(target_height: int = 50):
    """
    Returns a Tk-compatible image object (PhotoImage / ImageTk.PhotoImage)
    or None if no suitable image can be found.

    Search order:
      1) Files in Figures/ whose name contains "estcp" (PNG/GIF/JPG/BMP)
      2) Files in Figures/ whose name contains "logo"
      3) Any image file in Figures/
    Prefers PIL (Pillow) for resampling/JPG support; falls back to native
    Tk PhotoImage with subsample() if PIL is unavailable.
    """
    if not os.path.isdir(FIGURES_DIR):
        return None

    exts = (".png", ".gif", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff")
    all_files = [f for f in os.listdir(FIGURES_DIR) if f.lower().endswith(exts)]

    def score(name):
        low = name.lower()
        s = 0
        if "estcp" in low: s += 100
        if "logo"  in low: s += 50
        if low.endswith(".png"): s += 5  # prefer PNG
        return s

    candidates = sorted(all_files, key=score, reverse=True)

    for fname in candidates:
        path = os.path.join(FIGURES_DIR, fname)
        # Try Pillow first (handles all formats + smooth resize)
        try:
            from PIL import Image, ImageTk
            pil_img = Image.open(path)
            if pil_img.mode not in ("RGB", "RGBA"):
                pil_img = pil_img.convert("RGBA")
            w, h = pil_img.size
            if h <= 0:
                continue
            new_w = max(1, int(w * target_height / h))
            try:
                resample = Image.Resampling.LANCZOS
            except AttributeError:
                resample = Image.LANCZOS
            pil_img = pil_img.resize((new_w, target_height), resample)
            return ImageTk.PhotoImage(pil_img)
        except Exception:
            pass
        # Fallback: native Tk PhotoImage (PNG / GIF only)
        try:
            img = tk.PhotoImage(file=path)
            h = img.height()
            if h > target_height:
                factor = max(1, h // target_height)
                img = img.subsample(factor, factor)
            return img
        except Exception:
            continue

    return None


# ─────────────────────────────────────────────────────────────────────────────
# MAIN APPLICATION
# ─────────────────────────────────────────────────────────────────────────────
class REMFluorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("REMFluor-MD Model Input Screen  v2.6")
        self.configure(bg=BG_MAIN)
        self.resizable(True, True)

        # Match Tk's internal scaling to the actual display DPI so fonts/widgets
        # render at their intended physical size. Without this, the high-DPI
        # awareness call shrinks every widget to ~50-66% of its expected size.
        self._apply_tk_dpi_scaling()

        # ttk's default Windows theme ('vista') draws Combobox fields with the
        # native OS colors and ignores `fieldbackground` style settings.
        # Switching to 'clam' makes ttk honor our color customizations so the
        # PFAA / Clay dropdowns can actually appear in Pull-Down peach.
        try:
            s = ttk.Style()
            if "clam" in s.theme_names():
                s.theme_use("clam")
        except Exception:
            pass

        # Create the Tk named fonts (lets us zoom in/out at runtime).
        self._init_named_fonts()

        # Bind zoom shortcuts: Ctrl++ / Ctrl+- / Ctrl+0 / Ctrl+wheel
        self.bind_all("<Control-plus>",        self._zoom_in)
        self.bind_all("<Control-equal>",       self._zoom_in)   # plus w/o shift
        self.bind_all("<Control-KP_Add>",      self._zoom_in)
        self.bind_all("<Control-minus>",       self._zoom_out)
        self.bind_all("<Control-KP_Subtract>", self._zoom_out)
        self.bind_all("<Control-0>",           self._zoom_reset)
        self.bind_all("<Control-MouseWheel>",  self._zoom_wheel)

        self.active_sheet = "Simple"
        self._detailed_only_frames = []  # frames hidden in Simple mode
        self._simple_only_frames   = []  # frames hidden in Detailed mode
        global _app_ref
        _app_ref = self

        self._build_vars()

        outer = tk.Frame(self, bg=BG_MAIN)
        outer.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(outer, bg=BG_MAIN, bd=0, highlightthickness=0)
        # Classic Win9x-style scrollbars via ttk + clam theme:
        # light blue thumb on a gray track with dark arrow buttons and a
        # dotted grip in the middle of the thumb — matches user's
        # reference image.
        try:
            _ssty = ttk.Style()
            if "clam" in _ssty.theme_names() and _ssty.theme_use() != "clam":
                _ssty.theme_use("clam")
            for orient in ("Vertical", "Horizontal"):
                _ssty.configure(f"Big.{orient}.TScrollbar",
                                background="#A3C1E0",       # blue thumb
                                troughcolor="#E4E8EE",      # light gray track
                                bordercolor="#7B7B7B",
                                lightcolor="#FFFFFF",
                                darkcolor="#7B7B7B",
                                arrowcolor="#000000",
                                arrowsize=22,
                                gripcount=4)
                _ssty.map(f"Big.{orient}.TScrollbar",
                          background=[("active",  "#7AA8D0"),
                                      ("pressed", "#5C8AA0")])
        except Exception:
            pass

        vsb = ttk.Scrollbar(outer, orient="vertical",
                            style="Big.Vertical.TScrollbar",
                            command=self.canvas.yview)
        hsb = ttk.Scrollbar(outer, orient="horizontal",
                            style="Big.Horizontal.TScrollbar",
                            command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.inner = tk.Frame(self.canvas, bg=BG_MAIN)
        # Use natural height (inner expands to fit all content); width is
        # synced to the visible canvas in _on_canvas_configure so the
        # vertical scrollbar shows up when content overflows the window.
        self.canvas_win = self.canvas.create_window(
            (0, 0), window=self.inner, anchor="nw")

        self.inner.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        self._build_ui()
        # Scale the initial window with DPI so it isn't a tiny postage stamp
        # on a 4K display. _dpi_scale is 1.0 at 96 DPI, 1.5 at 144 DPI, etc.
        scale = getattr(self, "_dpi_scale", 1.0)
        # Open at a comfortable size; the surrounding canvas exposes
        # vertical and horizontal scrollbars when content overflows.
        scr_w = self.winfo_screenwidth()
        scr_h = self.winfo_screenheight()
        win_w = min(int(1500 * scale), scr_w - 60)
        win_h = min(int(950  * scale), scr_h - 100)
        self.geometry(f"{win_w}x{win_h}")

    # ── Named fonts + zoom ───────────────────────────────────────────────
    def _init_named_fonts(self):
        """Create the Tk named fonts described by _FONT_DEFS."""
        import tkinter.font as tkFont
        self._tk_fonts = {}
        self._base_font_sizes = {}
        self._zoom = 1.0
        for name, (family, size, weight, slant) in _FONT_DEFS.items():
            try:
                fnt = tkFont.nametofont(name)
                fnt.config(family=family, size=size, weight=weight, slant=slant)
            except tk.TclError:
                fnt = tkFont.Font(name=name, family=family, size=size,
                                  weight=weight, slant=slant)
            self._tk_fonts[name] = fnt
            self._base_font_sizes[name] = size

    def _apply_zoom(self, factor):
        """Resize every named font to (base_size * factor)."""
        self._zoom = max(0.5, min(3.5, factor))
        for name, base in self._base_font_sizes.items():
            try:
                self._tk_fonts[name].config(
                    size=max(4, int(round(base * self._zoom))))
            except Exception:
                pass

    def _zoom_in(self,    _=None): self._apply_zoom(self._zoom * 1.10)
    def _zoom_out(self,   _=None): self._apply_zoom(self._zoom / 1.10)
    def _zoom_reset(self, _=None): self._apply_zoom(1.0)

    def _zoom_wheel(self, event):
        if event.delta > 0:
            self._zoom_in()
        else:
            self._zoom_out()
        return "break"

    # ── Section divider helper ───────────────────────────────────────────
    def _hsep(self, parent, color="#000000", thickness=1, pady=8):
        """Thin horizontal black line, used between sections in a column.
        Default pady bumped 3 → 8 so each section in the right column
        has more breathing room above/below its divider line."""
        sep = tk.Frame(parent, bg=color, height=thickness)
        sep.pack(fill="x", padx=2, pady=pady)
        return sep

    # ── DPI / scaling ─────────────────────────────────────────────────────
    def _apply_tk_dpi_scaling(self):
        """
        Scale Tk's font/widget sizes to match the display DPI.

        Tk's default 'scaling' assumes 72 DPI; on a real Windows monitor
        that's typically 96 DPI (= scaling 1.333). On a 125% scaled display
        the OS reports 120 DPI (scaling 1.667); 150% reports 144 DPI
        (scaling 2.0); 175% reports 168 DPI (scaling 2.333). We compute the
        ratio from winfo_fpixels('1i') and apply it.
        """
        try:
            dpi = self.winfo_fpixels("1i")  # pixels per inch on the display
            if dpi and dpi > 0:
                self.tk.call("tk", "scaling", dpi / 72.0)
                # Cache the scale factor so geometry calls can use it too
                self._dpi_scale = max(1.0, dpi / 96.0)
            else:
                self._dpi_scale = 1.0
        except Exception:
            self._dpi_scale = 1.0

    # ── Layout helpers ────────────────────────────────────────────────────
    def _on_frame_configure(self, _):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        # When canvas is wider than natural content width, stretch inner
        # to fill (no h-scrollbar needed). When canvas is narrower, keep
        # inner at its natural width so the horizontal scrollbar appears.
        try:
            inner_w = max(self.inner.winfo_reqwidth(), 1300)
        except Exception:
            inner_w = 1300
        self.canvas.itemconfig(self.canvas_win,
                               width=max(event.width, inner_w))

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def show_calibration_panel(self):
        run_script("RunPythonScript")  # Auto-cal: just run model (Simple version)

    def show_main_panel(self):
        self.canvas.yview_moveto(0.0)

    def _on_model_version_change(self, *_):
        val = self.v_model_version.get()
        self.active_sheet = "Simple" if val == "Simple Version" else "Detailed_2"
        is_detailed = (self.active_sheet == "Detailed_2")
        self._toggle_widgets(
            getattr(self, "_detailed_only_frames", []), show=is_detailed)
        self._toggle_widgets(
            getattr(self, "_simple_only_frames", []), show=not is_detailed)
        # Recompute retardation for new version context
        self._on_pfaa_change()
        # The Detailed-only toggle above will have re-shown the MW row's
        # extra Detailed cells (mw_e3/mw_e4) — but those must stay
        # hidden when the current Kf unit isn't mol-based.  Re-run the
        # §9 visibility logic so the final state honours BOTH rules.
        cb = getattr(self, "_on_psb_kf_unit_change", None)
        if callable(cb):
            cb()
        # §7 cells whose color differs between Simple and Detailed —
        # recolor now that active_sheet is up-to-date.
        if hasattr(self, "_apply_s7_version_colors"):
            try:
                self._apply_s7_version_colors()
            except Exception:
                pass

    def _toggle_widgets(self, widgets, show):
        """Show or hide widgets based on the manager kind tagged on
        the widget at registration time.

        Each registered widget carries a ``_toggle_kind`` attribute:
        ``"grid"`` (default) or ``"pack"``.  No runtime detection is
        attempted — Tk's introspection (``winfo_manager``, ``grid_info``)
        is unreliable after a widget has been hidden.
        """
        for w in widgets:
            kind = getattr(w, "_toggle_kind", "grid")
            try:
                if kind == "pack":
                    if show:
                        info = getattr(w, "_pack_info", None)
                        if info is not None:
                            w.pack(**info)
                    else:
                        try:
                            if not hasattr(w, "_pack_info"):
                                w._pack_info = w.pack_info()
                        except Exception:
                            pass
                        w.pack_forget()
                else:  # "grid"
                    if show:
                        w.grid()         # restores grid_remove'd state
                    else:
                        w.grid_remove()
            except Exception:
                pass

    # Default retardation factor lookup per PFAA — matches what the
    # Excel CalculrateRetardationFactors macro produces with default
    # Section 4 inputs (Clay, porosity 0.48, tortuosity 0.56).  Used as
    # a placeholder until the popups_retardation.exe pipeline is wired.
    # ── PFAA Koc / diffusion lookup (verbatim from Sheet3.cls/Sheet7.cls
    #    HandlePFAS_Generic) — source of truth for Section 5 retardation.
    #    Tuple is (Koc, molecular diffusion m²/sec or None).
    PFAA_KOC = {
        "PFOS":           (631, 3.52e-6),
        "PFOA":           (200, 5.79e-6),
        "PFHxS":          (126, 4.5e-6),
        "PFHxA":          (40,  7.8e-6),
        "PFBS":           (35,  1.1e-5),
        "PFNA":           (398, None),
        "User-Specified": (None, None),
        "None":           (None, None),
    }
    # Section 4 K26 dropdown (Clay/Silt). Tortuosity formulas pulled
    # from Sheet3.cls Worksheet_Change handler:
    #   Clay : K27=0.48, K28 = 0.77 * exp(0.04 * ln(2.2 ** -10)) ≈ 0.56
    #   Silt : K27=0.48, K28 = 0.77 * exp(0.04 * ln(1.4 ** -7))  ≈ 0.70
    LOWK_DEFAULTS = {
        "Clay": (0.48, 0.56),
        "Silt": (0.48, 0.70),
    }

    def _read_retardation_file(self):
        """Read foc/rho_b/Koc list from retardation_inputs.txt in BASE_DIR.
        Returns (rho_b, foc_t, foc_l, koc_list) or defaults if file missing."""
        import os
        ret_file = os.path.join(BASE_DIR, "retardation_inputs.txt")
        rho_b, foc_t, foc_l, koc_list = 1.7, 0.001, 0.002, []
        if not ret_file or not os.path.exists(ret_file):
            return rho_b, foc_t, foc_l, koc_list
        try:
            with open(ret_file) as f:
                lines = [l.strip() for l in f if l.strip()]
            in_trans = in_lowk = False
            for line in lines:
                if "Transmissive Zone" in line:
                    in_trans, in_lowk = True, False
                elif "Low-K Zone" in line:
                    in_trans, in_lowk = False, True
                elif in_trans:
                    if line.startswith("Soil Bulk Density"):
                        try: rho_b = float(line.split(",")[1])
                        except Exception: pass
                    elif line.startswith("foc (-)"):
                        try: foc_t = float(line.split(",")[1])
                        except Exception: pass
                    elif line.startswith("Koc (L/kg)"):
                        for p in line.split(",")[1:]:
                            try: koc_list.append(float(p.strip()))
                            except Exception: koc_list.append(None)
                elif in_lowk:
                    if line.startswith("foc (-)"):
                        try: foc_l = float(line.split(",")[1])
                        except Exception: pass
        except Exception as e:
            print(f"Warning reading retardation_inputs.txt: {e}")
        return rho_b, foc_t, foc_l, koc_list

    def _on_pfaa_change(self, *_):
        """Recompute Section 5 retardation.
        Uses retardation_inputs.txt (Koc, foc, rho_b) when present;
        falls back to PFAA_KOC table defaults otherwise.
        Formula: R = 1 + (rho_b * foc * Koc) / porosity"""
        try:
            G22 = float(self.v_porf.get())
        except (ValueError, TypeError):
            G22 = 0.0
        try:
            K27 = float(self.v_lowk_por.get())
        except (ValueError, TypeError):
            K27 = 0.0

        rho_b, foc_t, foc_l, koc_list = self._read_retardation_file()

        pairs = [
            (self.v_pfaa1, self.v_ret_trans1, self.v_ret_lowk1),
            (self.v_pfaa2, self.v_ret_trans2, self.v_ret_lowk2),
        ]
        if self.active_sheet == "Detailed_2":
            pairs += [
                (self.v_pfaa3, self.v_ret_trans3, self.v_ret_lowk3),
                (self.v_pfaa4, self.v_ret_trans4, self.v_ret_lowk4),
            ]
        for idx, (src_var, ret_t, ret_l) in enumerate(pairs):
            # Koc: prefer retardation_inputs.txt, fall back to lookup table
            koc = None
            if idx < len(koc_list) and koc_list[idx] and koc_list[idx] > 0:
                koc = koc_list[idx]
            _, diff = self.PFAA_KOC.get(src_var.get(), (None, None))
            if koc is None:
                koc_lookup, _ = self.PFAA_KOC.get(src_var.get(), (None, None))
                koc = koc_lookup

            if koc is None or koc == 0:
                ret_t.set(""); ret_l.set("")
                continue
            r_high = 1.0 + (rho_b * foc_t * koc / G22) if G22 > 0 else 1.0
            r_low  = 1.0 + (rho_b * foc_l * koc / K27) if K27 > 0 else 1.0
            ret_t.set(f"{r_high:.1f}")
            ret_l.set(f"{r_low:.1f}")
            # Molecular diffusion from PFAA-1 only
            if idx == 0 and diff is not None:
                self.v_mol_diff.set(f"{diff:.2E}")

    # ── Decimal-formatting helpers ──────────────────────────────
    @staticmethod
    def _format_decimal_var(var, n: int):
        """Reformat the StringVar's value to *n* decimal places.
        Leaves the value untouched if it can't be parsed as a float
        or if it's empty."""
        s = (var.get() or "").strip()
        if not s:
            return
        try:
            f = float(s)
        except ValueError:
            return
        var.set(f"{f:.{n}f}")

    def _bind_decimal_format(self, entry, var, n: int):
        """Reformat *var* to *n* decimals when the user leaves the
        entry (FocusOut) or presses Return."""
        cb = lambda _=None, v=var, k=n: self._format_decimal_var(v, k)
        try:
            entry.bind("<FocusOut>", cb, add="+")
            entry.bind("<Return>",   cb, add="+")
        except Exception:
            pass

    def _on_lowk_media_change(self, *_):
        """K26 → K27 (porosity) + K28 (tortuosity), then cascade to
        retardation factors."""
        por, tort = self.LOWK_DEFAULTS.get(self.v_lowk_media.get(),
                                           ("", ""))
        self.v_lowk_por.set(f"{por:.2f}" if isinstance(por, float) else str(por))
        self.v_lowk_tort.set(f"{tort:.2f}" if isinstance(tort, float) else str(tort))
        self._on_pfaa_change()

    def _on_porf_change(self, *_):
        """Trans. porosity G22 changed → recalc retardation."""
        self._on_pfaa_change()

    # Section 6 – dispersivity presets (matches popups_dispersivity.py values)
    _DISPERSIVITY_PRESETS = {
        "High":   (7.5,  0.04, 0.004),
        "Medium": (3.2,  0.04, 0.004),
        "Weak":   (1.1,  0.04, 0.004),
    }

    def _on_het_change(self, *_):
        """Heterogeneity radio changed → update longitudinal/transverse/vertical."""
        preset = self._DISPERSIVITY_PRESETS.get(self.v_het.get())
        if preset is None:
            # "Enter Your Own Value Below" — clear cells so the user
            # types fresh values.  Cell colors stay sky blue; only
            # the text content is reset.
            self.v_alpha_l.set("")
            self.v_alpha_t.set("")
            self.v_alpha_v.set("")
            return
        al, at, av = preset
        self.v_alpha_l.set(f"{al:.3f}")
        self.v_alpha_t.set(f"{at:.3f}")
        self.v_alpha_v.set(f"{av:.4f}")

    # ── XLSM read-back ─────────────────────────────────────────────
    def refresh_from_xlsm(self):
        """Pull the cells listed in XLSM_CELL_MAP from the workbook
        back into the matching Tk StringVars.  Called after every
        dist/popups_*.exe returns, so cells those EXEs wrote are
        reflected in the GUI.

        Two flavors of var name supported:
          - "v_pfaa1"           -> single StringVar attribute on self
          - "v_src_pfaa1_3"     -> 4th element of self.v_src_pfaa1 list
        """
        addrs = list(XLSM_CELL_MAP.values())
        sheet = "Detailed_2" if self.active_sheet == "Detailed_2" else "Simple"
        cells = _read_xlsm_cells(addrs, sheet_name=sheet)
        if not cells:
            return
        for var_name, addr in XLSM_CELL_MAP.items():
            if addr not in cells:
                continue
            value = cells[addr]
            # Indexed list attribute? "<name>_<idx>"
            if "_" in var_name:
                base, _, tail = var_name.rpartition("_")
                if tail.isdigit():
                    seq = getattr(self, base, None)
                    if isinstance(seq, list) and int(tail) < len(seq):
                        try:
                            seq[int(tail)].set(value)
                        except Exception:
                            pass
                        continue
            var = getattr(self, var_name, None)
            if var is None:
                continue
            try:
                var.set(value)
            except Exception:
                pass

    # ── Variable initialisation ───────────────────────────────────────────
    def _build_vars(self):
        self.v_model_version = tk.StringVar(value="Simple Version")
        self.v_model_version.trace_add('write', self._on_model_version_change)
        self.v_units         = tk.StringVar(value="meters")
        self.v_site          = tk.StringVar(value="Test Case 1")
        self.v_date          = tk.StringVar(value=datetime.now().strftime("%b-%y"))

        # Section 2
        self.v_x_size   = tk.StringVar(value="500")
        self.v_y_size   = tk.StringVar(value="50")
        self.v_z_size   = tk.StringVar(value="10")
        self.v_sw_width = tk.StringVar(value="60")
        self.v_sw_thick = tk.StringVar(value="5")
        self.v_yr_start = tk.StringVar(value="1977")
        self.v_yr_end   = tk.StringVar(value="2077")
        self.v_run_time = tk.StringVar(value="0.0")

        # Section 3
        self.v_darcy   = tk.StringVar(value="10.00")
        self.v_porf    = tk.StringVar(value="0.2")
        self.v_porf.trace_add("write",
                              lambda *_: self._on_porf_change())

        # Section 4
        self.v_lowk_media = tk.StringVar(value="Clay")
        self.v_lowk_media.trace_add("write",
                                    lambda *_: self._on_lowk_media_change())
        self.v_lowk_por   = tk.StringVar(value="0.48")
        self.v_lowk_por.trace_add("write",
                                  lambda *_: self._on_pfaa_change())
        self.v_lowk_tort  = tk.StringVar(value="0.56")

        # Section 5
        self.v_pfaa1      = tk.StringVar(value="PFOS")
        self.v_pfaa2      = tk.StringVar(value="None")
        self.v_pfaa3      = tk.StringVar(value="PFAA 1-able")  # K38 Precursor 1 (Detailed)
        self.v_pfaa4      = tk.StringVar(value="PFAA 2-able")  # M38 Precursor 2 (Detailed)
        self.v_ret_trans1 = tk.StringVar(value="2.9")
        self.v_ret_lowk1  = tk.StringVar(value="2.6")
        # PFAA-2 retardation factors (matching PFAA-1 columns)
        self.v_ret_trans2 = tk.StringVar(value="")
        self.v_ret_lowk2  = tk.StringVar(value="")
        # Precursor 1 & 2 retardation (Detailed mode only)
        self.v_ret_trans3 = tk.StringVar(value="")
        self.v_ret_lowk3  = tk.StringVar(value="")
        self.v_ret_trans4 = tk.StringVar(value="")
        self.v_ret_lowk4  = tk.StringVar(value="")
        self.v_mol_diff   = tk.StringVar(value="3.5E-10")
        # When user picks a different PFAA, recompute the retardation
        # cells from a default lookup (mirrors what the Excel
        # CalculrateRetardationFactors macro does — until we wire the
        # real EXE).
        self.v_pfaa1.trace_add("write", lambda *_: self._on_pfaa_change())
        self.v_pfaa2.trace_add("write", lambda *_: self._on_pfaa_change())
        self.v_pfaa3.trace_add("write", lambda *_: self._on_pfaa_change())
        self.v_pfaa4.trace_add("write", lambda *_: self._on_pfaa_change())

        # Section 6 – Dispersivity (top header)
        self.v_het     = tk.StringVar(value="Medium")
        self.v_alpha_l = tk.StringVar(value="3.200")
        self.v_alpha_t = tk.StringVar(value="0.040")
        self.v_alpha_v = tk.StringVar(value="0.004")
        self.v_het.trace_add("write", self._on_het_change)

        # Section 7
        self.v_src_years = [tk.StringVar(value=str(1977 + i*10)) for i in range(11)]
        self.v_src_pfaa1 = [tk.StringVar(value="1,600.000") for _ in range(11)]
        self.v_src_pfaa2 = [tk.StringVar(value="0") for _ in range(11)]
        # Detailed-only Precursor 1 / Precursor 2 year-grid source values
        # (XLSM cols Z and AB on Detailed_2 sheet)
        self.v_src_pre1  = [tk.StringVar(value="1,600.000") for _ in range(11)]
        self.v_src_pre2  = [tk.StringVar(value="0") for _ in range(11)]
        self.v_total_mass      = tk.StringVar(value="#VALUE!")
        self.v_total_mass_p2   = tk.StringVar(value="0.0")
        self.v_total_mass_pre1 = tk.StringVar(value="#VALUE!")
        self.v_total_mass_pre2 = tk.StringVar(value="0.0")

        # Section 8
        self.v_src_rem_yr   = tk.StringVar(value="")
        self.v_src_conc_red = tk.StringVar(value="")

        # Section 9
        self.v_model_psb   = tk.BooleanVar(value=False)
        self.v_psb_a_1     = tk.StringVar(value="")
        self.v_psb_a_2     = tk.StringVar(value="")
        self.v_psb_a_3     = tk.StringVar(value="")
        self.v_psb_a_4     = tk.StringVar(value="")
        self.v_psb_kf_1    = tk.StringVar(value="")
        self.v_psb_kf_2    = tk.StringVar(value="")
        self.v_psb_kf_3    = tk.StringVar(value="")
        self.v_psb_kf_4    = tk.StringVar(value="")
        self.v_psb_kf_conv  = tk.StringVar(value="")
        self.v_psb_kf_conv2 = tk.StringVar(value="")
        self.v_psb_kf_conv3 = tk.StringVar(value="")
        self.v_psb_kf_conv4 = tk.StringVar(value="")
        self.v_psb_kf_unit  = tk.StringVar(value="(mg/kg)(mg/L)^(-a)")
        # S molecular weight (g/mol) — only visible when v_psb_kf_unit is mol-based
        self.v_psb_mw_1     = tk.StringVar(value="")
        self.v_psb_mw_2     = tk.StringVar(value="")
        self.v_psb_mw_3     = tk.StringVar(value="")
        self.v_psb_mw_4     = tk.StringVar(value="")
        # §5 Detailed-only Transformation Rate + Yield Factor
        # (XLSM K41/M41 = Precursor 1/2 transformation rates,
        #  K42/M42 = yield factors)
        self.v_trans_rate_3   = tk.StringVar(value="0.380")
        self.v_trans_rate_4   = tk.StringVar(value="")
        self.v_yield_factor_3 = tk.StringVar(value="0.750")
        self.v_yield_factor_4 = tk.StringVar(value="")
        self.v_psb_yr      = tk.StringVar(value="")
        self.v_psb_width   = tk.StringVar(value="4")
        self.v_psb_load    = tk.StringVar(value="")
        self.v_psb_dist    = tk.StringVar(value="")
        self.v_psb_cells   = tk.StringVar(value="")

        # Section 10
        self.v_sample_yr = tk.StringVar(value="2025")
        self.v_mw_names  = [tk.StringVar(value=n) for n in
                            ["MW-504","FS-MW504","FS-MW505","FS-MW506",
                             "FS-MW507","FS-MW508","FS-MW509"]]
        self.v_mw_conc   = [tk.StringVar(value=str(c)) for c in
                            [2000, 1950, 1900, 1700, 1300, 750, 200]]
        self.v_mw_dist   = [tk.StringVar(value=str(d)) for d in
                            [10, 50, 100, 200, 300, 400, 500]]
        self.v_mw_conc2  = [tk.StringVar(value="") for _ in range(7)]

        # Section 11
        self.v_see_every = tk.StringVar(value="100")

        # Image cache (Tk PhotoImage refs must be held to avoid GC)
        self._figures = {}

    # ── UI builder ───────────────────────────────────────────────────────
    def _build_ui(self):
        p = self.inner

        self._build_top_bar(p)          # 1. Dark teal banner with logo

        # 2. Two-column body with a thick black vertical bar between them.
        # Header strip (Site/Date + Sec1 + Legend) lives INSIDE the left
        # column so the legend's right edge snaps to the central divider.
        body = tk.Frame(p, bg=BG_MAIN)
        body.pack(fill="both", expand=True, padx=8, pady=4)

        left    = tk.Frame(body, bg=BG_MAIN)
        # 0.05" thick black vertical divider between the two halves.
        try:
            divider_w = max(2, int(self.winfo_pixels("0.05i")))
        except Exception:
            divider_w = 5
        divider = tk.Frame(body, bg="#000000", width=divider_w)
        right   = tk.Frame(body, bg=BG_MAIN)

        # padx around the central divider — small breathing room so
        # left/right content never touches the black line.
        left.grid(   row=0, column=0, sticky="new",  padx=(0, 14))
        divider.grid(row=0, column=1, sticky="ns")
        right.grid(  row=0, column=2, sticky="new",  padx=(14, 0))
        body.columnconfigure(0, weight=1, uniform="halves")
        body.columnconfigure(1, minsize=divider_w)
        body.columnconfigure(2, weight=1, uniform="halves")

        # ── LEFT HALF ──────────────────────────────────────────────────
        # Header strip (Site/Date + Section 1 + Legend) sits at top of
        # the left column.  Legend's right edge therefore snaps to the
        # central divider.
        self._build_header_strip(left)
        self._hsep(left)
        self._build_s2_model_config(left)
        self._hsep(left)
        self._build_s3_gw_velocity(left)
        self._hsep(left)
        self._build_s4_hydrogeologic(left)
        self._hsep(left)
        self._build_s5_transport(left)
        self._hsep(left)   # closing line at the end of Section 5

        # ── RIGHT HALF ─────────────────────────────────────────────────
        # Section 6 (Heterogeneity) spans the top of the right half
        self._build_s6_heterogeneity(right)
        self._hsep(right)   # between Section 6 and Section 7/8

        # Section 7 + thick black vertical divider + Section 8
        s78 = tk.Frame(right, bg=BG_MAIN)
        s78.pack(fill="x", anchor="nw", pady=(0, 4))

        s7_col = tk.Frame(s78, bg=BG_MAIN)
        s7_col.grid(row=0, column=0, sticky="nw", padx=(0, 20))
        self._build_s7_source(s7_col)

        # Vertical bar between Section 7 and Section 8 (same thickness
        # as the body's half-divider).
        divider78 = tk.Frame(s78, bg="#000000", width=divider_w)
        divider78.grid(row=0, column=1, sticky="ns")

        s8_col = tk.Frame(s78, bg=BG_MAIN)
        s8_col.grid(row=0, column=2, sticky="nw", padx=(20, 0))
        self._build_s8_source_rem(s8_col)

        s78.columnconfigure(1, minsize=divider_w)
        s78.columnconfigure(2, weight=1)

        # Sections 9-11 stacked beneath, each filling the full half-width.
        self._hsep(right)   # between Section 7/8 row and Section 9
        self._build_s9_psb(right)
        self._build_s10_field_data(right)
        self._hsep(right)   # between Section 10 and Section 11
        self._build_s11_output(right)

        # Action cluster is now built inline inside _build_s11_output
        # (see the blue panel there) — no separate _build_action_row call.

        # ── Mid horizontal black bar ─────────────────────────────────
        midbar = tk.Frame(p, bg="#000000", height=6)
        midbar.pack(fill="x", padx=0, pady=(2, 2))

        # ── Bottom-body 2-col: empty | calibration panel ─────────────
        bot = tk.Frame(p, bg=BG_MAIN)
        bot.pack(fill="both", expand=True, padx=8, pady=4)

        bot_left  = tk.Frame(bot, bg=BG_MAIN)
        bot_div   = tk.Frame(bot, bg="#000000", width=divider_w)
        bot_right = tk.Frame(bot, bg=BG_MAIN)

        bot_left .grid(row=0, column=0, sticky="new",  padx=(0, 8))
        bot_div  .grid(row=0, column=1, sticky="ns")
        bot_right.grid(row=0, column=2, sticky="new",  padx=(8, 0))
        bot.columnconfigure(0, weight=1, uniform="halves")
        bot.columnconfigure(1, minsize=divider_w)
        bot.columnconfigure(2, weight=1, uniform="halves")

        self._build_calibration_panel(bot_right)

    # ─────────────────────────────────────────────────────────────────────
    # 1.  TOP DARK-TEAL BANNER  (#074F69)
    # ─────────────────────────────────────────────────────────────────────
    def _build_top_bar(self, parent):
        bar = tk.Frame(parent, bg=BG_HEADER_BAR, pady=6)
        bar.pack(fill="x")

        # Title styled per the Excel storyboard reference: "REMFluor-MD"
        # (the product name) is greenish (#DAF2D0); " Model Input Screen"
        # and "Vers. 2.6" are white.  All three labels use the bold
        # AppTitle / AppVersion fonts.
        title_fr = tk.Frame(bar, bg=BG_HEADER_BAR)
        title_fr.pack(side="left", padx=18)
        tk.Label(title_fr, text="REMFluor-MD",
                 font=FONT_TITLE, fg=FG_TITLE_GREEN, bg=BG_HEADER_BAR
                 ).pack(side="left")
        tk.Label(title_fr, text=" Model Input Screen",
                 font=FONT_TITLE, fg=FG_TITLE, bg=BG_HEADER_BAR
                 ).pack(side="left")

        tk.Label(bar, text="Vers. 2.6",
                 font=FONT_VERSION, fg=FG_TITLE, bg=BG_HEADER_BAR
                 ).pack(side="left", padx=10)

        # ── Zoom hint + buttons (between version and logo) ─────────────
        zoom_fr = tk.Frame(bar, bg=BG_HEADER_BAR)
        zoom_fr.pack(side="right", padx=12)
        tk.Label(zoom_fr, text="Zoom:",
                 font=FONT_LABEL_SM, fg=FG_TITLE, bg=BG_HEADER_BAR
                 ).pack(side="left", padx=(0, 4))
        for txt, cmd in [("−", self._zoom_out),
                         ("⟲", self._zoom_reset),
                         ("+", self._zoom_in)]:
            tk.Button(zoom_fr, text=txt, command=cmd,
                      font=FONT_BTN_CALIB, fg=FG_BTN_NAVY, bg=BTN_FILL,
                      width=2, padx=2, pady=0, bd=1,
                      relief="raised", cursor="hand2"
                      ).pack(side="left", padx=1)
        tk.Label(zoom_fr, text="(Ctrl±wheel)",
                 font=FONT_LABEL_XS, fg=FG_TITLE_GREEN, bg=BG_HEADER_BAR
                 ).pack(side="left", padx=(4, 0))

        # ── ESTCP logo from Figures/ ───────────────────────────────────
        # Render at the display's native pixel resolution so the logo stays
        # crisp on high-DPI screens. target_height is in pixels, not points.
        scale = getattr(self, "_dpi_scale", 1.0)
        self._logo_img = _load_logo_image(target_height=int(54 * scale))
        if self._logo_img is not None:
            tk.Label(bar, image=self._logo_img,
                     bg=BG_HEADER_BAR, bd=0).pack(side="right", padx=18)
        else:
            # Fallback "ESTCP" text badge if no image was found
            badge = tk.Label(bar, text="  ESTCP  ",
                             font=("Arial Black", 14), fg="#0F8B4C",
                             bg="#FFFFFF", padx=10, pady=2,
                             bd=2, relief="raised")
            badge.pack(side="right", padx=18)

    # ─────────────────────────────────────────────────────────────────────
    # 2.  HEADER STRIP – Site/ID + Date  (Legend & Heterogeneity moved into
    #                                    the body columns)
    # ─────────────────────────────────────────────────────────────────────
    def _build_header_strip(self, parent):
        # Top band of upper-left quadrant — TWO columns per Excel:
        #   LEFT  : Site/Date row + Section 1 STARTING INFO row stacked
        #   RIGHT : Legend block (rows 1-5)
        strip = tk.Frame(parent, bg=BG_MAIN, pady=4)
        strip.pack(fill="x", padx=8, pady=(2, 4))

        # ── LEFT column: Site/Date on top, Section 1 below ───────────
        left_col = tk.Frame(strip, bg=BG_MAIN)
        left_col.grid(row=0, column=0, sticky="nw")

        # --- Site / Date row ---
        sd = tk.Frame(left_col, bg=BG_MAIN)
        sd.pack(anchor="w", fill="x")
        tk.Label(sd, text="Site Location and ID:",
                 font=FONT_LABEL, bg=BG_MAIN).grid(row=0, column=0, sticky="w")
        tk.Label(sd, text="Date:", font=FONT_LABEL, bg=BG_MAIN
                 ).grid(row=0, column=1, sticky="w", padx=(14, 0))
        make_entry(sd, self.v_site, width=22, justify="left").grid(
            row=1, column=0, sticky="w", pady=2)
        make_entry(sd, self.v_date, width=10, justify="left").grid(
            row=1, column=1, sticky="w", padx=(14, 0), pady=2)

        # --- Section 1 STARTING INFO row (header + bordered radios) ---
        s1 = tk.Frame(left_col, bg=BG_MAIN)
        s1.pack(anchor="w", fill="x", pady=(8, 0))
        tk.Label(s1, text="1.  STARTING INFORMATION",
                 font=FONT_SECTION, fg=FG_SECTION, bg=BG_MAIN, anchor="w"
                 ).grid(row=0, column=0, sticky="w", columnspan=2)
        tk.Label(s1, text="Units?", font=FONT_LABEL_BI, bg=BG_MAIN
                 ).grid(row=0, column=2, sticky="w", padx=(40, 0))

        # Bordered Simple/Detailed radios block (Excel rectangle) —
        # thick black border per Excel reference.
        ver_box = tk.Frame(s1, bg=BG_MAIN, bd=2, relief="solid",
                           highlightbackground="#000000",
                           highlightthickness=0,
                           padx=8, pady=3)
        ver_box.grid(row=1, column=0, sticky="w", pady=2)
        big_radio(ver_box, "Simple Version",
                  self.v_model_version, "Simple Version").pack(anchor="w")
        big_radio(ver_box, "Detailed Version",
                  self.v_model_version, "Detailed Version").pack(anchor="w")
        help_link(s1, "OpenTable1").grid(row=1, column=1,
                                         sticky="w", padx=(2, 0))

        # Bordered feet/meters block — thick black border per Excel.
        u_box = tk.Frame(s1, bg=BG_MAIN, bd=2, relief="solid",
                         highlightbackground="#000000",
                         highlightthickness=0,
                         padx=8, pady=3)
        u_box.grid(row=1, column=2, sticky="w", padx=(40, 0), pady=2)
        big_radio(u_box, "feet",   self.v_units, "feet").pack(anchor="w")
        big_radio(u_box, "meters", self.v_units, "meters").pack(anchor="w")
        help_link(s1, "OpenTable1").grid(row=1, column=3,
                                         sticky="w", padx=(2, 0))

        # ── Legend block — left edge aligned with the central black
        #    divider that separates upper-left from upper-right halves.
        #    Spacer column 1 expands so legend snaps to the far right
        #    of the upper-left half.
        strip.columnconfigure(1, weight=1)
        legend_holder = tk.Frame(strip, bg=BG_MAIN)
        legend_holder.grid(row=0, column=2, sticky="ne")
        self._build_legend(legend_holder)

    # ─────────────────────────────────────────────────────────────────────
    # LEGEND BLOCK  (top-right of upper-left quadrant, per Excel storyboard)
    # Layout matches the Excel reference:
    #   Row 0: "Legend" title  ............................  ☐ Check Box
    #   Row 1: [white  ] Enter value directly             ..  [Button]
    #   Row 2: [blue   ] Cell with formula or default, but ok to overwrite**
    #   Row 3: [peach  ] Pull Down Menu
    #   Row 4: [black  ] Calculated value or taken from other cell.
    # ─────────────────────────────────────────────────────────────────────
    def _build_legend(self, parent):
        outer = tk.Frame(parent, bg=BG_LEGEND_BAR, bd=1, relief="solid")
        outer.pack(anchor="nw")

        # ── Title bar (dark teal, white italic) ──────────────────────
        title = tk.Frame(outer, bg=BG_HEADER_BAR, padx=4, pady=1)
        title.pack(fill="x")
        tk.Label(title, text="Legend",
                 font=FONT_LABEL_BI, bg=BG_HEADER_BAR, fg=FG_TITLE
                 ).pack(side="left", padx=(2, 12))
        tk.Label(title, text="☑", font=FONT_LABEL_B,
                 bg=BG_HEADER_BAR, fg="#7FCB7B"
                 ).pack(side="left")
        tk.Label(title, text="Check Box", font=FONT_LABEL_I,
                 bg=BG_HEADER_BAR, fg=FG_TITLE
                 ).pack(side="left", padx=(2, 12))
        tk.Label(title, text="Button", font=FONT_LABEL_BI,
                 bg=BG_HEADER_BAR, fg="#9CC3E8"
                 ).pack(side="left", padx=(2, 4))

        # ── Body (white) — 4 swatch rows ─────────────────────────────
        body = tk.Frame(outer, bg=BG_LEGEND_BAR, padx=4, pady=2)
        body.pack(fill="x")

        legend_items = [
            (BG_INPUT_BLUE, "",             "Enter value directly"),
            (BG_FORMULA,    "",             "Cell with formula or default, but ok to overwrite**"),
            (BG_PULLDOWN,   "Pull Down",    "Pull Down Menu"),
            (BG_LOCKED,     "Can't change", "Calculated value or taken from other cell."),
        ]
        for i, (color, swatch_text, descr) in enumerate(legend_items):
            fg_swatch = "#FFFFFF" if color == "#000000" else FG_INPUT
            sw = tk.Label(body, text=swatch_text,
                          font=FONT_LABEL_BI, bg=color, fg=fg_swatch,
                          bd=1, relief="solid", width=12, anchor="center")
            sw.grid(row=i, column=0, sticky="w", padx=2, pady=1)
            tk.Label(body, text=descr, font=FONT_LABEL_I,
                     bg=BG_LEGEND_BAR, fg=FG_INPUT, anchor="w"
                     ).grid(row=i, column=1, sticky="w", padx=6)

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 6 – GEOLOGIC HETEROGENEITY  (right column, top)
    # ─────────────────────────────────────────────────────────────────────
    def _build_s6_heterogeneity(self, parent):
        section_header(parent, "6",
                       "PLUME TRANSPORT – DISPERSIVITY"
                       ).pack(anchor="w", pady=(2, 1))

        # All controls live in a single grid:
        #   row 0  – heterogeneity radios + "Enter Your Own Value Below"
        #   row 1  – column titles (Longitudinal / Transverse / Vertical)
        #   row 2  – "Values:" label + the three cells
        #
        # Column layout:
        #   col 0   – row labels  ("Geologic Heterogeneity is:" / "Values:")
        #   col 1-3 – content (radios on row 0, titles on row 1, cells on row 2)
        #   col 4   – "Enter Your Own Value Below" hint  (row 0 only)
        #   col 5   – How to Decide?  (rowspan=3 → vertically centered)
        #   col 6   – Section5_1.png  (rowspan=3 → vertically centered)
        form = tk.Frame(parent, bg=BG_MAIN)
        form.pack(fill="x", pady=(0, 4))

        # ── Row 0 – Heterogeneity radios ───────────────────────────
        tk.Label(form, text="Geologic Heterogeneity is:",
                 font=FONT_LABEL, bg=BG_MAIN
                 ).grid(row=0, column=0, sticky="e", padx=(0, 4))

        het_fr = tk.Frame(form, bg=BG_MAIN)
        het_fr.grid(row=0, column=1, columnspan=4, sticky="w")
        for val in ["High", "Medium", "Weak", "Enter Your Own Value Below"]:
            big_radio(het_fr, val, self.v_het, val
                      ).pack(side="left", padx=4)

        # ── Row 1 – Titles ABOVE each cell ─────────────────────────
        titles = ["Longitudinal (m)", "Transverse (m)", "Vertical (m)"]
        for i, lbl in enumerate(titles):
            tk.Label(form, text=lbl, font=FONT_LABEL_SM, bg=BG_MAIN
                     ).grid(row=1, column=1+i, sticky="w", padx=(8, 4))

        # ── Row 2 – Values label + entry cells (start at same X as
        #             the heterogeneity radios — both in column 1). ──
        tk.Label(form, text="Values:", font=FONT_LABEL, bg=BG_MAIN
                 ).grid(row=2, column=0, sticky="e", padx=(0, 4))
        for i, var in enumerate([self.v_alpha_l,
                                  self.v_alpha_t,
                                  self.v_alpha_v]):
            make_entry(form, var, width=8, bg=BG_FORMULA
                       ).grid(row=2, column=1+i, sticky="w", padx=(8, 4))

        # ── How to Decide? button — vertically centered between
        #     heterogeneity row and values row (rowspan=3 with no
        #     sticky lets the grid auto-center it). ──────────────
        make_btn(form, "How to Decide?",
                 "OpenAppendix_6_1_Relative", quarto=True,
                 font=FONT_BTN_SM, width=14, bg=BTN_FILL
                 ).grid(row=0, column=5, rowspan=3, padx=(12, 0))

        # ── Section5_1.png next to the How to Decide button ───────
        scale = getattr(self, "_dpi_scale", 1.0)
        img = _load_figure("Section5_1.png", target_height=int(80 * scale))
        if img is None:
            # Fallback: previous filename in case the file is still
            # named Section6_1.png in the Figures/ folder.
            img = _load_figure("Section6_1.png",
                               target_height=int(80 * scale))
        self._figures["s5_1"] = img
        if img is not None:
            tk.Label(form, image=img, bg=BG_MAIN, bd=0
                     ).grid(row=0, column=6, rowspan=3, padx=(8, 0))

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 1 – STARTING INFORMATION
    # ─────────────────────────────────────────────────────────────────────
    def _build_s1_starting(self, parent):
        # One grid frame holds everything so the "Units" title can sit on
        # row 0 — the same vertical line as "STARTING INFORMATION".
        #
        # Column map:
        #   0 : "STARTING INFORMATION" header
        #   1 : 1-inch spacer (forced via columnconfigure minsize)
        #   2 : Simple / Detailed radios          (rows 1-2)
        #   3 : '?' for versions  (rowspan=2  -> vertically centered)
        #   4 : 2-inch spacer (forced via columnconfigure minsize)
        #   5 : "Units" title  (row 0)  +  feet / meters  (rows 1-2)
        #   6 : '?' for Units    (rowspan=2  -> vertically centered)
        f = tk.Frame(parent, bg=BG_MAIN)
        f.pack(fill="x", pady=(2, 4))

        # Force the spacer columns to be exactly 0.5" / 0.5" wide.
        f.columnconfigure(1, minsize="0.5i")
        f.columnconfigure(4, minsize="0.5i")

        # Row 0 – titles on the same horizontal line
        tk.Label(f, text="1.  STARTING INFORMATION",
                 font=FONT_SECTION, fg=FG_SECTION, bg=BG_MAIN, anchor="w"
                 ).grid(row=0, column=0, sticky="w")
        tk.Label(f, text="Units", font=FONT_LABEL_BI, bg=BG_MAIN
                 ).grid(row=0, column=5, sticky="w")

        # Rows 1-2 – Simple / Detailed radios in col 2
        big_radio(f, "Simple Version",
                  self.v_model_version, "Simple Version"
                  ).grid(row=1, column=2, sticky="w")
        big_radio(f, "Detailed Version",
                  self.v_model_version, "Detailed Version"
                  ).grid(row=2, column=2, sticky="w")

        # '?' for versions – col 3, rowspan covers feet/meters rows
        help_link(f, "OpenTable1").grid(
            row=1, column=3, rowspan=2, padx=(2, 0))

        # Rows 1-2 – feet / meters in col 5 (same column as the Units title)
        big_radio(f, "feet",   self.v_units, "feet"
                  ).grid(row=1, column=5, sticky="w")
        big_radio(f, "meters", self.v_units, "meters"
                  ).grid(row=2, column=5, sticky="w")

        # '?' for Units – col 6, rowspan covers feet/meters rows
        help_link(f, "OpenTable1").grid(
            row=1, column=6, rowspan=2, padx=(2, 0))

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 2 – MODEL CONFIGURATION
    # ─────────────────────────────────────────────────────────────────────
    def _build_s2_model_config(self, parent):
        section_header(parent, "2", "MODEL CONFIGURATION").pack(
            anchor="w", pady=(4, 1))

        container = tk.Frame(parent, bg=BG_MAIN)
        container.pack(fill="x", pady=(0, 4))

        # ── LEFT: just the form. Labels on a SINGLE line (no wraplength)
        #         so the form takes its natural full width and pushes the
        #         images further to the right. ─────────────────────────
        left_col = tk.Frame(container, bg=BG_MAIN)
        left_col.pack(side="left", anchor="n")

        f = tk.Frame(left_col, bg=BG_MAIN)
        f.pack(fill="x", pady=(0, 4))

        # "Model Size" sub-label sits on row 0, in the column where the
        # value cells live (column 1) — so it labels the input column.
        tk.Label(f, text="Model Size", font=FONT_LABEL_I,
                 bg=BG_MAIN, fg=FG_GREY
                 ).grid(row=0, column=1, sticky="w", padx=2)

        rows = [
            ("Model Size in Direction of Groundwater Flow (X Direction)", self.v_x_size, "(m)",        "OpenTable2_1_XDirection"),
            ("Model Width Perpendicular to Flow (Y Direction)",           self.v_y_size, "(m)",        "OpenTable2_1_YDirection"),
            ("Model Depth Below Water Table (Z Direction)",               self.v_z_size, "(m)",        "OpenTable2_1_ZDirection"),
            ("Source Width (REMChlor-MD will round to nearest whole cell)", self.v_sw_width, "(m)",    "OpenTable2_3"),
            ("Thickness of Source Below Water Table",                     self.v_sw_thick, "(m)",      "OpenTable2_4"),
            ("Starting Year of Simulation (year the source started)",     self.v_yr_start, "(YYYY year)", "OpenTable2_5"),
            ("Ending Year of Simulation",                                 self.v_yr_end,   "(YYYY year)", "OpenTable2_6"),
        ]
        for i, (lbl, var, unit, helpm) in enumerate(rows):
            # Extra top-padding (~3× previous) to visually separate the
            # three logical groups (model dims | source | sim years):
            #   i = 3 → above "Source Width"   (after Model Depth)
            #   i = 5 → above "Starting Year"  (after Thickness)
            pady_top = 30 if i in (3, 5) else 1
            pady     = (pady_top, 1)

            # Right-align the descriptive label so all entries line up.
            tk.Label(f, text=lbl, font=FONT_LABEL, bg=BG_MAIN, anchor="e"
                     ).grid(row=i+1, column=0, sticky="e",
                            pady=pady, padx=(0, 4))
            make_entry(f, var, width=8).grid(
                row=i+1, column=1, sticky="w", pady=pady)
            tk.Label(f, text=unit, font=FONT_LABEL_SM, bg=BG_MAIN
                     ).grid(row=i+1, column=2, sticky="w",
                            padx=2, pady=pady)
            help_link(f, helpm).grid(row=i+1, column=3, sticky="w",
                                      pady=pady)

        # ── RIGHT: image area + controls, all positioned with place()
        #          so each control can sit RIGHT of the image it
        #          relates to (Optional next to Section2_1, run-time +
        #          Estimating next to Section2_2). Because Section2_2
        #          is narrower than Section2_1, run-time / Estimating
        #          start further LEFT than Optional. ────────────────
        right_area = tk.Frame(container, bg=BG_MAIN)
        right_area.pack(side="left", anchor="n", padx=(14, 0))

        scale = getattr(self, "_dpi_scale", 1.0)
        img_h = int(110 * scale)
        try:
            overlap_px = int(self.winfo_pixels("0.5i"))
            gap_px     = int(self.winfo_pixels("0.15i"))
            shift_px   = int(self.winfo_pixels("0.2i"))   # vertical nudge
        except Exception:
            overlap_px = int(48 * scale)
            gap_px     = int(14 * scale)
            shift_px   = int(19 * scale)

        self._figures["s2_1"] = _load_figure("Section2_1.png", target_height=img_h)
        self._figures["s2_2"] = _load_figure("Section2_2.png", target_height=img_h)

        img1 = self._figures.get("s2_1")
        img2 = self._figures.get("s2_2")
        w1 = img1.width()  if img1 else 200
        h1 = img1.height() if img1 else img_h
        w2 = img2.width()  if img2 else 150
        h2 = img2.height() if img2 else img_h

        # Container needs explicit width / height for place() to work.
        # Reserve ~380 px on the right of the wider image for buttons —
        # the rounded "Optional: Enter user defined size of grid cells"
        # button is wider than the previous flat tk.Button.
        ctrl_w = 380
        box_w  = max(w1, w2) + gap_px + ctrl_w
        box_h  = h1 + h2 - overlap_px          # vertical overlap of 0.5"

        img_box = tk.Frame(right_area, bg=BG_MAIN, width=box_w, height=box_h)
        img_box.pack(anchor="nw")
        img_box.pack_propagate(False)

        # Section2_1 at top-left
        if img1 is not None:
            tk.Label(img_box, image=img1, bg=BG_MAIN, bd=0
                     ).place(x=0, y=0)

        # Section2_2 below Section2_1, overlapping by 0.5", drawn on top
        if img2 is not None:
            lbl2 = tk.Label(img_box, image=img2, bg=BG_MAIN, bd=0)
            lbl2.place(x=0, y=h1 - overlap_px)
            lbl2.lift()

        # ── Optional button: top-right, next to Section2_1 ───────────
        # Starts at x = w1 + gap (right edge of Section2_1 + gap).
        opt_btn = make_btn(img_box,
                           "Optional: Enter user defined\nsize of grid cells",
                           "OpenAppendix_2_1_Relative", quarto=True,
                           font=FONT_BTN_SM, width=24, bg=BTN_FILL)
        opt_btn.place(x=w1 + gap_px, y=8 + shift_px)   # 0.2" lower

        # ── approx run time: next to Section2_2.  Because Section2_2
        #    is narrower, this starts further LEFT than the Optional
        #    button does (x = w2 + gap). ──────────────────────────
        rt = tk.Frame(img_box, bg=BG_MAIN)
        tk.Label(rt, text="approx. run time:", font=FONT_LABEL, bg=BG_MAIN,
                 fg=FG_GREY).pack(side="left")
        tk.Label(rt, textvariable=self.v_run_time, font=FONT_LABEL,
                 bg=BG_LOCKED, fg=FG_LOCKED, width=14, relief="solid", bd=1
                 ).pack(side="left", padx=4)
        tk.Label(rt, text="minutes", font=FONT_LABEL_SM, bg=BG_MAIN,
                 fg=FG_GREY).pack(side="left")
        # 0.2" lower than its previous position
        rt_y = h1 - overlap_px + 4 + shift_px
        rt.place(x=w2 + gap_px, y=rt_y)

        # ── Estimating the Source Start Time: under run-time, also at
        #    x = w2 + gap (next to Section2_2). Single line, no wrap.
        #    0.2" gap between the run-time label and this button.
        est_btn = make_btn(img_box, "Estimating the Source Start Time",
                           "OpenAppendix_2_2_Relative", quarto=True,
                           font=FONT_BTN_SM, width=30, bg=BTN_FILL)
        est_btn.place(x=w2 + gap_px, y=rt_y + 22 + shift_px)

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 3 – GW DARCY VELOCITY
    # ─────────────────────────────────────────────────────────────────────
    def _build_s3_gw_velocity(self, parent):
        section_header(parent, "3", "GROUNDWATER DARCY VELOCITY (Vd) (m/yr)").pack(
            anchor="w", pady=(6, 1))

        f = tk.Frame(parent, bg=BG_MAIN)
        f.pack(fill="x", pady=(0, 4))

        # ── Row 0: short titles ON TOP of the cells, with '?' next ──
        # Col 0: empty (long Vd description sits here on row 1)
        tk.Label(f, text="Vd (m/yr)",
                 font=FONT_LABEL, bg=BG_MAIN, anchor="w"
                 ).grid(row=0, column=1, sticky="w")
        help_link(f, "OpenTable3_1").grid(row=0, column=2,
                                          sticky="w", padx=(2, 0))

        tk.Label(f, text="Trans. Effective Porosity (-)",
                 font=FONT_LABEL, bg=BG_MAIN, anchor="w"
                 ).grid(row=0, column=4, sticky="w", padx=(20, 0))
        help_link(f, "OpenTable3_2").grid(row=0, column=5,
                                          sticky="w", padx=(2, 0))

        # ── Row 1: long Vd description (LEFT of cell), Vd cell,
        #          Trans cell, OR, Calculator button ───────────────
        tk.Label(f, text="Bulk Groundwater Darcy Velocity (Vd)  (Vd = K·dh/dx)",
                 font=FONT_LABEL, bg=BG_MAIN, anchor="e"
                 ).grid(row=1, column=0, sticky="e", padx=(0, 6),
                        pady=(2, 0))

        # Vd cell sized to roughly match the "Vd (m/yr)" title above it
        make_entry(f, self.v_darcy, width=12).grid(
            row=1, column=1, columnspan=2, sticky="w", pady=(2, 0))

        # Trans. Effective Porosity cell — doubled width
        make_entry(f, self.v_porf, width=20).grid(
            row=1, column=4, columnspan=2, sticky="w",
            padx=(20, 0), pady=(2, 0))

        # OR between Trans. Effective Porosity cell and Calculator button
        tk.Label(f, text="OR", font=FONT_LABEL_BI,
                 bg=BG_MAIN, fg=FG_GREY
                 ).grid(row=1, column=6, padx=8, pady=(2, 0))

        # Groundwater Velocity Calculator button (right of OR)
        make_btn(f, "Groundwater Velocity Calculator",
                 "GWVelocityCalculator", font=FONT_BTN_SM, width=30,
                 bg=BTN_FILL).grid(row=1, column=7, padx=(0, 0),
                                    pady=(2, 0))

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 4 – HYDROGEOLOGIC / MATRIX DIFFUSION
    # ─────────────────────────────────────────────────────────────────────
    def _build_s4_hydrogeologic(self, parent):
        # Whole section indented 0.25" from the left edge.
        s4 = tk.Frame(parent, bg=BG_MAIN)
        s4.pack(fill="x", padx=("0.25i", 0), pady=(0, 4))

        section_header(s4, "4",
                       "HYDROGEOLOGIC SETTING AND MATRIX DIFFUSION"
                       ).pack(anchor="w", pady=(2, 1))

        f = tk.Frame(s4, bg=BG_MAIN)
        f.pack(fill="x", pady=(0, 4))

        # ── LEFT: heterogeneity calculator buttons, vertically
        #          centered against the image's height. The two
        #          weight=1 spacer rows (above & below the buttons)
        #          push the button pair to the middle of bcol. ────
        bcol = tk.Frame(f, bg=BG_MAIN)
        bcol.grid(row=0, column=0, sticky="ns")
        bcol.rowconfigure(0, weight=1)   # top spacer
        bcol.rowconfigure(3, weight=1)   # bottom spacer

        make_btn(bcol, "Use Unconsolidated Media\nHeterogeneity Calculator",
                 "HeterogeneityCalculator_Unconsolidated_Media",
                 font=FONT_BTN_SM, width=30, bg=BTN_FILL, padx=4, pady=4
                 ).grid(row=1, column=0, sticky="w")
        # 0.1" gap between the two "Use" buttons
        make_btn(bcol, "Use Fractured Rock\nHeterogeneity Calculator",
                 "HeterogeneityCalculator_Fractured_Rock",
                 font=FONT_BTN_SM, width=30, bg=BTN_FILL, padx=4, pady=4
                 ).grid(row=2, column=0, sticky="w", pady=("0.1i", 0))

        # ── MIDDLE: Section4_1.png with 1.5" gaps on either side
        #            (6× the previous 0.25" — separating it from the
        #            buttons on the left and the Low-k Media block on
        #            the right). ──────────────────────────────────
        mcol = tk.Frame(f, bg=BG_MAIN)
        mcol.grid(row=0, column=1, sticky="n",
                  padx=("1.5i", "1.5i"))

        scale = getattr(self, "_dpi_scale", 1.0)
        self._figures["s4_1"] = _load_figure("Section4_1.png",
                                             target_height=int(140 * scale))
        if self._figures["s4_1"] is not None:
            tk.Label(mcol, image=self._figures["s4_1"],
                     bg=BG_MAIN, bd=0).pack(anchor="n")

        # ── RIGHT: Low-k Media Details + "What is a Low-k" button ──
        rcol = tk.Frame(f, bg=BG_MAIN)
        rcol.grid(row=0, column=2, sticky="nw")

        tk.Label(rcol, text="Low-k Media Details",
                 font=FONT_LABEL_BI,
                 bg=BG_SECTION_HDR, anchor="w", padx=4
                 ).grid(row=0, column=0, columnspan=4, sticky="ew", pady=(0, 2))

        rows = [
            ("Low-k Zone Media",          self.v_lowk_media, None,  "OpenTable4_1"),
            ("Low-k Zone Total Porosity", self.v_lowk_por,   "(-)", "OpenTable4_2"),
            ("Low-k Zone Pore Tortuosity",self.v_lowk_tort,  "(-)", "OpenTable4_3"),
        ]
        for i, (lbl, var, unit, helpm) in enumerate(rows):
            tk.Label(rcol, text=lbl, font=FONT_LABEL, bg=BG_MAIN
                     ).grid(row=i+1, column=0, sticky="w", pady=1)
            if i == 0:
                # Excel data-validation for K26: Clay or Silt only.
                dropdown(rcol, var,
                         ["Clay", "Silt"],
                         width=10, bg=BG_PULLDOWN
                         ).grid(row=i+1, column=1, padx=4, sticky="w")
            else:
                # Total Porosity / Pore Tortuosity – Formula legend color
                e_s4 = make_entry(rcol, var, width=8, bg=BG_FORMULA)
                e_s4.grid(row=i+1, column=1, padx=4, sticky="w")
                # Section 4: enforce 2-decimal display on focus-out / Return.
                self._bind_decimal_format(e_s4, var, 2)
            if unit:
                tk.Label(rcol, text=unit, font=FONT_LABEL_SM, bg=BG_MAIN
                         ).grid(row=i+1, column=2, sticky="w")
            help_link(rcol, helpm).grid(row=i+1, column=3, sticky="w", padx=2)

        # 0.1" gap between Low-k Zone Pore Tortuosity (row 3) and the
        # "What is a Low-k Geologic Unit?" button (row 4).
        make_btn(rcol, 'What is a "Low-k" Geologic Unit?',
                 "OpenAppendix_4_2_Relative", quarto=True,
                 font=FONT_BTN_SM, width=32, bg=BTN_FILL).grid(
            row=4, column=0, columnspan=4, sticky="w", pady=("0.1i", 0))

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 5 – PFAS TRANSPORT PROPERTIES
    # ─────────────────────────────────────────────────────────────────────
    def _build_s5_transport(self, parent):
        section_header(parent, "5", "PFAS TRANSPORT PROPERTIES").pack(
            anchor="w", pady=(4, 1))

        # Centered form — labels right-aligned in column 0, entries
        # share columns 1 + 2.  Per Excel storyboard the whole block is
        # centered horizontally, not flush-left.
        f = tk.Frame(parent, bg=BG_MAIN)
        f.pack(anchor="center", pady=(0, 4))

        # Excel data-validation lists for E38 / G38 (PFAA-1 / PFAA-2)
        PFAA1_CHOICES = ["PFOS", "PFOA", "PFHxS", "PFHxA", "PFBS", "PFNA", "User-Specified"]
        PFAA2_CHOICES = ["None"] + PFAA1_CHOICES

        # ── Column headers ────────────────────────────────────────────────
        tk.Label(f, text="", bg=BG_MAIN).grid(row=0, column=0)
        tk.Label(f, text="PFAA 1", font=FONT_LABEL_B,
                 bg=BG_MAIN).grid(row=0, column=1, padx=4)
        tk.Label(f, text="PFAA 2", font=FONT_LABEL_B,
                 bg=BG_MAIN).grid(row=0, column=2, padx=4)
        # Detailed-only headers
        lbl_pre1 = tk.Label(f, text="Precursor 1", font=FONT_LABEL_B, bg=BG_MAIN)
        lbl_pre1.grid(row=0, column=3, padx=4)
        lbl_pre2 = tk.Label(f, text="Precursor 2", font=FONT_LABEL_B, bg=BG_MAIN)
        lbl_pre2.grid(row=0, column=4, padx=4)
        self._detailed_only_frames += [lbl_pre1, lbl_pre2]

        # ── PFAA dropdowns ─────────────────────────────────────────────
        tk.Label(f, text="PFAA (use dropdown menu)",
                 font=FONT_LABEL, bg=BG_MAIN, anchor="e"
                 ).grid(row=1, column=0, sticky="e", padx=(0, 6))
        dropdown(f, self.v_pfaa1, PFAA1_CHOICES, width=8, bg=BG_PULLDOWN
                 ).grid(row=1, column=1, padx=4)
        dropdown(f, self.v_pfaa2, PFAA2_CHOICES, width=8, bg=BG_PULLDOWN
                 ).grid(row=1, column=2, padx=4)
        PRE_CHOICES = ["PFAA 1-able", "PFAA 2-able"] + PFAA1_CHOICES
        dd3 = dropdown(f, self.v_pfaa3, PRE_CHOICES, width=8, bg=BG_PULLDOWN)
        dd3.grid(row=1, column=3, padx=4)
        dd4 = dropdown(f, self.v_pfaa4, PRE_CHOICES, width=8, bg=BG_PULLDOWN)
        dd4.grid(row=1, column=4, padx=4)
        self._detailed_only_frames += [dd3, dd4]
        help_link(f, "OpenTable5_1").grid(row=1, column=6)

        # ── Retardation rows ───────────────────────────────────────────
        ret_rows = [
            ("Retardation Factor for Transmissive Zones",
             self.v_ret_trans1, self.v_ret_trans2,
             self.v_ret_trans3, self.v_ret_trans4, "(-)", "OpenTable5_3"),
            ("Retardation Factor for Low-k Units",
             self.v_ret_lowk1,  self.v_ret_lowk2,
             self.v_ret_lowk3,  self.v_ret_lowk4,  "(-)", "OpenTable5_4"),
        ]
        for i, (lbl, v1, v2, v3, v4, unit, helpm) in enumerate(ret_rows):
            tk.Label(f, text=lbl, font=FONT_LABEL, bg=BG_MAIN, anchor="e"
                     ).grid(row=2+i, column=0, sticky="e", pady=1, padx=(0, 6))
            e1 = make_entry(f, v1, width=8, bg=BG_LOCKED)
            e1.grid(row=2+i, column=1, padx=4)
            e2 = make_entry(f, v2, width=8, bg=BG_LOCKED)
            e2.grid(row=2+i, column=2, padx=4)
            e3 = make_entry(f, v3, width=8, bg=BG_LOCKED)
            e3.grid(row=2+i, column=3, padx=4)
            e4 = make_entry(f, v4, width=8, bg=BG_LOCKED)
            e4.grid(row=2+i, column=4, padx=4)
            # Section 5 (black retardation cells): enforce 1-decimal display.
            for _e, _v in ((e1, v1), (e2, v2), (e3, v3), (e4, v4)):
                self._bind_decimal_format(_e, _v, 1)
            self._detailed_only_frames += [e3, e4]
            tk.Label(f, text=unit, font=FONT_LABEL_SM, bg=BG_MAIN
                     ).grid(row=2+i, column=5, sticky="w")
            help_link(f, helpm).grid(row=2+i, column=6, padx=2)

        # Start in correct state for current version
        if self.active_sheet != "Detailed_2":
            for w in [lbl_pre1, lbl_pre2, dd3, dd4]:
                w.grid_remove()
            # entry widgets for pre cols already added above, remove them too
            for w in self._detailed_only_frames[-4:]:
                w.grid_remove()

        # ── Detailed-only: Transformation Rate + Yield Factor rows ─
        # XLSM J41/K41/M41/N41 (Trans. Rate) and J42/K42/M42/N42 (Yield).
        # White cells with red text in Precursor 1/2 columns.
        for r_off, lbl, v3, v4, unit in [
            (4, "Transformation Rate",
             self.v_trans_rate_3, self.v_trans_rate_4, "(years)"),
            (5, "Yield Factor",
             self.v_yield_factor_3, self.v_yield_factor_4, "(-)"),
        ]:
            tr_lbl = tk.Label(f, text=lbl, font=FONT_LABEL,
                              bg=BG_MAIN, anchor="e")
            tr_lbl.grid(row=r_off, column=0, sticky="e",
                        pady=1, padx=(0, 6))
            # Empty placeholder cells in PFAA1/PFAA2 columns
            tr_e3 = tk.Entry(f, textvariable=v3, width=8, font=FONT_INPUT,
                             bg="#FFFFFF", fg="#C00000",
                             relief="solid", bd=1, justify="right")
            tr_e3.grid(row=r_off, column=3, padx=4)
            tr_e4 = tk.Entry(f, textvariable=v4, width=8, font=FONT_INPUT,
                             bg="#FFFFFF", fg="#C00000",
                             relief="solid", bd=1, justify="right")
            tr_e4.grid(row=r_off, column=4, padx=4)
            tr_unit = tk.Label(f, text=unit, font=FONT_LABEL_SM,
                               bg=BG_MAIN, fg=FG_GREY)
            tr_unit.grid(row=r_off, column=5, sticky="w")
            self._detailed_only_frames += [tr_lbl, tr_e3, tr_e4, tr_unit]
            # Hide on init if currently Simple
            if self.active_sheet != "Detailed_2":
                for w in (tr_lbl, tr_e3, tr_e4, tr_unit):
                    try: w.grid_remove()
                    except Exception: pass

        # "Calculate Retardation Factors" + "Modeling Transformation Low K"
        # buttons live in the same row.  Use grid layout inside cb so the
        # detailed-only Modeling Transformation button can be hidden via
        # grid_remove (consistent with the existing toggle).
        cb = tk.Frame(parent, bg=BG_MAIN); cb.pack(anchor="center", pady=(3, 3))
        make_btn(cb, "Calculate Retardation Factors",
                 "CalculrateRetardationFactors",
                 font=FONT_BTN_SM, width=24, bg=BTN_FILL
                 ).grid(row=0, column=0, padx=2)
        help_link(cb, "OpenTable5_2").grid(row=0, column=1, padx=4)
        mt_btn = make_btn(cb, "Modeling Transformation Low K",
                          "ModelingTransformationLowK",
                          font=FONT_BTN_SM, width=28, bg=BTN_FILL)
        mt_btn.grid(row=0, column=2, padx=(12, 2))
        self._detailed_only_frames.append(mt_btn)
        if self.active_sheet != "Detailed_2":
            try: mt_btn.grid_remove()
            except Exception: pass

        # General Molecular Diffusion Coefficient on its own row — centered
        df = tk.Frame(parent, bg=BG_MAIN)
        df.pack(anchor="center", pady=(2, 4))
        tk.Label(df, text="General molecular Diffusion Coefficient for all Constituents",
                 font=FONT_LABEL, bg=BG_MAIN, anchor="e"
                 ).grid(row=0, column=0, sticky="e", padx=(0, 6))
        make_entry(df, self.v_mol_diff, width=10, bg=BG_FORMULA).grid(
            row=0, column=1, padx=4)
        tk.Label(df, text="(m²/sec)", font=FONT_LABEL_SM, bg=BG_MAIN
                 ).grid(row=0, column=2, sticky="w")
        help_link(df, "OpenTable5_7").grid(row=0, column=3)

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 7 – PFAS SOURCE TERM (middle column)
    #
    # XLSM-authoritative layout (Detailed_2 sheet):
    #   col Q   – Option 1:/2:/3: labels + button frames
    #   col T   – "Go to Section 2 to Change Years" rotated label
    #             (BOTH versions; spans years 1977…2077)
    #   col U   – Years (1977…2077)
    #   col V   – PFAA 1 values   (header V6, species V7=E38)
    #   col X   – PFAA 2 values   (header X6, species X7=G38)
    #   col Y   – "(ug/L)" units / "(kg)" on totals row
    #   col Z   – Precursors 1 values  (DETAILED ONLY, header Z6, Z7=K38)
    #   col AB  – Precursors 2 values  (DETAILED ONLY, header AB6, AB7=M38)
    #   row 19  – Total PFAS Mass Out of Source per column
    # ─────────────────────────────────────────────────────────────────────
    def _build_s7_source(self, parent):
        section_header(parent, "7", "PFAS SOURCE TERM").pack(
            anchor="w", pady=(2, 1))

        f = tk.Frame(parent, bg=BG_MAIN)
        f.pack(fill="x", pady=(0, 4))

        PFAA1_CHOICES = ["PFOS", "PFOA", "PFHxS", "PFHxA", "PFBS", "PFNA",
                         "User-Specified"]
        PFAA2_CHOICES = ["None"] + PFAA1_CHOICES
        PRE_CHOICES   = ["PFAA 1-able", "PFAA 2-able"] + PFAA1_CHOICES

        # Track Detailed-only widgets locally for reliable hide-on-init.
        det_only = []

        def _mark_detailed(*widgets):
            for w in widgets:
                det_only.append(w)
                self._detailed_only_frames.append(w)

        # ── Row 0 – Column headers (italic gray text).  Cell bg is
        #    version-specific: gray (BG_MAIN) for Simple, white
        #    (BG_INPUT_BLUE) for Detailed — see _apply_s7_version_colors.
        hdr_pfaa1 = tk.Label(f, text="PFAA 1", font=FONT_LABEL_BI,
                             bg=BG_MAIN, fg=FG_GREY)
        hdr_pfaa1.grid(row=0, column=3, padx=4)
        hdr_pfaa2 = tk.Label(f, text="PFAA 2", font=FONT_LABEL_BI,
                             bg=BG_MAIN, fg=FG_GREY)
        hdr_pfaa2.grid(row=0, column=4, padx=4)
        hdr_pre1 = tk.Label(f, text="Precursors 1", font=FONT_LABEL_BI,
                            bg=BG_MAIN, fg=FG_GREY)
        hdr_pre1.grid(row=0, column=6, padx=4)
        hdr_pre2 = tk.Label(f, text="Precursors 2", font=FONT_LABEL_BI,
                            bg=BG_MAIN, fg=FG_GREY)
        hdr_pre2.grid(row=0, column=7, padx=4)
        _mark_detailed(hdr_pre1, hdr_pre2)
        self._s7_hdr_labels = [hdr_pfaa1, hdr_pfaa2, hdr_pre1, hdr_pre2]

        # ── Row 1 – "How to Get Source Concentration?" + dropdowns ────
        make_btn(f, "How to Get Source Concentration?",
                 "OpenAppendix_7_1_Relative", quarto=True,
                 font=FONT_BTN_SM, width=28, bg=BTN_FILL
                 ).grid(row=1, column=0, sticky="w",
                        padx=(0, 8), pady=(0, 4))
        # §7 species cells are display-only (XLSM V7=E38, X7=G38, Z7=K38,
        # AB7=M38).  Use locked black labels bound via textvariable so
        # they auto-update when the §5 dropdowns change.
        tk.Label(f, textvariable=self.v_pfaa1, font=FONT_LABEL_B,
                 bg=BG_LOCKED, fg="#FFFFFF", width=8,
                 relief="solid", bd=1
                 ).grid(row=1, column=3, padx=4, sticky="ew")
        tk.Label(f, textvariable=self.v_pfaa2, font=FONT_LABEL_B,
                 bg=BG_LOCKED, fg="#FFFFFF", width=8,
                 relief="solid", bd=1
                 ).grid(row=1, column=4, padx=4, sticky="ew")
        help_link(f, "OpenTable5_1").grid(row=1, column=5, padx=(2, 0))
        dd_pre1 = tk.Label(f, textvariable=self.v_pfaa3, font=FONT_LABEL_B,
                           bg=BG_LOCKED, fg="#FFFFFF", width=8,
                           relief="solid", bd=1)
        dd_pre1.grid(row=1, column=6, padx=4, sticky="ew")
        dd_pre2 = tk.Label(f, textvariable=self.v_pfaa4, font=FONT_LABEL_B,
                           bg=BG_LOCKED, fg="#FFFFFF", width=8,
                           relief="solid", bd=1)
        dd_pre2.grid(row=1, column=7, padx=4, sticky="ew")
        _mark_detailed(dd_pre1, dd_pre2)

        # ── Col 1 rowspan – rotated "Go to Section 2 to Change Years" ─
        # XLSM cell T8 — present in BOTH Simple and Detailed sheets.
        # Bigger font + wider canvas for legibility; spans only the 11
        # year rows (rows 2..12), leaving the totals row 13 untouched.
        rot_h = 28 * 11
        rot = tk.Canvas(f, width=26, height=rot_h,
                        bg=BG_MAIN, highlightthickness=0)
        rot.create_text(13, rot_h // 2,
                        text="Go to Section 2 to Change Years",
                        font=("Calibri", 11, "italic"),
                        fill=FG_BTN_NAVY, angle=90)
        rot.grid(row=2, column=1, rowspan=11, sticky="ns",
                 padx=(0, 6), pady=(0, 0))

        # ── Col 0 – Option 1/2/3 stacked frames ───────────────────────
        # Option 1: rowspan=4 covers years 1977…2007 (i 0..3)
        opt1 = tk.Frame(f, bg=BG_MAIN)
        tk.Label(opt1, text="Option 1:", font=FONT_LABEL_I, bg=BG_MAIN
                 ).pack(anchor="w")
        make_btn(opt1, "Assume Constant Source\n(populates same values\n"
                       "from the first row)",
                 "SourceOption1", font=FONT_BTN_SM, width=24, bg=BTN_FILL,
                 padx=4, pady=2
                 ).pack(anchor="w", pady=(2, 0))
        opt1.grid(row=2, column=0, rowspan=4, padx=(0, 8), sticky="n")

        # Option 2: rowspan=5 covers years 2017…2057 (i 4..8) — DETAILED
        opt2 = tk.Frame(f, bg=BG_MAIN)
        tk.Label(opt2, text="Option 2:", font=FONT_LABEL_I, bg=BG_MAIN
                 ).pack(anchor="w")
        make_btn(opt2, "Upload PFAS\nConcentrations\nfrom PFAS-LEACH",
                 "SourceOption2", font=FONT_BTN_SM, width=24,
                 bg=BTN_FILL, padx=4, pady=2
                 ).pack(anchor="w", pady=(2, 0))
        opt2.grid(row=6, column=0, rowspan=5, padx=(0, 8), sticky="n")
        _mark_detailed(opt2)

        # Option 3: rowspan=2 covers years 2067…2077 (i 9..10) — DETAILED
        opt3 = tk.Frame(f, bg=BG_MAIN)
        tk.Label(opt3, text="Option 3: Enter\nconcs. for each year.",
                 font=FONT_LABEL_I, bg=BG_MAIN, justify="left"
                 ).pack(anchor="w")
        opt3.grid(row=11, column=0, rowspan=2, padx=(0, 8), sticky="n")
        _mark_detailed(opt3)

        # ── Year × value rows (i = 0..10  →  grid rows 2..12) ──────────
        # Cell colors:
        #   * Year column            → SKY BLUE in BOTH versions
        #   * PFAA 1/2 conc columns  → version-specific (Simple WHITE,
        #                               Detailed SKY BLUE)
        #   * Precursors 1/2 columns → SKY BLUE (Detailed-only)
        self._s7_conc12_cells = []
        for i in range(11):
            r = 2 + i
            make_entry(f, self.v_src_years[i], width=8,
                       bg=BG_FORMULA, justify="right"
                       ).grid(row=r, column=2, padx=2, pady=1)
            e_p1 = make_entry(f, self.v_src_pfaa1[i], width=10,
                              bg=BG_INPUT_BLUE)
            e_p1.grid(row=r, column=3, padx=2, pady=1)
            e_p2 = make_entry(f, self.v_src_pfaa2[i], width=10,
                              bg=BG_INPUT_BLUE)
            e_p2.grid(row=r, column=4, padx=2, pady=1)
            self._s7_conc12_cells.extend([e_p1, e_p2])
            tk.Label(f, text="(ug/L)", font=FONT_LABEL_SMI,
                     bg=BG_MAIN, fg=FG_GREY
                     ).grid(row=r, column=5, sticky="w", padx=2)
            e_pre1 = make_entry(f, self.v_src_pre1[i], width=10,
                                bg=BG_FORMULA)
            e_pre1.grid(row=r, column=6, padx=2, pady=1)
            e_pre2 = make_entry(f, self.v_src_pre2[i], width=10,
                                bg=BG_FORMULA)
            e_pre2.grid(row=r, column=7, padx=2, pady=1)
            _mark_detailed(e_pre1, e_pre2)

        # ── Row 13 – "Total PFAS Mass Out of Source:" totals ───────────
        tk.Label(f, text="Total PFAS Mass Out of Source:",
                 font=FONT_LABEL, bg=BG_MAIN, fg=FG_BTN_NAVY,
                 anchor="e"
                 ).grid(row=13, column=0, columnspan=3, sticky="e",
                        padx=(0, 4), pady=(6, 0))
        tk.Label(f, textvariable=self.v_total_mass, font=FONT_LABEL,
                 bg=BG_LOCKED, fg=FG_LOCKED, width=8,
                 relief="solid", bd=1
                 ).grid(row=13, column=3, padx=2, pady=(6, 0))
        tk.Label(f, textvariable=self.v_total_mass_p2, font=FONT_LABEL,
                 bg=BG_LOCKED, fg=FG_LOCKED, width=8,
                 relief="solid", bd=1
                 ).grid(row=13, column=4, padx=2, pady=(6, 0))
        tk.Label(f, text="(kg)", font=FONT_LABEL_SM,
                 bg=BG_MAIN, fg=FG_GREY
                 ).grid(row=13, column=5, sticky="w", padx=2,
                        pady=(6, 0))
        tot_pre1 = tk.Label(f, textvariable=self.v_total_mass_pre1,
                            font=FONT_LABEL, bg=BG_LOCKED, fg=FG_LOCKED,
                            width=8, relief="solid", bd=1)
        tot_pre1.grid(row=13, column=6, padx=2, pady=(6, 0))
        tot_pre2 = tk.Label(f, textvariable=self.v_total_mass_pre2,
                            font=FONT_LABEL, bg=BG_LOCKED, fg=FG_LOCKED,
                            width=8, relief="solid", bd=1)
        tot_pre2.grid(row=13, column=7, padx=2, pady=(6, 0))
        _mark_detailed(tot_pre1, tot_pre2)

        # ── Auto-compute "Total PFAS Mass Out of Source:" per column ──
        # XLSM V19 formula (Detailed):
        #   Total_kg = 1e-6 * Vd * thickness * width * Σ(C_i * Δt_i)
        # where:
        #   Vd        = §3 Bulk Groundwater Darcy Velocity (v_darcy, m/yr)
        #   thickness = §2 Thickness of Source Below WT  (v_sw_thick, m)
        #   width     = §2 Source Width                  (v_sw_width, m)
        #   Δt_i      = year[i+1] - year[i]  for i=0..9
        #               end_year - year[10]  for i=10  (end_year = §2 v_yr_end)
        #   C_i       = column-specific concentration   (ug/L)
        # Verified against XLSM Detailed_2 sample: 1e-6 * 0.011992 * 5 * 60
        # * 112000 = 0.4029 kg, matches V19.
        def _to_float_pt(var, default=None):
            try:
                s = (var.get() or "").replace(",", "").strip()
                return float(s) if s else default
            except (ValueError, AttributeError):
                return default

        def _fmt_kg(v):
            if v is None:
                return "#VALUE!"
            if v == 0:
                return "0.0"
            return f"{v:.4f}"

        def _recompute_total_mass(*_):
            vd    = _to_float_pt(self.v_darcy)
            thick = _to_float_pt(self.v_sw_thick)
            width = _to_float_pt(self.v_sw_width)
            yr_e  = _to_float_pt(self.v_yr_end)
            years = [_to_float_pt(v) for v in self.v_src_years]
            if (vd is None or thick is None or width is None or yr_e is None
                    or any(y is None for y in years)):
                # Missing geometry/year input → mirror Excel's #VALUE! result
                self.v_total_mass.set("#VALUE!")
                self.v_total_mass_p2.set("#VALUE!")
                self.v_total_mass_pre1.set("#VALUE!")
                self.v_total_mass_pre2.set("#VALUE!")
                return
            # Δt: 10 forward differences + (end_year - last_year) for index 10
            dt = [years[i+1] - years[i] for i in range(10)]
            dt.append(yr_e - years[10])
            base = 1e-6 * vd * thick * width

            def _col_total(varlist):
                tot = 0.0
                for i, v in enumerate(varlist):
                    c = _to_float_pt(v)
                    if c is None:
                        return None
                    tot += c * dt[i]
                return base * tot

            self.v_total_mass.set(_fmt_kg(_col_total(self.v_src_pfaa1)))
            self.v_total_mass_p2.set(_fmt_kg(_col_total(self.v_src_pfaa2)))
            self.v_total_mass_pre1.set(_fmt_kg(_col_total(self.v_src_pre1)))
            self.v_total_mass_pre2.set(_fmt_kg(_col_total(self.v_src_pre2)))

        # Wire the trace on every input the formula reads.
        _trace_vars = [self.v_darcy, self.v_sw_thick, self.v_sw_width,
                       self.v_yr_end]
        _trace_vars += list(self.v_src_years)
        _trace_vars += list(self.v_src_pfaa1) + list(self.v_src_pfaa2)
        _trace_vars += list(self.v_src_pre1)  + list(self.v_src_pre2)
        for _v in _trace_vars:
            _v.trace_add("write", _recompute_total_mass)
        self._recompute_total_mass = _recompute_total_mass
        _recompute_total_mass()  # initial value

        # Apply the Simple/Detailed-specific colors to the §7 widgets
        # whose bg differs between versions (Row 0 header labels +
        # PFAA 1/2 conc cells).  Re-runs from _on_model_version_change.
        self._apply_s7_version_colors()

        # In Simple mode, hide every Detailed-only widget registered above.
        if self.active_sheet != "Detailed_2":
            for w in det_only:
                try:
                    w.grid_remove()
                except Exception:
                    pass

    # ─────────────────────────────────────────────────────────────────────
    # §7 cell-color helper — keeps Simple/Detailed visual differences
    # ─────────────────────────────────────────────────────────────────────
    def _apply_s7_version_colors(self):
        """Set §7 cells whose color depends on Simple vs Detailed.

        Simple   → Row 0 hdr bg = BG_MAIN  ; PFAA 1/2 conc bg = WHITE
        Detailed → Row 0 hdr bg = WHITE    ; PFAA 1/2 conc bg = SKY BLUE
        """
        is_det = (getattr(self, "active_sheet", "Simple") == "Detailed_2")
        hdr_bg    = BG_INPUT_BLUE if is_det else BG_MAIN
        conc12_bg = BG_FORMULA    if is_det else BG_INPUT_BLUE
        for w in getattr(self, "_s7_hdr_labels", []):
            try:
                w.configure(bg=hdr_bg)
            except Exception:
                pass
        for w in getattr(self, "_s7_conc12_cells", []):
            try:
                w.configure(bg=conc12_bg, readonlybackground=conc12_bg,
                            disabledbackground=conc12_bg)
            except Exception:
                pass

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 8 – SOURCE REMEDIATION
    # ─────────────────────────────────────────────────────────────────────
    def _build_s8_source_rem(self, parent):
        section_header(parent, "8", "SOURCE REMEDIATION").pack(
            anchor="w", pady=(2, 1))

        f = tk.Frame(parent, bg=BG_MAIN)
        f.pack(fill="x", pady=(0, 4))

        tk.Label(f, text="Source Treatment Start Year",
                 font=FONT_LABEL, bg=BG_MAIN).grid(row=0, column=0, sticky="w")
        make_entry(f, self.v_src_rem_yr, width=8).grid(row=0, column=1, padx=4)
        tk.Label(f, text="(YYYY)", font=FONT_LABEL_SM, bg=BG_MAIN
                 ).grid(row=0, column=2, sticky="w")
        help_link(f, "OpenTable8_2").grid(row=0, column=3)

        tk.Label(f, text="Source Concentration Reduction",
                 font=FONT_LABEL, bg=BG_MAIN).grid(row=1, column=0, sticky="w", pady=2)
        make_entry(f, self.v_src_conc_red, width=8).grid(row=1, column=1, padx=4)
        tk.Label(f, text="(%)", font=FONT_LABEL_SM, bg=BG_MAIN
                 ).grid(row=1, column=2, sticky="w")
        help_link(f, "OpenTable8_1").grid(row=1, column=3)

        bf = tk.Frame(parent, bg=BG_MAIN)
        bf.pack(fill="x", pady=(2, 4))
        make_btn(bf, "Apply Remediation", "SourceRemediation",
                 font=FONT_BTN_SM, width=18, bg=BTN_FILL).pack(side="left", padx=(0, 4))
        make_btn(bf, "HELP", "OpenAppendix_8_1_Relative", quarto=True,
                 font=FONT_BTN_SM, width=8, bg=BTN_FILL).pack(side="left")

        # ── Section8_1.png below the buttons ─────────────────────────
        scale = getattr(self, "_dpi_scale", 1.0)
        self._figures["s8_1"] = _load_figure("Section8_1.png",
                                             target_height=int(120 * scale))
        if self._figures["s8_1"] is not None:
            tk.Label(parent, image=self._figures["s8_1"],
                     bg=BG_MAIN, bd=0).pack(anchor="w", pady=(2, 4))

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 9 – PSB
    # ─────────────────────────────────────────────────────────────────────
    def _build_s9_psb(self, parent):
        section_header(parent, "9",
                       "PLUME REMEDIATION: INSTALL PERMEABLE SORPTION BARRIER (PSB)"
                       ).pack(anchor="w", pady=(2, 1))

        det_only = []
        def _mark_detailed(*ws):
            for w in ws:
                w._toggle_kind = "grid"
                det_only.append(w)
                self._detailed_only_frames.append(w)

        # ═══ TOP — leftcol(form + buttons) | sep | PSB Dist | rcluster ═══
        top = tk.Frame(parent, bg=BG_MAIN)
        top.pack(fill="x", pady=(0, 2))

        # ── LEFT column: Freundlich grid + the two action buttons
        #    stacked underneath it (so the buttons sit flush against the
        #    bottom of the form, no empty gap between the form cell and
        #    the button row).
        leftcol = tk.Frame(top, bg=BG_MAIN)
        leftcol.pack(side="left", anchor="nw")
        f = tk.Frame(leftcol, bg=BG_MAIN)
        f.pack(anchor="nw")

        # Headers row 0
        tk.Checkbutton(f, text="Model PSB?", variable=self.v_model_psb,
                       font=FONT_LABEL, bg=BG_MAIN,
                       activebackground=BG_MAIN
                       ).grid(row=0, column=0, sticky="w", padx=(0, 4))
        tk.Label(f, text="Unit", font=FONT_LABEL_B, bg=BG_PULLDOWN,
                 relief="solid", bd=1, padx=4
                 ).grid(row=0, column=1, padx=2, sticky="ew")
        tk.Label(f, textvariable=self.v_pfaa1, font=FONT_LABEL_B,
                 bg=BG_LOCKED, fg="#FFFFFF",
                 relief="solid", bd=1, width=8
                 ).grid(row=0, column=2, padx=2, sticky="ew")
        tk.Label(f, textvariable=self.v_pfaa2, font=FONT_LABEL_B,
                 bg=BG_LOCKED, fg="#FFFFFF",
                 relief="solid", bd=1, width=8
                 ).grid(row=0, column=3, padx=2, sticky="ew")
        h_pre1 = tk.Label(f, textvariable=self.v_pfaa3,
                          font=FONT_LABEL_B, bg=BG_LOCKED, fg="#FFFFFF",
                          relief="solid", bd=1, width=10)
        h_pre1.grid(row=0, column=4, padx=2, sticky="ew")
        h_pre2 = tk.Label(f, textvariable=self.v_pfaa4,
                          font=FONT_LABEL_B, bg=BG_LOCKED, fg="#FFFFFF",
                          relief="solid", bd=1, width=10)
        h_pre2.grid(row=0, column=5, padx=2, sticky="ew")
        _mark_detailed(h_pre1, h_pre2)

        # Row 1: Freundlich "a" — Unit cell GREY
        tk.Label(f, text='PSB\'s Freundlich "a"', font=FONT_LABEL,
                 bg=BG_MAIN, anchor="e"
                 ).grid(row=1, column=0, sticky="e", pady=1, padx=(0, 4))
        tk.Label(f, text="-", font=FONT_LABEL_SM, bg=BTN_FILL,
                 fg=FG_INPUT, relief="solid", bd=1, padx=2
                 ).grid(row=1, column=1, padx=2, sticky="ew")
        # Sky-blue (BG_FORMULA) data cells per Excel storyboard — these
        # rows are "formula or default, but ok to overwrite" cells.
        for c, var in enumerate([self.v_psb_a_1, self.v_psb_a_2]):
            make_entry(f, var, width=8, bg=BG_FORMULA
                       ).grid(row=1, column=2 + c, padx=2)
        e_a3 = make_entry(f, self.v_psb_a_3, width=8, bg=BG_FORMULA)
        e_a3.grid(row=1, column=4, padx=2)
        e_a4 = make_entry(f, self.v_psb_a_4, width=8, bg=BG_FORMULA)
        e_a4.grid(row=1, column=5, padx=2)
        _mark_detailed(e_a3, e_a4)
        help_link(f, "OpenTable9_1").grid(row=1, column=6, padx=(2, 0))

        # Row 2: Freundlich Kf — Unit cell DROPDOWN (6 options from XLSM)
        tk.Label(f, text="PSB's Freundlich Kf", font=FONT_LABEL,
                 bg=BG_MAIN, anchor="e"
                 ).grid(row=2, column=0, sticky="e", pady=1, padx=(0, 4))
        KF_UNITS = [
            "(ng/kg)(ng/L)^(-a)",
            "(ug/kg)(ug/L)^(-a)",
            "(mg/kg)(mg/L)^(-a)",
            "(nmol/kg)(nmol/L)^(-a)",
            "(umol/kg)(umol/L)^(-a)",
            "(mmol/kg)(mmol/L)^(-a)",
        ]
        dropdown(f, self.v_psb_kf_unit, KF_UNITS,
                 width=20, bg=BG_PULLDOWN
                 ).grid(row=2, column=1, padx=2, sticky="ew")
        # Sky-blue (BG_FORMULA) data cells per Excel storyboard.
        for c, var in enumerate([self.v_psb_kf_1, self.v_psb_kf_2]):
            make_entry(f, var, width=8, bg=BG_FORMULA
                       ).grid(row=2, column=2 + c, padx=2)
        e_kf3 = make_entry(f, self.v_psb_kf_3, width=8, bg=BG_FORMULA)
        e_kf3.grid(row=2, column=4, padx=2)
        e_kf4 = make_entry(f, self.v_psb_kf_4, width=8, bg=BG_FORMULA)
        e_kf4.grid(row=2, column=5, padx=2)
        _mark_detailed(e_kf3, e_kf4)
        help_link(f, "OpenTable9_2").grid(row=2, column=6, padx=(2, 0))

        # Row 3: PFAS molecular weight (mol/g) — only visible when the
        # Kf unit selected on row 2 contains "mol".  Layout matches the
        # XLSM (U25 = "PFAS molecular weight (mol/g)" inside a BLACK
        # locked cell; V25/X25/Z25/AB25 are SKY-BLUE EDITABLE input
        # cells the user types the molecular weight into).
        mw_unit = tk.Label(f, text="PFAS molecular weight (mol/g)",
                           font=FONT_LABEL_SM,
                           bg=BG_LOCKED, fg="#FFFFFF",
                           relief="solid", bd=1, padx=2)
        mw_unit.grid(row=3, column=1, padx=2, sticky="ew")
        # Editable sky-blue (BG_FORMULA) cells — user types MW in g/mol.
        mw_e1 = make_entry(f, self.v_psb_mw_1, width=8, bg=BG_FORMULA)
        mw_e1.grid(row=3, column=2, padx=2)
        mw_e2 = make_entry(f, self.v_psb_mw_2, width=8, bg=BG_FORMULA)
        mw_e2.grid(row=3, column=3, padx=2)
        mw_e3 = make_entry(f, self.v_psb_mw_3, width=8, bg=BG_FORMULA)
        mw_e3.grid(row=3, column=4, padx=2)
        mw_e4 = make_entry(f, self.v_psb_mw_4, width=8, bg=BG_FORMULA)
        mw_e4.grid(row=3, column=5, padx=2)
        _mark_detailed(mw_e3, mw_e4)
        # Track which widgets are mol-only and which are also detailed-only.
        # No left-side label any more — column 0 stays empty for this row,
        # mirroring the XLSM where row 25 has no T25 caption.
        self._psb_mw_widgets     = [mw_unit, mw_e1, mw_e2]
        self._psb_mw_widgets_det = [mw_e3, mw_e4]

        # Row 4: Converted Kf — BLACK Unit + BLACK locked values
        tk.Label(f, text="Converted PSB's Freundlich Kf",
                 font=FONT_LABEL, bg=BG_MAIN, anchor="e"
                 ).grid(row=4, column=0, sticky="e", pady=1, padx=(0, 4))
        tk.Label(f, text="(ug/kg)(ug/L)^(-a)", font=FONT_LABEL_SM,
                 bg=BG_LOCKED, fg="#FFFFFF",
                 relief="solid", bd=1, padx=2
                 ).grid(row=4, column=1, padx=2, sticky="ew")
        for c, var in enumerate([self.v_psb_kf_conv, self.v_psb_kf_conv2]):
            make_entry(f, var, width=8, bg=BG_LOCKED
                       ).grid(row=4, column=2 + c, padx=2)
        e_cv3 = make_entry(f, self.v_psb_kf_conv3, width=8, bg=BG_LOCKED)
        e_cv3.grid(row=4, column=4, padx=2)
        e_cv4 = make_entry(f, self.v_psb_kf_conv4, width=8, bg=BG_LOCKED)
        e_cv4.grid(row=4, column=5, padx=2)
        _mark_detailed(e_cv3, e_cv4)

        # MW row visibility — driven by current Kf unit selection.
        def _on_kf_unit_change(*_):
            unit   = (self.v_psb_kf_unit.get() or "").lower()
            is_mol = "mol" in unit
            is_det = (self.active_sheet == "Detailed_2")
            for w in self._psb_mw_widgets:
                try:
                    w.grid() if is_mol else w.grid_remove()
                except Exception:
                    pass
            for w in self._psb_mw_widgets_det:
                try:
                    w.grid() if (is_mol and is_det) else w.grid_remove()
                except Exception:
                    pass
        self.v_psb_kf_unit.trace_add("write", _on_kf_unit_change)
        self._on_psb_kf_unit_change = _on_kf_unit_change   # exposed so the
        #  model-version toggle handler can re-run it.
        _on_kf_unit_change()  # initial state — hidden by default (mg-based)

        # ── Auto-compute Converted PSB's Freundlich Kf ────────────
        # Conversion to (ug/kg)(ug/L)^(-a):
        #     Kf_ug = Kf × m^(1-a)
        # where m is the unit-prefix factor (mass1 → ug):
        #     ng/kg     m = 1e-3
        #     ug/kg     m = 1.0
        #     mg/kg     m = 1e3
        #     nmol/kg   m = MW × 1e-3
        #     umol/kg   m = MW
        #     mmol/kg   m = MW × 1e3
        # `a` and `MW` come from row 1 / row 3 of the same column.
        def _to_float(var, default=None):
            try:
                s = (var.get() or "").strip().replace(",", "")
                return float(s) if s else default
            except (ValueError, AttributeError):
                return default

        def _unit_prefix_factor(unit_text, mw):
            u = (unit_text or "").lower()
            if "ng/kg"   in u: return 1e-3
            if "ug/kg"   in u: return 1.0
            if "mg/kg"   in u: return 1e3
            # Mol-based units need a positive MW (g/mol) to be defined
            if mw is None or mw <= 0:
                return None
            if "nmol/kg" in u: return mw * 1e-3
            if "umol/kg" in u: return mw
            if "mmol/kg" in u: return mw * 1e3
            return None

        def _recompute_psb_conv_kf(*_):
            unit = self.v_psb_kf_unit.get()
            for var_a, var_kf, var_mw, var_out in [
                (self.v_psb_a_1, self.v_psb_kf_1, self.v_psb_mw_1, self.v_psb_kf_conv),
                (self.v_psb_a_2, self.v_psb_kf_2, self.v_psb_mw_2, self.v_psb_kf_conv2),
                (self.v_psb_a_3, self.v_psb_kf_3, self.v_psb_mw_3, self.v_psb_kf_conv3),
                (self.v_psb_a_4, self.v_psb_kf_4, self.v_psb_mw_4, self.v_psb_kf_conv4),
            ]:
                a   = _to_float(var_a)
                kf  = _to_float(var_kf)
                mw  = _to_float(var_mw)
                if a is None or kf is None:
                    var_out.set("")
                    continue
                m = _unit_prefix_factor(unit, mw)
                if m is None:
                    var_out.set("")
                    continue
                try:
                    conv = kf * (m ** (1 - a))
                except (ValueError, OverflowError, ZeroDivisionError):
                    var_out.set("")
                    continue
                # Format with thousands separators — never scientific.
                # Whole numbers display without decimals ("1,227,951"),
                # otherwise up to 4 decimals are kept and trailing zeros
                # stripped ("0.33", "0.5617", "1,234.56").
                if conv == 0:
                    var_out.set("0")
                elif abs(conv - round(conv)) < 1e-9:
                    var_out.set(f"{int(round(conv)):,}")
                else:
                    var_out.set(f"{conv:,.4f}".rstrip("0").rstrip("."))

        # Wire traces — any of these StringVar changes triggers a recompute.
        for _v in (self.v_psb_kf_unit,
                   self.v_psb_a_1, self.v_psb_a_2, self.v_psb_a_3, self.v_psb_a_4,
                   self.v_psb_kf_1, self.v_psb_kf_2, self.v_psb_kf_3, self.v_psb_kf_4,
                   self.v_psb_mw_1, self.v_psb_mw_2, self.v_psb_mw_3, self.v_psb_mw_4):
            _v.trace_add("write", _recompute_psb_conv_kf)
        self._recompute_psb_conv_kf = _recompute_psb_conv_kf
        _recompute_psb_conv_kf()  # initial value

        # ── Vertical separator ────────────────────────────────────
        sep = tk.Frame(top, bg="#000000", width=1)
        sep.pack(side="left", fill="y", padx=(8, 8))

        # ── PSB Dist From Source — multi-line BLACK ───────────────
        psbd = tk.Frame(top, bg=BG_MAIN)
        psbd.pack(side="left", anchor="n", padx=(0, 8))
        for w in ("PSB", "Dist.", "From", "Source"):
            tk.Label(psbd, text=w, font=FONT_LABEL, bg=BG_MAIN
                     ).pack(anchor="w")
        dist_row = tk.Frame(psbd, bg=BG_MAIN)
        dist_row.pack(anchor="w", pady=(4, 0))
        make_entry(dist_row, self.v_psb_dist, width=6
                   ).pack(side="left")
        tk.Label(dist_row, text="(m)", font=FONT_LABEL_SM, bg=BG_MAIN
                 ).pack(side="left", padx=(2, 0))
        help_link(dist_row, "OpenTable9_3").pack(side="left", padx=(2, 0))

        # ── RIGHT cluster: image + Total Width stack + Year PSB,
        # vertically grouped & centered ───────────────────────────
        rcluster = tk.Frame(top, bg=BG_MAIN)
        rcluster.pack(side="left", anchor="n", padx=(4, 0))

        # Image on a Canvas so we can overlay a curved gray connector
        # line that links the "PSB Dist From Source" help "?" chicklet
        # (the last widget in `psbd.dist_row`, just outside this
        # canvas's bottom-left corner) to the red SOURCE square on the
        # 3D-box figure.  The canvas is intentionally taller than the
        # image — the extra strip beneath the image gives the curve
        # vertical room to drop down to "?"-button height.
        scale = getattr(self, "_dpi_scale", 1.0)
        self._figures["s9_1"] = _load_figure("Section9_1.png",
                                             target_height=int(70 * scale))
        img = self._figures["s9_1"]
        if img is not None:
            iw = img.width()
            ih = img.height()
            # Tighter below-image strip: ~14 px (was 34 px).  Reduces the
            # vertical gap between the figure and the Year PSB / Total
            # Width / fcac / # cells stack so §10 can move up.  The
            # curve still droops below the image enough to read as
            # pointing at the "?" chicklet sitting in psbd.dist_row.
            extra = int(14 * scale)
            ch = ih + extra                     # canvas height
            img_canvas = tk.Canvas(rcluster, width=iw, height=ch,
                                   bg=BG_MAIN, highlightthickness=0, bd=0)
            img_canvas.pack(anchor="nw", pady=(0, 0))
            img_canvas.create_image(0, 0, anchor="nw", image=img)
            # Source rectangle on the original 1165×365 png is at roughly
            # x≈55, y≈140 (left red square).  Scale into the resized
            # image and use that as the curve's terminus.
            sx = max(2, int(iw * 0.06))
            sy = int(ih * 0.42)
            # Curve start = bottom-left of canvas, which now sits at the
            # same vertical level as the "?" help chicklet inside `psbd`.
            img_canvas.create_line(
                0,                  ch - 2,           # start at "?" level
                int(iw * 0.04),     ih + int(extra * 0.4),
                int(iw * 0.06),     int(ih * 0.80),
                sx,                 sy,               # land on source
                smooth=True,
                fill="#7F7F7F", width=2, capstyle="round",
            )

        # Year PSB Installed — packed ABOVE the Total Width stack and
        # left-anchored so it appears slightly to the left of the
        # "Total Width of PSB in X-Direction" label that sits below.
        # pady tightened from (0, 4) → (0, 0) so the cell hugs the
        # bottom of the image canvas above it.
        yr = tk.Frame(rcluster, bg=BG_MAIN)
        yr.pack(anchor="w", pady=(0, 0))
        tk.Label(yr, text="Year PSB Installed", font=FONT_LABEL,
                 bg=BG_MAIN
                 ).pack(side="left")
        make_entry(yr, self.v_psb_yr, width=8
                   ).pack(side="left", padx=(4, 2))
        help_link(yr, "OpenTable9_3").pack(side="left", padx=(2, 0))

        # Total Width / fcac / # cells stack (right-anchored so the
        # right-edge of these long labels stays aligned with the
        # rcluster column; YR above it visually pokes out to the left).
        rg = tk.Frame(rcluster, bg=BG_MAIN)
        rg.pack(anchor="e", pady=(0, 0))
        for i, (lbl, var, unit, helpm) in enumerate([
            ("Total Width of PSB in X-Direction",
             self.v_psb_width, "(m) (Typical 4)", "OpenTable9_4"),
            ("PSB Loading 'fcac'",
             self.v_psb_load,  "(%)",              "OpenTable9_5"),
            ("# of cells in PSB in x direction",
             self.v_psb_cells, "(-)",              ""),
        ]):
            tk.Label(rg, text=lbl, font=FONT_LABEL, bg=BG_MAIN,
                     anchor="e"
                     ).grid(row=i, column=0, sticky="e", padx=(0, 4),
                            pady=1)
            make_entry(rg, var, width=6
                       ).grid(row=i, column=1, padx=2)
            tk.Label(rg, text=unit, font=FONT_LABEL_SM, bg=BG_MAIN
                     ).grid(row=i, column=2, sticky="w", padx=(2, 0))
            if helpm:
                help_link(rg, helpm).grid(row=i, column=3, padx=(2, 0))

        # ═══ Buttons — packed under the form in `leftcol` so they
        # sit flush against the bottom of the Freundlich grid (no
        # empty gap between the form and the button row).
        bf = tk.Frame(leftcol, bg=BG_MAIN)
        bf.pack(anchor="w", pady=(8, 0))
        make_btn(bf, "Where to Get Freundlich\nParameters",
                 "OpenAppendix_9_1_Relative", quarto=True,
                 font=FONT_BTN, width=28, bg=BTN_FILL,
                 padx=10, pady=16
                 ).pack(side="left", padx=(0, 12))
        make_btn(bf, "Simple CAC Barrier\nLongevity Tool",
                 "LongevityTool", font=FONT_BTN, width=28, bg=BTN_FILL,
                 padx=10, pady=16
                 ).pack(side="left")

        # In Simple, hide registered Detailed-only widgets
        if self.active_sheet != "Detailed_2":
            for w in det_only:
                try:
                    w.grid_remove()
                except Exception:
                    pass

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 10 – FIELD DATA TO CALIBRATE
    #
    # Layout depends on model version:
    #   SIMPLE  : 3-column inline table — header + Sample Year + helper
    #             notes | MW table | separator | Location Info + Distance
    #   DETAILED: header + a single big "Enter Monitoring Well Data" button
    # ─────────────────────────────────────────────────────────────────────
    def _build_s10_field_data(self, parent):
        outer = tk.Frame(parent, bg=BG_MAIN)
        # Tighter top padding — pulls §10 closer to §9 so the §9 button
        # row no longer leaves a visible empty band above it.
        outer.pack(fill="x", pady=(0, 4))

        # ═══ Section header (multi-line) — ALWAYS VISIBLE ═══════════════
        leftcol = tk.Frame(outer, bg=BG_MAIN)
        leftcol.pack(side="left", anchor="nw", padx=(0, 12))
        tk.Label(leftcol, text="10. FIELD\nDATA TO\nCALIBRATE",
                 font=FONT_SECTION, fg=FG_SECTION, bg=BG_MAIN,
                 justify="left", anchor="w"
                 ).pack(anchor="w", pady=(0, 12))

        # Sample Year + helper notes — SIMPLE ONLY
        sy_block = tk.Frame(leftcol, bg=BG_MAIN)
        sy_block.pack(anchor="w")
        tk.Label(sy_block, text="Sample Year (XXXX)",
                 font=FONT_LABEL_I, bg=BG_MAIN
                 ).pack(anchor="w")
        sy_row = tk.Frame(sy_block, bg=BG_MAIN); sy_row.pack(anchor="w")
        # Black value text on white bg — matches Excel (red is reserved
        # for the Distance-from-Source column only).
        tk.Entry(sy_row, textvariable=self.v_sample_yr, width=6,
                 font=FONT_INPUT, bg=BG_INPUT_BLUE, fg=FG_INPUT,
                 relief="solid", bd=1, justify="right"
                 ).pack(side="left", padx=(0, 4))
        help_link(sy_row, "OpenTable10_1").pack(side="left")
        helper_fr = tk.Frame(sy_block, bg=BG_MAIN)
        helper_fr.pack(anchor="e", pady=(16, 0))
        # Helper notes — black italic (Excel uses red only on the
        # Distance-from-Source column, not on these helper notes).
        for txt in ["Enter up to 7 shallow",
                    "centerline wells",
                    "for key monitoring event"]:
            tk.Label(helper_fr, text=txt, font=FONT_LABEL_SMI,
                     bg=BG_MAIN, fg=FG_INPUT, anchor="e"
                     ).pack(anchor="e")
        sy_block._toggle_kind = "pack"
        self._simple_only_frames.append(sy_block)

        # ═══ DETAILED-ONLY: just a big button ═══════════════════════════
        detailed_block = tk.Frame(outer, bg=BG_MAIN)
        detailed_block.pack(side="left", anchor="n", padx=(60, 0))
        # Big rounded button — italic blue text, gray fill.  Matches the
        # Excel storyboard reference: visually dominant, with whitespace
        # around it.
        make_btn(detailed_block, "Enter Monitoring\nWell Data",
                 "CalibrationDataLoader",
                 font=FONT_BTN_LG, width=30, bg=BTN_FILL,
                 padx=40, pady=24
                 ).pack(anchor="center", padx=20, pady=30)
        detailed_block._toggle_kind = "pack"
        self._detailed_only_frames.append(detailed_block)

        # ═══ SIMPLE-ONLY: MW table + separator + Location Info ══════════
        simple_block = tk.Frame(outer, bg=BG_MAIN)
        simple_block.pack(side="left", anchor="nw")
        simple_block._toggle_kind = "pack"
        self._simple_only_frames.append(simple_block)

        mid = tk.Frame(simple_block, bg=BG_MAIN)
        mid.pack(side="left", anchor="nw")

        # Row 0 — italic group titles
        tk.Label(mid, text="Monitoring", font=FONT_LABEL_BI,
                 bg=BG_MAIN, fg=FG_GREY).grid(row=0, column=0, padx=4)
        tk.Label(mid, text="Monitoring Well", font=FONT_LABEL_BI,
                 bg=BG_MAIN, fg=FG_GREY
                 ).grid(row=0, column=1, columnspan=2, padx=4)

        # Row 1 — sub-titles
        tk.Label(mid, text="Well Name", font=FONT_LABEL_BI,
                 bg=BG_MAIN, fg=FG_GREY).grid(row=1, column=0, padx=4)
        sub_fr = tk.Frame(mid, bg=BG_MAIN)
        sub_fr.grid(row=1, column=1, columnspan=2, padx=4, sticky="ew")
        tk.Label(sub_fr, text="Concentration Data (ug/L)",
                 font=FONT_LABEL_BI, bg=BG_MAIN, fg=FG_GREY
                 ).pack(side="left")
        help_link(sub_fr, "OpenTable10_3").pack(side="left", padx=(4, 0))

        # Row 2 — column headers (Event | PFOS | None) + Event help.
        # "Event" sits inside its own BLACK locked cell to mirror the
        # PFOS / None header style (per Excel reference).
        hdr_event = tk.Frame(mid, bg=BG_MAIN)
        hdr_event.grid(row=2, column=0, padx=4)
        tk.Label(hdr_event, text="Event", font=FONT_LABEL_B,
                 bg=BG_LOCKED, fg="#FFFFFF",
                 relief="solid", bd=1, padx=10
                 ).pack(side="left")
        help_link(hdr_event, "OpenTable10_2").pack(side="left", padx=(4, 0))
        tk.Label(mid, text="PFOS", font=FONT_LABEL_B,
                 bg=BG_LOCKED, fg="#FFFFFF",
                 relief="solid", bd=1, width=8
                 ).grid(row=2, column=1, padx=4, sticky="ew")
        tk.Label(mid, text="None", font=FONT_LABEL_B,
                 bg=BG_LOCKED, fg="#FFFFFF",
                 relief="solid", bd=1, width=8
                 ).grid(row=2, column=2, padx=4, sticky="ew")

        # Rows 3-9 — 7 monitoring well rows.
        # Well names: WHITE bg + BLACK text (right-justified, read-only)
        #   — matches Excel; previous black-bg/white-text was wrong.
        # Concentration values: BLACK text on white (red is reserved for
        # the Distance column only).
        for i in range(7):
            tk.Entry(mid, textvariable=self.v_mw_names[i], width=10,
                     font=FONT_INPUT, bg=BG_INPUT_BLUE, fg=FG_INPUT,
                     relief="solid", bd=1, justify="right",
                     state="readonly", readonlybackground=BG_INPUT_BLUE
                     ).grid(row=3+i, column=0, padx=2, pady=1)
            tk.Entry(mid, textvariable=self.v_mw_conc[i], width=8,
                     font=FONT_INPUT, bg=BG_INPUT_BLUE, fg=FG_INPUT,
                     relief="solid", bd=1, justify="right"
                     ).grid(row=3+i, column=1, padx=2, pady=1)
            tk.Entry(mid, textvariable=self.v_mw_conc2[i], width=8,
                     font=FONT_INPUT, bg="#FFFFFF", fg=FG_INPUT,
                     relief="solid", bd=1, justify="right"
                     ).grid(row=3+i, column=2, padx=2, pady=1)
            tk.Label(mid, text="(ug/L)", font=FONT_LABEL_SMI,
                     bg=BG_MAIN, fg=FG_INPUT
                     ).grid(row=3+i, column=3, sticky="w", padx=2)

        # Vertical separator
        sep = tk.Frame(simple_block, bg="#000000", width=2)
        sep.pack(side="left", fill="y", padx=(16, 16))

        # Right — Location Info + Distance
        rt = tk.Frame(simple_block, bg=BG_MAIN)
        rt.pack(side="left", anchor="nw")
        loc_hdr = tk.Frame(rt, bg=BG_MAIN)
        loc_hdr.grid(row=0, column=0, columnspan=2, sticky="w",
                     padx=4, pady=(0, 2))
        tk.Label(loc_hdr, text="Monitoring Well Location Info:",
                 font=FONT_LABEL_BI, bg=BG_MAIN, fg=FG_GREY
                 ).pack(side="left")
        dist_hdr = tk.Frame(rt, bg=BG_MAIN)
        dist_hdr.grid(row=1, column=1, sticky="e", padx=4)
        tk.Label(dist_hdr, text="Distance\nfrom\nSource (m)",
                 font=FONT_LABEL_BI, bg=BG_MAIN, fg=FG_GREY,
                 justify="center"
                 ).pack(side="left")
        help_link(dist_hdr, "OpenTable10_4").pack(side="left", padx=(4, 0))

        descr = [
            "For simple model use only",
            "monitoring wells on or",
            "near the top of the  plume",
            "",
            "and on the",
            "centerline of the plume",
            "",
        ]
        # Right-side descriptive notes — BLACK italic per Excel.
        # Distances stay RED on white — that's the only red column in §10.
        for i in range(7):
            tk.Label(rt, text=descr[i], font=FONT_LABEL_SMI,
                     bg=BG_MAIN, fg=FG_INPUT, anchor="e"
                     ).grid(row=2+i, column=0, sticky="e",
                            padx=(0, 8), pady=1)
            tk.Entry(rt, textvariable=self.v_mw_dist[i], width=6,
                     font=FONT_INPUT, bg="#FFFFFF", fg="#C00000",
                     relief="solid", bd=1, justify="right"
                     ).grid(row=2+i, column=1, padx=2, pady=1)

        # Hide the wrong block on first paint based on current version.
        if self.active_sheet == "Detailed_2":
            self._toggle_widgets([sy_block, simple_block], show=False)
        else:
            self._toggle_widgets([detailed_block], show=False)

    # SECTION 11
    def _build_s11_output(self, parent):
        # Single horizontal row containing 3 columns:
        #   LEFT   – multi-line header + Change Numerical Parameters btn
        #   MIDDLE – "See Results Every: [val] (years) ?"
        #   RIGHT  – blue action panel (Run Model + 3x2 button grid)
        outer = tk.Frame(parent, bg=BG_MAIN)
        outer.pack(fill="x", anchor="w", pady=(4, 4))

        # ── LEFT — header + Change Numerical Parameters ──────────────
        leftcol = tk.Frame(outer, bg=BG_MAIN)
        leftcol.pack(side="left", anchor="nw", padx=(0, 16))
        tk.Label(leftcol,
                 text="11. MODEL OUTPUT AND\nNUMERICAL PARAMETERS",
                 font=FONT_SECTION, fg=FG_SECTION, bg=BG_MAIN,
                 justify="left", anchor="w"
                 ).pack(anchor="w", pady=(0, 8))
        make_btn(leftcol, "Change Numerical Parameters",
                 "ChangeNumericalParameters",
                 font=FONT_BTN_SM, width=26, bg=BTN_FILL
                 ).pack(anchor="w")

        # ── MIDDLE — See Results Every: [entry] (years) ? ────────────
        midcol = tk.Frame(outer, bg=BG_MAIN)
        midcol.pack(side="left", anchor="n", padx=(0, 16))
        tk.Label(midcol, text="See Results", font=FONT_LABEL_I,
                 bg=BG_MAIN, justify="center"
                 ).pack(anchor="center")
        tk.Label(midcol, text="Every:", font=FONT_LABEL_I,
                 bg=BG_MAIN, justify="center"
                 ).pack(anchor="center")
        sr_row = tk.Frame(midcol, bg=BG_MAIN); sr_row.pack(anchor="center")
        make_entry(sr_row, self.v_see_every, width=6
                   ).pack(side="left", padx=(0, 2))
        tk.Label(sr_row, text="(years)", font=FONT_LABEL_SM, bg=BG_MAIN
                 ).pack(side="left")
        help_link(sr_row, "OpenTable11_1").pack(side="left", padx=(2, 0))

        # ── RIGHT — blue action panel (looser internal spacing) ────
        # Bumped panel padx/pady 14 → 20 to give the buttons more breathing
        # room against the panel border, matching the Excel reference.
        bar = tk.Frame(outer, bg=BTN_FILL_BLUE, bd=3, relief="solid",
                       padx=20, pady=20)
        bar.pack(side="left", anchor="n", fill="x", expand=True)

        # Two big Run Model buttons — gray fill (Excel storyboard)
        # while the surrounding panel stays blue.  side-pad 10 → 16.
        make_btn(bar, "Run Model", "RunPythonScript",
                 fg=FG_BTN_NAVY, font=FONT_BTN_LG, padx=18, pady=14,
                 bg=BTN_FILL, width=14).pack(side="left", padx=16)
        make_btn(bar, "Run Model with\nAuto-Calibration",
                 "Show_Visualization",
                 fg=FG_BTN_NAVY, font=FONT_BTN_LG, padx=18, pady=14,
                 bg=BTN_FILL, width=18).pack(side="left", padx=16)

        # 3x2 action grid on the right (5 buttons; bottom-left empty).
        # Gap to grid bumped 20 → 32 so the action group is clearly
        # separated from the Run buttons.  Cell padding bumped 6 → 10
        # for visibly looser spacing between buttons in both rows.
        actfr = tk.Frame(bar, bg=BTN_FILL_BLUE)
        actfr.pack(side="left", padx=(32, 0))
        cells = [
            (0, 0, "Authors",        "Authors",       BTN_FILL_GREEN, FG_BTN_GREEN),
            (0, 1, "Load Data",      "Load_Data",     "#FFFFFF",      FG_BTN_NAVY),
            (0, 2, "Save Data",      "Save_Data",     BTN_FILL_GREEN, FG_BTN_GREEN),
            (1, 1, "Clear All Data", "Clear_Data",    BTN_FILL,       FG_BTN_NAVY),
            (1, 2, "Paste Example",  "Paste_Example", BTN_FILL,       FG_BTN_NAVY),
        ]
        for r, c, txt, macro, bg, fg in cells:
            make_btn(actfr, txt, macro,
                     fg=fg, font=FONT_BTN, padx=14, pady=10,
                     bg=bg, width=14
                     ).grid(row=r, column=c, padx=10, pady=10, sticky="ew")

    # BOTTOM BAR
    # ─────────────────────────────────────────────────────────────────────
    # ACTION ROW — Run / Authors / Load / Save / Help / Clear / Paste.
    # Lives inside the right column of the top body (per PDF storyboard).
    # ─────────────────────────────────────────────────────────────────────
    def _build_action_row(self, parent):
        # Per Excel: the whole action cluster sits inside a solid blue
        # rectangle (BTN_FILL_BLUE) at the bottom-right of upper page.
        bar = tk.Frame(parent, bg=BTN_FILL_BLUE, bd=1, relief="solid",
                       padx=6, pady=6)
        bar.pack(fill="x", anchor="ne", pady=(8, 0))

        # Two big Run Model buttons (blue filled, navy italic)
        make_btn(bar, "Run Model", "RunPythonScript",
                 fg=FG_BTN_NAVY, font=FONT_BTN_LG, padx=10, pady=4,
                 bg=BTN_FILL_BLUE, width=12).pack(side="left", padx=4)
        make_btn(bar, "Run Model with\nAuto-Calibration",
                 "Show_Visualization",
                 fg=FG_BTN_NAVY, font=FONT_BTN_LG, padx=10, pady=4,
                 bg=BTN_FILL_BLUE, width=18).pack(side="left", padx=4)

        # 3x2 action grid on the right.  Authors + Save Data = green;
        # Load Data white-ish; Clear/Paste/Help neutral gray.
        actfr = tk.Frame(bar, bg=BTN_FILL_BLUE)
        actfr.pack(side="right", padx=4)
        cells = [
            (0, 0, "Authors",        "Authors",       BTN_FILL_GREEN, FG_BTN_GREEN),
            (0, 1, "Load Data",      "Load_Data",     BG_INPUT_BLUE,  FG_BTN_NAVY),
            (0, 2, "Save Data",      "Save_Data",     BTN_FILL_GREEN, FG_BTN_GREEN),
            (1, 0, "Help",           "Help",          BTN_FILL,       FG_BTN_NAVY),
            (1, 1, "Clear All Data", "Clear_Data",    BTN_FILL,       FG_BTN_NAVY),
            (1, 2, "Paste Example",  "Paste_Example", BTN_FILL,       FG_BTN_NAVY),
        ]
        for r, c, txt, macro, bg, fg in cells:
            make_btn(actfr, txt, macro,
                     fg=fg, font=FONT_BTN, padx=8, pady=3,
                     bg=bg, width=12).grid(
                row=r, column=c, padx=3, pady=2, sticky="ew")

    # ─────────────────────────────────────────────────────────────────────
    # CALIBRATION PANEL — bottom-right quadrant per the Excel storyboard.
    # Layout matches the Simple + Detailed reference screenshots exactly:
    #   • Header band: title + 5 step lines (left)  |  iterations + time
    #     remaining + caveat (right)
    #   • Three-box row: Step 2 (Calibrate using) | Step 3 (Weighting
    #     factor table) | LEGEND
    #   • Step 4: Enter Calibration Ranges — table with parameter
    #     checkboxes / Lowest / Mid-Range (black, locked) / Highest
    #     plus units and Section refs.  Detailed mode adds 6 Precursor
    #     rows after a small gap.
    #   • Action buttons centered in 3 rows.
    # ─────────────────────────────────────────────────────────────────────
    def _build_calibration_panel(self, parent):
        outer = tk.Frame(parent, bg=BG_MAIN)
        outer.pack(fill="both", expand=True, padx=8, pady=6)

        # ── Top header band: title + step lines (left) | iter/time (right)
        header_band = tk.Frame(outer, bg=BG_MAIN)
        header_band.pack(fill="x", pady=(0, 4))

        head_l = tk.Frame(header_band, bg=BG_MAIN)
        head_l.pack(side="left", anchor="nw", fill="x", expand=True)
        tk.Label(head_l,
                 text="REMFluor-MD MACHINE-CALIBRATION (Singh et al., 2025)",
                 font=FONT_LABEL_B, bg=BG_MAIN, fg=FG_INPUT, anchor="w"
                 ).pack(anchor="w")
        tk.Label(head_l,
            text="This allows you to let the computer perform a simple "
                 "calibration of your REMFluor-MD model by:",
            font=FONT_LABEL, bg=BG_MAIN, fg=FG_INPUT, anchor="w",
            wraplength=720, justify="left"
            ).pack(anchor="w", pady=(2, 0))
        # 5 Step lines — red "Step N)" tag, black body
        steps = [
            ("Step 1)", "Entering monitoring well data in Section 9."),
            ("Step 2)", "Select which parameters to calibrate against."),
            ("Step 3)", "Enter calibration options:  decide if you want "
                        "to weight any monitoring data more or less "
                        "during calibration."),
            ("Step 4)", "Enter calibration data, the likely minimum and "
                        "maximum values for the selected parameters below."),
            ("Step 5)", 'Hitting the "Run Machine-Based Calibration" '
                        "button to the right."),
        ]
        for tag, body_txt in steps:
            srow = tk.Frame(head_l, bg=BG_MAIN)
            srow.pack(anchor="w", padx=14)
            tk.Label(srow, text=tag, font=FONT_LABEL, bg=BG_MAIN,
                     fg=FG_HELP).pack(side="left")
            tk.Label(srow, text=" " + body_txt, font=FONT_LABEL,
                     bg=BG_MAIN, fg=FG_INPUT, wraplength=620,
                     justify="left", anchor="w"
                     ).pack(side="left", fill="x")

        # Right side: iterations + time remaining
        head_r = tk.Frame(header_band, bg=BG_MAIN)
        head_r.pack(side="right", anchor="ne", padx=(8, 4))
        nf = tk.Frame(head_r, bg=BG_MAIN); nf.pack(anchor="e")
        tk.Label(nf, text="Number of iteration", font=FONT_LABEL,
                 bg=BG_MAIN).pack(side="left")
        self.v_n_iter = tk.StringVar(value="50")
        tk.Entry(nf, textvariable=self.v_n_iter, width=6,
                 font=FONT_INPUT, bg=BG_INPUT_BLUE, fg=FG_INPUT,
                 relief="solid", bd=1, justify="right"
                 ).pack(side="left", padx=(6, 0))
        tf = tk.Frame(head_r, bg=BG_MAIN); tf.pack(anchor="e", pady=(2, 0))
        tk.Label(tf, text="Estim.Time Remaining", font=FONT_LABEL,
                 bg=BG_MAIN).pack(side="left")
        self.v_t_remain = tk.StringVar(value="2")
        tk.Entry(tf, textvariable=self.v_t_remain, width=6,
                 font=FONT_INPUT, bg=BG_LOCKED, fg=FG_LOCKED,
                 relief="solid", bd=1, justify="right",
                 readonlybackground=BG_LOCKED, state="readonly"
                 ).pack(side="left", padx=(6, 0))
        time_unit = "(hours)" if self.active_sheet == "Detailed_2" \
                    else "(minutes)"
        tk.Label(tf, text=time_unit, font=FONT_LABEL_SM,
                 bg=BG_MAIN).pack(side="left", padx=(2, 0))
        tk.Label(head_r,
                 text="(add explainer text here about\npossible run times)",
                 font=FONT_LABEL_SMI, fg=FG_HELP, bg=BG_MAIN,
                 justify="left").pack(anchor="e", pady=(4, 0))

        # ── Three-box row: Step 2 | Step 3 | LEGEND ─────────────────────
        body = tk.Frame(outer, bg=BG_MAIN)
        body.pack(fill="x", expand=True, pady=(8, 0))

        # --- Step 2: Calibrate using ---
        s2 = tk.Frame(body, bg=BG_MAIN, bd=1, relief="solid",
                      padx=12, pady=10)
        s2.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        tk.Label(s2, text="Step 2:  Calibrate using:",
                 font=FONT_LABEL_B, bg=BG_MAIN, anchor="w"
                 ).grid(row=0, column=0, columnspan=2, sticky="w",
                        pady=(0, 6))
        self.v_calib_pfoa = tk.BooleanVar(value=True)
        self.v_calib_none = tk.BooleanVar(value=False)
        tk.Label(s2, text="PFOS", font=FONT_LABEL, bg=BG_MAIN
                 ).grid(row=1, column=0, sticky="e", padx=(8, 4),
                        pady=2)
        tk.Checkbutton(s2, variable=self.v_calib_pfoa, bg=BG_MAIN
                       ).grid(row=1, column=1, sticky="w", pady=2)
        tk.Label(s2, text="None", font=FONT_LABEL, bg=BG_MAIN
                 ).grid(row=2, column=0, sticky="e", padx=(8, 4),
                        pady=2)
        tk.Checkbutton(s2, variable=self.v_calib_none, bg=BG_MAIN
                       ).grid(row=2, column=1, sticky="w", pady=2)
        # Detailed-only: PFAA 1-able + None for precursors
        self.v_calib_pre      = tk.BooleanVar(value=True)
        self.v_calib_pre_none = tk.BooleanVar(value=False)
        det_lbl1 = tk.Label(s2, text="PFAA 1-able", font=FONT_LABEL,
                            bg=BG_MAIN)
        det_chk1 = tk.Checkbutton(s2, variable=self.v_calib_pre,
                                   bg=BG_MAIN)
        det_lbl2 = tk.Label(s2, text="None", font=FONT_LABEL,
                            bg=BG_MAIN)
        det_chk2 = tk.Checkbutton(s2, variable=self.v_calib_pre_none,
                                   bg=BG_MAIN)
        det_lbl1.grid(row=3, column=0, sticky="e", padx=(8, 4), pady=2)
        det_chk1.grid(row=3, column=1, sticky="w", pady=2)
        det_lbl2.grid(row=4, column=0, sticky="e", padx=(8, 4), pady=2)
        det_chk2.grid(row=4, column=1, sticky="w", pady=2)
        for w in (det_lbl1, det_chk1, det_lbl2, det_chk2):
            w._toggle_kind = "grid"
            self._detailed_only_frames.append(w)

        # --- Step 3: Calibration Options (weighting) ---
        s3 = tk.Frame(body, bg=BG_MAIN, bd=1, relief="solid",
                      padx=12, pady=10)
        s3.grid(row=0, column=1, sticky="nsew", padx=(0, 6))
        tk.Label(s3, text="Step 3:  Calibration Options - "
                          "enter any weighting factors.",
                 font=FONT_LABEL_B, bg=BG_MAIN, anchor="w"
                 ).pack(anchor="w", pady=(0, 4))
        s3g = tk.Frame(s3, bg=BG_MAIN); s3g.pack()
        for j, h in enumerate(["Monitoring\nPoint\nName",
                                "Distance\nfrom\nSource (m)",
                                "Weighting\nFactor\nfor Calb:",
                                "Source:"]):
            tk.Label(s3g, text=h, font=FONT_LABEL_BI, bg=BG_MAIN,
                     fg=FG_INPUT
                     ).grid(row=0, column=j, sticky="w", padx=4,
                            pady=(0, 2))
        self.v_calib_w = []
        defaults = ["1.0"]*6 + ["2.0"]
        for i, w in enumerate(defaults):
            tk.Entry(s3g, textvariable=self.v_mw_names[i], width=10,
                     font=FONT_INPUT, bg=BG_LOCKED, fg=FG_LOCKED,
                     relief="solid", bd=1, justify="center",
                     state="readonly", readonlybackground=BG_LOCKED
                     ).grid(row=i+1, column=0, sticky="w", padx=2,
                            pady=1)
            tk.Entry(s3g, textvariable=self.v_mw_dist[i], width=8,
                     font=FONT_INPUT, bg=BG_LOCKED, fg=FG_LOCKED,
                     relief="solid", bd=1, justify="right",
                     state="readonly", readonlybackground=BG_LOCKED
                     ).grid(row=i+1, column=1, sticky="w", padx=2,
                            pady=1)
            wv = tk.StringVar(value=w); self.v_calib_w.append(wv)
            tk.Entry(s3g, textvariable=wv, width=8, font=FONT_INPUT,
                     bg=BG_FORMULA, fg=FG_INPUT, relief="solid", bd=1,
                     justify="right"
                     ).grid(row=i+1, column=2, sticky="w", padx=2,
                            pady=1)
            tk.Label(s3g, text="Section 10", font=FONT_LABEL_I,
                     bg=BG_MAIN
                     ).grid(row=i+1, column=3, sticky="w", padx=4,
                            pady=1)
        tk.Label(s3,
                 text="(Often downgradient monitoring wells are "
                      "weighted higher)",
                 font=FONT_LABEL_SMI, fg=FG_HELP, bg=BG_MAIN
                 ).pack(anchor="center", pady=(6, 0))

        # --- LEGEND ---
        legend = tk.Frame(body, bg=BG_MAIN, bd=1, relief="solid",
                          padx=12, pady=10)
        legend.grid(row=0, column=2, sticky="nsew")
        tk.Label(legend, text="LEGEND", font=FONT_LABEL_B, bg=BG_MAIN
                 ).grid(row=0, column=0, columnspan=2, sticky="w",
                        pady=(0, 4))
        legend_items = [
            (BG_INPUT_BLUE, "Enter value directly"),
            (BG_FORMULA,    "Formula or default, but ok to overwrite**"),
            (BG_PULLDOWN,   "Pull Down Menu"),
            (BG_LOCKED,     "Calculated value or taken from other cell."),
        ]
        for r, (color, txt) in enumerate(legend_items, start=1):
            tk.Frame(legend, bg=color, width=22, height=18, bd=1,
                     relief="solid"
                     ).grid(row=r, column=0, sticky="w", padx=(0, 8),
                            pady=3)
            tk.Label(legend, text=txt, font=FONT_LABEL_SMI,
                     bg=BG_MAIN, anchor="w", justify="left",
                     wraplength=220
                     ).grid(row=r, column=1, sticky="w", pady=3)

        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=2)
        body.columnconfigure(2, weight=1)

        # ── Step 4: Enter Calibration Ranges ─────────────────────────────
        s4_hdr = tk.Frame(outer, bg=BG_MAIN)
        s4_hdr.pack(fill="x", pady=(12, 0))
        tk.Label(s4_hdr, text="Step 4:  Enter Calibration Ranges:",
                 font=FONT_LABEL_B, bg=BG_MAIN, anchor="w"
                 ).pack(side="left")
        rs_note = tk.Label(s4_hdr,
                           text="are taken from in the input cells in "
                                "these sections:",
                           font=FONT_LABEL_I, bg=BG_MAIN, fg=FG_INPUT)
        rs_y    = tk.Label(s4_hdr, text="Mid-Range Values",
                           font=FONT_LABEL_BI, bg=FG_YELLOW, fg=FG_INPUT,
                           padx=4)
        rs_note.pack(side="right", padx=(0, 4))
        rs_y.pack(side="right", padx=(0, 4))

        make_btn(outer,
                 "Enter Default Low-High Ranges from Literature, "
                 "Experience (default +/- x2)",
                 "DefaultRanges",
                 fg=FG_BTN_NAVY, font=FONT_BTN_SM, padx=10, pady=4,
                 bg=BTN_FILL).pack(anchor="w", pady=(2, 6))

        tbl = tk.Frame(outer, bg=BG_MAIN); tbl.pack(anchor="w", pady=2,
                                                     fill="x")
        tk.Label(tbl, text="Use this\nParameter?", font=FONT_LABEL_BI,
                 bg=BG_MAIN, fg=FG_INPUT
                 ).grid(row=0, column=1, padx=2, pady=2)
        tk.Label(tbl, text="Lowest\nLikely Value", font=FONT_LABEL_BI,
                 bg=BG_MAIN, fg=FG_INPUT
                 ).grid(row=0, column=2, padx=2, pady=2)
        tk.Label(tbl, text="Mid-Range\nValue", font=FONT_LABEL_BI,
                 bg=FG_YELLOW, fg=FG_INPUT, padx=4
                 ).grid(row=0, column=3, padx=2, pady=2, sticky="ew")
        tk.Label(tbl, text="Highest\nLikely Value", font=FONT_LABEL_BI,
                 bg=BG_MAIN, fg=FG_INPUT
                 ).grid(row=0, column=4, padx=2, pady=2)

        simple_rows = [
            ("Source Start Year (nt)",                     "1965",   "1977",   "1980",  "(xxxx)",     "Section 2"),
            ("Hydraulic Conductivity (k)",                 "0.31536","3.15",   "31.536","(m/yr)",     "Section 3"),
            ("Hydraulic Gradient (i)",                     "0.0004", "0.0038", "0.0380","(-)",        "Section 3"),
            ("Effective Porosity (porf)",                  "0.16",   "0.20",   "0.24",  "(-)",        "Section 3"),
            ("Transmissive Fraction of Model (volfrac)",   "0.6",    "1.00",   "0.85",  "(-)",        "Section 4"),
            ("Average Diffusion Length (difflen)",         "1.5",    "0.00",   "6",     "(m)",        "Section 4"),
            ("Retardation Factor of PFAA-1 (ock(2))",      "1.6",    "2.9",    "6.4",   "(-)",        "Section 5"),
            ("Retardation Factor of PFAA-2 (ock(4))",      "1",      "0.0",    "4",     "(-)",        "Section 5"),
            ("Longitudinal Dispersivity (alphax (m))",     "1",      "3.2",    "1.5",   "(m)",        "Section 6"),
            ("Multiplier to PFAA-1 Source Concentration in #7 (czero(2,n))",
                                                            "0.5",   "1",      "2",     "x(Ct)",      "Section 7"),
            ("Multiplier to PFAA-2 Source Concentration in #7 (czero(4,n))",
                                                            "0.5",   "#DIV/0!","2",     "x(Ct)",      "Section 7"),
        ]
        precursor_rows = [
            ("First order decay rate coefficient for Precursors-1 (decayf(1))",
                                                            "0.5",   "0.4",    "4.5",   "(per year)", "Section 5"),
            ("First order decay rate coefficient for Precursors-2 (decayf(3))",
                                                            "0.5",   "0.0",    "4.5",   "(per year)", "Section 5"),
            ("Retardation Factor of Precursors-1 (ock(1))",
                                                            "2",     "4.7",    "4",     "(-)",        "Section 5"),
            ("Retardation Factor of Precursors-2 (ock(3))",
                                                            "2",     "0.0",    "4",     "(-)",        "Section 5"),
            ("Multiplier to Precursor-1 Source Concentration in #7 (czero(1,n))",
                                                            "0.5",   "1.0",    "2",     "x(Ct)",      "Section 7"),
            ("Multiplier to Precursor-2 Source Concentration in #7 (czero(3,n))",
                                                            "0.5",   "#DIV/0!","2",     "x(Ct)",      "Section 7"),
        ]
        preselected = {
            "Hydraulic Conductivity (k)",
            "Hydraulic Gradient (i)",
            "Multiplier to PFAA-1 Source Concentration in #7 (czero(2,n))",
            "Multiplier to Precursor-1 Source Concentration in #7 (czero(1,n))",
        }
        self.v_calib_chk  = []
        self.v_calib_low  = []
        self.v_calib_mid  = []
        self.v_calib_high = []

        def _add_row(i, lbl, lo, mid, hi, unit, section,
                     detailed_only=False):
            wname = tk.Label(tbl, text=lbl, font=FONT_LABEL_I,
                              bg=BG_MAIN, anchor="e", justify="right")
            wname.grid(row=i, column=0, sticky="e", padx=(2, 4),
                       pady=1)
            cv = tk.BooleanVar(value=(lbl in preselected))
            self.v_calib_chk.append(cv)
            wchk = tk.Checkbutton(tbl, variable=cv, bg=BG_MAIN)
            wchk.grid(row=i, column=1, pady=1)
            cells = []
            for col, vlist, val, bg in [
                (2, self.v_calib_low,  lo,  BG_FORMULA),
                (3, self.v_calib_mid,  mid, BG_LOCKED),
                (4, self.v_calib_high, hi,  BG_FORMULA),
            ]:
                vv = tk.StringVar(value=val); vlist.append(vv)
                state_kw = {}
                fg = FG_INPUT
                if bg == BG_LOCKED:
                    fg = FG_LOCKED
                    state_kw = dict(state="readonly",
                                    readonlybackground=BG_LOCKED)
                e = tk.Entry(tbl, textvariable=vv, width=10,
                             font=FONT_INPUT, bg=bg, fg=fg,
                             relief="solid", bd=1, justify="right",
                             **state_kw)
                e.grid(row=i, column=col, padx=1, pady=1)
                cells.append(e)
            wunit = tk.Label(tbl, text=unit, font=FONT_LABEL,
                             bg=BG_MAIN, anchor="w")
            wunit.grid(row=i, column=5, sticky="w", padx=(4, 2),
                       pady=1)
            wsec = tk.Label(tbl, text=section, font=FONT_LABEL_I,
                             bg=BG_MAIN, anchor="w")
            wsec.grid(row=i, column=6, sticky="w", padx=(2, 4),
                      pady=1)
            if detailed_only:
                for w in (wname, wchk, *cells, wunit, wsec):
                    w._toggle_kind = "grid"
                    self._detailed_only_frames.append(w)

        ridx = 1
        for r in simple_rows:
            _add_row(ridx, *r)
            ridx += 1
        note_pfaa = tk.Label(tbl,
            text="(All 11 source concentrations for PFAA-1\n"
                 "are changed by the same amount during\n"
                 "calibration runs).",
            font=FONT_LABEL_SMI, fg=FG_HELP, bg=BG_MAIN,
            justify="left")
        note_pfaa.grid(row=ridx - 2, column=7, rowspan=2, sticky="nw",
                       padx=(10, 0))

        gap = tk.Frame(tbl, bg=BG_MAIN, height=10)
        gap.grid(row=ridx, column=0, columnspan=8)
        gap._toggle_kind = "grid"
        self._detailed_only_frames.append(gap)
        ridx += 1
        for r in precursor_rows:
            _add_row(ridx, *r, detailed_only=True)
            ridx += 1
        note_pre = tk.Label(tbl,
            text="(All 11 source concentrations for Precursor-1\n"
                 "are changed by the same amount during\n"
                 "calibration runs).",
            font=FONT_LABEL_SMI, fg=FG_HELP, bg=BG_MAIN,
            justify="left")
        note_pre.grid(row=ridx - 2, column=7, rowspan=2, sticky="nw",
                      padx=(10, 0))
        note_pre._toggle_kind = "grid"
        self._detailed_only_frames.append(note_pre)

        actwrap = tk.Frame(outer, bg=BG_MAIN)
        actwrap.pack(pady=(16, 4))

        actrow1 = tk.Frame(actwrap, bg=BG_MAIN); actrow1.pack(pady=3)
        make_btn(actrow1, "1. Save Calibration Data",
                 "Save_Data_Calibration",
                 fg=FG_BTN_NAVY, font=FONT_BTN_CALIB, padx=12, pady=10,
                 bg=BTN_FILL, width=24).pack(side="left", padx=10)
        make_btn(actrow1, "2. Run Machine Based Calibration",
                 "Run_Machine_Based_Calibration",
                 fg=FG_BTN_RED, font=FONT_BTN_CALIB, padx=12, pady=10,
                 bg=BTN_FILL, width=30).pack(side="left", padx=10)
        make_btn(actrow1, "Go Back to Main\nInterface",
                 "Show_MainInterface",
                 fg=FG_BTN_NAVY, font=FONT_BTN_CALIB, padx=12, pady=10,
                 bg=BTN_FILL, width=18).pack(side="left", padx=10)

        actrow2 = tk.Frame(actwrap, bg=BG_MAIN); actrow2.pack(pady=3)
        make_btn(actrow2, "3. See All Calibration Data",
                 "See_Calibration_Data",
                 fg=FG_BTN_NAVY, font=FONT_BTN_CALIB, padx=12, pady=10,
                 bg=BTN_FILL, width=24).pack(side="left", padx=10)
        make_btn(actrow2, "4. Load Optimal\nData",
                 "Load_Optimal_Data",
                 fg=FG_BTN_NAVY, font=FONT_BTN_CALIB, padx=12, pady=10,
                 bg=BTN_FILL, width=16).pack(side="left", padx=10)
        make_btn(actrow2, "5. Run Optimal\nModel",
                 "Run_Optimal_Model",
                 fg=FG_BTN_NAVY, font=FONT_BTN_CALIB, padx=12, pady=10,
                 bg=BTN_FILL, width=16).pack(side="left", padx=10)
        make_btn(actrow2, "Help",
                 "Help_Calibration",
                 fg=FG_BTN_NAVY, font=FONT_BTN_CALIB, padx=12, pady=10,
                 bg=BTN_FILL, width=14).pack(side="left", padx=10)

        actrow3 = tk.Frame(actwrap, bg=BG_MAIN); actrow3.pack(pady=3)
        make_btn(actrow3, "6. Save Optimal Model",
                 "Save_Optimal_Model",
                 fg=FG_BTN_NAVY, font=FONT_BTN_CALIB, padx=12, pady=10,
                 bg=BTN_FILL, width=24).pack(padx=10)

        if self.active_sheet != "Detailed_2":
            for w in (det_lbl1, det_chk1, det_lbl2, det_chk2,
                      gap, note_pre):
                try: w.grid_remove()
                except Exception: pass
            for w in self._detailed_only_frames:
                try:
                    if w.master is tbl:
                        w.grid_remove()
                except Exception: pass


    def _build_bottom_bar(self, parent):
        return None


# ENTRY POINT
if __name__ == "__main__":
    app = REMFluorApp()
    app.mainloop()
