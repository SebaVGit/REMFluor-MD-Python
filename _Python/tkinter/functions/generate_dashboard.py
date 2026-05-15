#!/usr/bin/env python
# generate_dashboard.py
#
# PURPOSE
# generate dashboard based on the results 
#
# PROJECT
# 10213
#
# NOTES
#   1. 
#
# AUTHOR(S)
# 
#
# HISTORY
#  Date      Remarks
# ---------- -------------------------------------------------------------
# 2025-09-07 
# ==========================================================================

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from dash import Dash, dcc, html, Input, Output, State, callback
import webbrowser
from datetime import datetime
import numpy as np
import glob
import os
import time
import sys
import json
import signal
import threading
import re

# v86: openpyxl/win32com/xlwings imports removed.  The dashboard reads
# cell values from dashboard_state.json (written by run_model.py with
# hard-coded fallback defaults so no None ever leaks through) instead
# of opening the .xlsm Storyboard workbook.  This makes the dashboard
# fully standalone — no Excel installation required, works in any
# folder layout, no .xlsm dependency.
#
# The optional calibration .xlsx (Model Data sheet, observed PFAS
# time-series) is still read via openpyxl with a deferred import —
# its absence just disables observed-data overlays.

# v86: state accessors backed by dashboard_state.json (no openpyxl ws).
# `_STATE` is loaded by main() from the JSON dump that run_model.py
# writes before launching this subprocess.  The legacy (ws, ref) call
# signatures are preserved so all ~25 dashboard call sites work
# unchanged — `ws` is now a no-op placeholder.
_STATE = {}


def _to_number(v):
    """Coerce a state value to a number where possible."""
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if s == "" or s.lower() == "none":
            return None
        try:
            f = float(s)
            return int(f) if f.is_integer() else f
        except (TypeError, ValueError):
            return v
    return v


def get_cell_value(ws, cell_ref):
    """Read a cell value from the JSON state dict (ws ignored)."""
    if not cell_ref:
        return None
    return _to_number(_STATE.get(cell_ref.upper()))


def get_range_values(ws, range_ref):
    """Read a precomputed range list from the JSON state dict.

    run_model._dump_dashboard_state pre-flattens each Excel range under
    its range string key (e.g. "V8:V18" -> [c1,...,c11]).  Tolerates
    +/-1 row mismatches at the call site (V34:V41 vs dump V34:V40).
    """
    if not range_ref:
        return None
    raw = _STATE.get(range_ref.upper())
    if raw is None:
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_ref.upper())
        if m:
            sc, sr, ec, er = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            for er_alt in (er - 1, er + 1):
                key = f"{sc}{sr}:{ec}{er_alt}"
                raw = _STATE.get(key)
                if raw is not None:
                    break
    if raw is None:
        return None
    return [_to_number(x) for x in raw]

# Read calibration_inputs.txt to get Excel file path
def read_calibration_inputs(filepath):
    """Read calibration_inputs.txt and extract Excel file path.

    Returns None silently when the sidecar is missing — the dashboard
    can still launch with simulated-only plots."""
    if not os.path.exists(filepath):
        return None
    excel_path = None
    try:
        with open(filepath, "r") as f:
            for line in f:
                if "Excel File Path:" in line:
                    excel_path = line.split("Excel File Path:")[1].strip()
                    excel_path = os.path.normpath(excel_path)
                    break
    except Exception:
        return None
    return excel_path

def read_model_data_sheet(excel_app, calibration_file_path):
    """Read the Model Data sheet from a calibration .xlsx into a DataFrame.

    v86: openpyxl-based — no Excel COM dependency.  excel_app is ignored
    (kept in signature for backward compat).  Returns columns
    [date, analyte, concentration, unit, well] or None.
    """
    if not calibration_file_path or not os.path.exists(calibration_file_path):
        return None
    try:
        from openpyxl import load_workbook  # deferred optional dep
    except ImportError:
        print("Warning: openpyxl not available - calibration data skipped")
        return None

    try:
        wb_cal = load_workbook(calibration_file_path, data_only=True)
        ws_model = None
        for nm in ("Model Data", "ModelData"):
            if nm in wb_cal.sheetnames:
                ws_model = wb_cal[nm]
                break
        if ws_model is None:
            wb_cal.close()
            return None

        header_row = None
        for row in range(1, 20):
            ca = ws_model.cell(row=row, column=1).value
            cc = ws_model.cell(row=row, column=3).value
            if ca is not None and cc is not None:
                a_s = str(ca).strip().lower()
                c_s = str(cc).strip().lower()
                if ("date" in a_s or "time" in a_s) and ("analyte" in c_s or "compound" in c_s):
                    header_row = row
                    break
        if header_row is None:
            header_row = 1

        max_row = ws_model.max_row or 1000
        data_rows = []
        for row in range(header_row + 1, min(header_row + 1000, max_row + 1)):
            date_val    = ws_model.cell(row=row, column=1).value
            well_val    = ws_model.cell(row=row, column=2).value
            analyte_val = ws_model.cell(row=row, column=3).value
            conc_val    = ws_model.cell(row=row, column=4).value
            unit_val    = ws_model.cell(row=row, column=5).value
            if date_val is None and analyte_val is None and conc_val is None:
                break
            if date_val is not None and analyte_val is not None and conc_val is not None:
                data_rows.append({
                    "date": date_val,
                    "analyte": str(analyte_val).strip(),
                    "concentration": conc_val,
                    "unit": str(unit_val).strip() if unit_val else "",
                    "well": str(well_val).strip() if well_val else None,
                })
        wb_cal.close()

        if not data_rows:
            return None
        for r in data_rows:
            d = r["date"]
            if hasattr(d, "replace") and hasattr(d, "tzinfo"):
                try:
                    r["date"] = d.replace(tzinfo=None)
                except Exception:
                    pass
        return pd.DataFrame(data_rows)
    except Exception as e:
        print(f"Error reading Model Data sheet: {e}")
        return None

def find_closest_date(target_date, date_array):
    """Find the closest date in date_array to target_date"""
    if len(date_array) == 0:
        return None
    date_diffs = np.abs(date_array - target_date)
    closest_idx = np.argmin(date_diffs)
    return date_array[closest_idx]

def calculate_rmsle(observed, predicted):
    """Calculate Root Mean Squared Log Error"""
    # Filter out zeros and negative values
    mask = (observed > 0) & (predicted > 0)
    if mask.sum() == 0:
        return None
    
    obs_filtered = observed[mask]
    pred_filtered = predicted[mask]
    
    # Calculate log errors
    log_errors = np.log1p(obs_filtered) - np.log1p(pred_filtered)
    
    # Calculate RMSLE
    rmsle = np.sqrt(np.mean(log_errors ** 2))
    return rmsle

def main(workbook_path=None, sheet_name=None):
    # v86: locate working directory.  argv[1] is still passed for
    # backwards compat with run_model.py's existing contract — we use
    # its parent dir as cwd so all .out files / dashboard_state.json
    # resolve relatively.
    if workbook_path is None:
        if len(sys.argv) > 1:
            workbook_path = sys.argv[1]
            if len(sys.argv) > 2:
                sheet_name = sys.argv[2]

    if workbook_path and os.path.exists(workbook_path):
        workbook_dir = os.path.dirname(os.path.abspath(workbook_path))
    elif workbook_path:
        parent = os.path.dirname(os.path.abspath(workbook_path))
        workbook_dir = parent if os.path.isdir(parent) else os.getcwd()
    else:
        workbook_dir = os.getcwd()
    os.chdir(workbook_dir)

    # Load JSON state dump (zero .xlsm dependency).  run_model.py
    # writes this with hard-coded defaults for every cell, so even
    # if the user's tkinter form is blank, _STATE has sensible values.
    global _STATE
    _STATE = {}
    state_path = os.path.join(workbook_dir, "dashboard_state.json")
    try:
        with open(state_path, "r", encoding="utf-8") as fh:
            _STATE = json.load(fh) or {}
    except Exception as e:
        print(f"Warning: could not load dashboard_state.json ({e}); "
              f"using built-in defaults")

    # `ws` is no longer an openpyxl worksheet — kept for legacy call sites.
    ws = None
    if not sheet_name:
        sheet_name = "Simple"

    # Read initial values from JSON state dict (via legacy helpers).
    # These have defaults baked in by run_model._dump_dashboard_state,
    # so they are guaranteed non-None.
    version_flag = get_cell_value(ws, "A8") or 1   # 1=simple, 2=detailed
    unit_flag = get_cell_value(ws, "AD1") or 1     # 1=ft, 2=m
    if unit_flag == 1:
        convert_to_meters = True
    else:
        convert_to_meters = False
    feet_to_meters = 0.3048  # Conversion factor: 1 foot = 0.3048 meters
    show_in_feet = True if unit_flag == 1 else False
    meters_to_feet = (1.0 / feet_to_meters) if feet_to_meters else 1.0
    length_display_scale = meters_to_feet if show_in_feet else 1.0
    length_unit_label = "ft" if show_in_feet else "m"
    
    k38_value = get_cell_value(ws, "K38")
    if k38_value is None or (isinstance(k38_value, str) and not str(k38_value).strip()) or k38_value == 'None':
        ipre = 0
    else:
        ipre = 1

    precursor_flag = True if version_flag == 2 and ipre ==1 else False  # True if version_flag is 2, otherwise blank

    # Get See Results Every Values
    # Prefer input.inp (always rewritten by the Run Model pipeline)
    # over the workbook (which is static and doesn't reflect §11 cell
    # changes the user made in the standalone tkinter app).
    # input.inp's "dt (yr), nt, npt" line gives see_every = npt × dt.
    See_Results_Every = None
    try:
        with open("input.inp") as _fp:
            _lines = _fp.readlines()
        for _i, _ln in enumerate(_lines):
            if "dt (yr)" in _ln and "nt" in _ln and "npt" in _ln:
                _parts = [p.strip() for p
                          in _lines[_i + 1].split(",") if p.strip()]
                if len(_parts) >= 3:
                    _dt  = float(_parts[0])
                    _npt = int(float(_parts[2]))
                    See_Results_Every = _npt * _dt
                break
    except Exception:
        pass
    if See_Results_Every is None:
        # Fallback to workbook if input.inp can't be parsed.
        See_Results_Every = get_cell_value(ws, "V47")

    # v90: calibration .xlsx is loaded for BOTH Simple and Detailed
    # versions.  The legacy code gated it behind `if version_flag != 1`
    # so users in Simple mode lost their §10 calibration overlay even
    # after using the import button.  Now whichever version + button
    # combination the user chose, the dashboard tries to load it.
    excel = None  # legacy variable, kept for read_model_data_sheet sig
    calibration_file = read_calibration_inputs("calibration_inputs.txt")

    # MW_names: pulled from state dump (§10 first column, U34:U40),
    # which run_model.py guarantees has placeholder names ("Well 1",
    # ...) if the user's form was blank.
    MW_names = []
    mw_raw = get_range_values(ws, "U34:U40") or []
    for v in mw_raw:
        if v is not None and str(v).strip() and str(v).strip().lower() != "none":
            MW_names.append(str(v).strip())
        
    PFAS_names = [
        str(get_cell_value(ws, "K38") or ""), #Precursor 1
        str(get_cell_value(ws, "E38") or ""), #PFAA1
        str(get_cell_value(ws, "M38") or ""), #Precursor 2
        str(get_cell_value(ws, "G38") or "") #PFAA2
    ]
    def is_valid_pfas_name(name):
        if name is None:
            return False
        name_str = str(name).strip()
        return name_str != "" and name_str.lower() != "none"

    # v86: defensive defaults — though run_model.py provides values,
    # we still defend in case of truly missing JSON or unexpected None.
    def _scalar(addr, default):
        v = get_cell_value(ws, addr)
        if v is None or not isinstance(v, (int, float)):
            return default
        return v
    StartY = _scalar("E18", 2025)
    Total_Depth = _scalar("E13", 10.0)
    # v102: REMFluor-MD.out's Z column is in METRES (Fortran works in
    # metres regardless of the UI unit).  Total_Depth comes from state
    # E13 which is in the user's unit (ft or m).  Convert to metres
    # here so the `Total_Depth - df_obs['Z']` arithmetic below is
    # unit-consistent.  Without this, feet-mode users saw depth values
    # multiplied effectively twice (once mismatched, then once on
    # display) — e.g. 75-108 ft for a 10 m / 33 ft model.
    if unit_flag == 1:
        Total_Depth = Total_Depth * feet_to_meters   # ft → m

    # Check G38 to determine ncomp: if blank, ncomp = 1, otherwise ncomp = 2
    g38_value = get_cell_value(ws, "G38")
    if g38_value is None or (isinstance(g38_value, str) and not str(g38_value).strip()) or g38_value == 'None':
        ncomp = 1  # G38 is blank
    else:
        ncomp = 2  # G38 has a value
    


    # v89: §10 monitoring-well observed concentrations are read for
    # BOTH Simple and Detailed versions.  The legacy code gated this
    # behind `if version_flag == 1`, leaving Detailed runs without
    # observed-data overlays unless a separate calibration .xlsx was
    # provided.  Now V34:V40 (PFAA-1) and X34:X40 (PFAA-2) flow through
    # for both versions, with the sample year coming from R36 (Simple)
    # or Y74 (Detailed) — falling back to StartY if blank.
    excel_PFAA1_data = {}  # {well_name: concentration}
    excel_PFAA2_data = {}  # {well_name: concentration} — used if ncomp == 2
    excel_data_year = None

    try:
        # v101: Y74 is the authoritative source — v_sample_yr only
        # maps to Y74 in state.CELL_MAP (R36 is never populated by
        # state.snapshot, so it always defaulted to start_year and
        # made obs markers appear at the simulation start year
        # instead of the user's sample year).  Try Y74 first, fall
        # back to R36 only if Y74 is somehow blank, and StartY only
        # as the final fallback.
        sy_val = get_cell_value(ws, "Y74")
        if sy_val is None or (isinstance(sy_val, str) and not str(sy_val).strip()):
            sy_val = get_cell_value(ws, "R36")
        if sy_val is not None:
            try:
                excel_data_year = float(sy_val)
            except (ValueError, TypeError):
                pass
        if excel_data_year is None:
            excel_data_year = float(StartY)
        # Diagnostic log so the user can confirm what value the
        # dashboard sees (visible in the dashboard log file).
        try:
            print(f"[dashboard] excel_data_year (sample year) = "
                  f"{excel_data_year} (Y74={get_cell_value(ws, 'Y74')!r}, "
                  f"R36={get_cell_value(ws, 'R36')!r})")
        except Exception:
            pass

        PFAA1_range = get_range_values(ws, "V34:V40") or []
        PFAA2_range = get_range_values(ws, "X34:X40") if ncomp == 2 else []

        def _flatten(rng):
            out = []
            for item in rng or []:
                if isinstance(item, (list, tuple)):
                    out.append(item[0] if len(item) > 0 else None)
                else:
                    out.append(item)
            return out

        PFAA1_values = _flatten(PFAA1_range)
        PFAA2_values = _flatten(PFAA2_range)

        for idx in range(min(len(MW_names), 8)):
            well = MW_names[idx]
            if idx < len(PFAA1_values) and PFAA1_values[idx] is not None:
                try:
                    v = float(PFAA1_values[idx])
                    if v > 0:
                        excel_PFAA1_data[well] = v
                except (ValueError, TypeError):
                    pass
            if ncomp == 2 and idx < len(PFAA2_values) and PFAA2_values[idx] is not None:
                try:
                    v = float(PFAA2_values[idx])
                    if v > 0:
                        excel_PFAA2_data[well] = v
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"Warning: Could not read §10 observed concentrations: {e}")

    # Read Model Data sheet from calibration file for RMSLE calculation
    df_model_data = None
    if calibration_file and os.path.exists(calibration_file):
        df_model_data = read_model_data_sheet(excel, calibration_file)
    # Load data
    df_obs = pd.read_csv("REMFluor-MD.out")
    df_obs['Z'] = Total_Depth - df_obs['Z'] # convert to Depth
    # Find all files matching obs_well*.outl
    obs_well_files = sorted(glob.glob("obs_well*.out"))

    # v86: pad MW_names so an empty/short list never causes
    # IndexError when the Fortran solver wrote N obs_well*.out files.
    while len(MW_names) < len(obs_well_files):
        MW_names.append(f"Well {len(MW_names) + 1}")

    # Read and combine all observation well files into one DataFrame.
    # v102: hardened reader + diagnostics.  The Fortran solver writes
    # the column header ONLY in obs_well1.out — files 2..7 start
    # straight at data.  We capture the header columns from file 1 and
    # reuse them for the rest.  skipinitialspace=True normalises the
    # multi-space alignment the solver uses ("   0.100,  0.11522...").
    CANON_COLS = ['Time yr', 'C1well', 'C2well', 'C3well', 'C4well']
    dfs = []
    captured_cols = None
    for idx, fname in enumerate(obs_well_files):
        try:
            if idx == 0:
                df = pd.read_csv(fname, skipinitialspace=True)
                # Strip leading/trailing whitespace from header names
                df.columns = [str(c).strip() for c in df.columns]
                captured_cols = list(df.columns)
            else:
                use_cols = captured_cols if captured_cols else CANON_COLS
                df = pd.read_csv(fname, header=None, names=use_cols,
                                 skipinitialspace=True)
                df.columns = [str(c).strip() for c in df.columns]
            df['well_file'] = MW_names[idx]
            dfs.append(df)
            # Diagnostic: per-file row count + C2well max so user can
            # compare to the .out at a glance.
            try:
                c2_max = (df['C2well'].astype(float).max()
                          if 'C2well' in df.columns else float('nan'))
                print(f"[obs_well] {fname} -> {MW_names[idx]!r}: "
                      f"{len(df)} rows, C2well max={c2_max:.6g}")
            except Exception:
                pass
        except Exception as exc:
            print(f"[obs_well] FAILED to read {fname}: {exc}")

    df_obs_MW = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    # Clean column names by removing all spaces (defence against any
    # remaining whitespace that slipped past read_csv).
    df_obs_MW.columns = df_obs_MW.columns.str.strip()
    # If the captured header had non-canonical names (e.g. solver wrote
    # different labels in some build), force the canonical names.  This
    # guarantees downstream code that references "C2well" etc. works.
    if len(df_obs_MW.columns) >= 5:
        rename_map = {}
        for old_col, canon in zip(df_obs_MW.columns[:5], CANON_COLS):
            if old_col != canon:
                rename_map[old_col] = canon
        if rename_map:
            print(f"[obs_well] renaming columns: {rename_map}")
            df_obs_MW = df_obs_MW.rename(columns=rename_map)
    print(f"[obs_well] combined frame: {len(df_obs_MW)} rows, "
          f"columns={list(df_obs_MW.columns)}")
    df_obs_MassD = pd.read_csv("discharge.out") # mass discharge (kg/yr)
    # Clean column names by removing all spaces
    df_obs_MassD.columns = df_obs_MassD.columns.str.strip()
    df_obs_MassD.columns = ['Time', 'X', 'Discharge1', 'Discharge2', 'Discharge3', 'Discharge4']
    df_obs_D = pd.read_csv("plume_mass.out") # discharge (kg)
    # Clean column names by removing all spaces
    df_obs_D.columns = df_obs_D.columns.str.strip()

    # v90: RMSLE subtitle was removed — user sees per-analyte RMSLE
    # in the Observation Wells analysis tab instead.

    # Define options (filter time dropdown to every See_Results_Every years)
    time_options_full = np.unique(np.round(df_obs[' Time'].unique() + StartY, 0))
    if See_Results_Every and See_Results_Every > 0:
        time_options = time_options_full[(time_options_full - StartY) % See_Results_Every == 0]
        if len(time_options) == 0:
            time_options = time_options_full
    else:
        time_options = time_options_full

    # Restrict observation data to times in the dropdown (time_options)
    time_calendar = (df_obs[' Time'] + StartY)
    df_obs = df_obs[time_calendar.isin(time_options)]
    df_obs_MassD = df_obs_MassD[(df_obs_MassD['Time'] + StartY).isin(time_options)]
    df_obs_D = df_obs_D[(df_obs_D['Time'] + StartY).isin(time_options)]
    #df_obs_MW = df_obs_MW[(df_obs_MW['Time yr'] + StartY).isin(time_options)]

    x_options = df_obs['X'].unique()
    y_options = df_obs['Y'].unique()
    z_options = df_obs['Z'].unique()

    # Define options for Mass Discharge analysis
    x_options_massd = df_obs_MassD['X'].unique() if 'X' in df_obs_MassD.columns else []

    # Calculate Source Mass from the spreadsheet (v86: defensive defaults)
    Source_Width = _scalar("E15", 10.0)
    Source_Thick = _scalar("E16", 5.0)
    Velocity     = _scalar("E22", 1.0)
    EndYear      = _scalar("E19", StartY + 100)
    SL_Time_raw = get_range_values(ws, "U8:U18")
    SL_Time = [v if not isinstance(v, list) else v[0] for v in SL_Time_raw] if SL_Time_raw else []
    SL_Time = [t for t in SL_Time if t is not None]
    SL_Time_org = list(SL_Time)
    SL_Time.append(EndYear)
    if len(SL_Time) >= 2:
        Time_diff = np.diff(SL_Time)
    else:
        Time_diff = np.array([0.0])

    # calculate Conc (ug/L)*yr*m*m*m/yr
    def calc_SL_conc(rng):
        range_vals = get_range_values(ws, rng)
        if range_vals:
            vals = np.array([v if not isinstance(v, list) else v[0] for v in range_vals])
        else:
            vals = np.array([])
        return vals, np.insert(np.cumsum(vals * Time_diff * Source_Width * Source_Thick * Velocity * 1000) / 1e9, 0, 0)

    # Source Loading for PFAA1
    SL_Conc_PFAA1, SL_Mass_PFAA1      = calc_SL_conc("V8:V18")

    # Only calculate PFAA2 if ncomp == 2
    if ncomp == 2:
        SL_Conc_PFAA2, SL_Mass_PFAA2      = calc_SL_conc("X8:X18")
    else:
        SL_Conc_PFAA2 = [0] * 11
        SL_Mass_PFAA2 = [0] * 11

    if precursor_flag == True:
        SL_Conc_Precursor1, SL_Mass_Precursor1 = calc_SL_conc("Z8:Z18")
        if ncomp ==2:
            SL_Conc_Precursor2, SL_Mass_Precursor2 = calc_SL_conc("AB8:AB18")
        else:
            SL_Conc_Precursor2 = [0] * 11
    else:
        SL_Conc_Precursor1 = [0] * 11   
        SL_Conc_Precursor2 = [0] * 11
        SL_Mass_Precursor1 = [0] * 11
        SL_Mass_Precursor2 = [0] * 11


    # App init with external stylesheets
    app = Dash(__name__, external_stylesheets=[
        'https://codepen.io/chriddyp/pen/bWLwgP.css',
        'https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css'
    ])

    # Custom CSS for modern styling
    app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
            
            * {
                font-family: 'Inter', sans-serif;
            }
            
            body {
                margin: 0;
                padding: 0;
                background: linear-gradient(135deg, #1C6EAC 0%, #5798C1 100%);
                min-height: 100vh;
            }
            
            .main-container {
                background: linear-gradient(135deg, #1C6EAC 0%, #5798C1 100%);
                min-height: 100vh;
                padding: 20px;
            }
            
            .dashboard-header {
                background: rgba(255, 255, 255, 0.95);
                backdrop-filter: blur(20px);
                border-radius: 25px;
                padding: 40px;
                margin-bottom: 30px;
                box-shadow: 0 25px 50px rgba(0, 0, 0, 0.15);
                border: 1px solid rgba(255, 255, 255, 0.3);
                text-align: center;
            }
            
            .dashboard-title {
                font-size: 3.5rem;
                font-weight: 800;
                background: linear-gradient(135deg, #1C6EAC, #5798C1);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                margin: 0;
                text-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
            }
            
            .dashboard-subtitle {
                text-align: center;
                color: #666;
                margin-top: 15px;
                font-size: 1.3rem;
                font-weight: 400;
            }
            
            .content-wrapper {
                display: flex;
                gap: 30px;
                align-items: flex-start;
            }
            
            .control-panel {
                background: rgba(255, 255, 255, 0.95);
                backdrop-filter: blur(20px);
                border-radius: 25px;
                padding: 35px;
                box-shadow: 0 25px 50px rgba(0, 0, 0, 0.15);
                border: 1px solid rgba(255, 255, 255, 0.3);
                width: 350px;
                flex-shrink: 0;
            }
            
            .control-title {
                font-size: 1.5rem;
                font-weight: 700;
                color: #333;
                margin-bottom: 25px;
                display: flex;
                align-items: center;
                gap: 12px;
                background: linear-gradient(135deg, #1C6EAC, #5798C1);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
            }
            
            .control-group {
                margin-bottom: 30px;
            }
            
            .control-label {
                font-weight: 600;
                color: #555;
                margin-bottom: 10px;
                display: block;
                font-size: 1rem;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }
            
            .dropdown-container {
                position: relative;
            }
            
            .dropdown-container .Select-control {
                border-radius: 15px !important;
                border: 2px solid #e1e5e9 !important;
                box-shadow: 0 8px 15px rgba(0, 0, 0, 0.08) !important;
                transition: all 0.3s ease !important;
                background: white !important;
                min-height: 50px !important;
            }
            
            .dropdown-container .Select-control:hover {
                border-color: #1C6EAC !important;
                box-shadow: 0 12px 25px rgba(28, 110, 172, 0.15) !important;
            }
            
            .dropdown-container .Select-control.is-focused {
                border-color: #1C6EAC !important;
                box-shadow: 0 0 0 4px rgba(28, 110, 172, 0.1) !important;
            }
            
            .dropdown-container .Select-value {
                font-weight: 500 !important;
                color: #333 !important;
            }
            
            .plot-container {
                background: rgba(255, 255, 255, 0.95);
                backdrop-filter: blur(20px);
                border-radius: 25px;
                padding: 35px;
                box-shadow: 0 25px 50px rgba(0, 0, 0, 0.15);
                border: 1px solid rgba(255, 255, 255, 0.3);
                flex: 1;
                min-height: 600px;
                max-width: 1000px;!calc(100% - 380px);
                overflow: hidden;
            }
            
            .plot-title {
                font-size: 1.6rem;
                font-weight: 700;
                color: #333;
                margin-bottom: 25px;
                text-align: center;
                background: linear-gradient(135deg, #1C6EAC, #5798C1);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
            }
            
            .stats-container {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                margin-bottom: 30px;
            }
            
            .stat-card {
                background: linear-gradient(135deg, #1C6EAC, #5798C1);
                color: white;
                padding: 25px;
                border-radius: 20px;
                text-align: center;
                box-shadow: 0 15px 30px rgba(28, 110, 172, 0.3);
                transition: all 0.3s ease;
                position: relative;
                overflow: hidden;
            }
            
            .stat-card:hover {
                transform: translateY(-5px);
                box-shadow: 0 20px 40px rgba(28, 110, 172, 0.4);
            }
            
            .stat-value {
                font-size: 2.5rem;
                font-weight: 800;
                margin-bottom: 8px;
                text-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }
            
            .stat-label {
                font-size: 1rem;
                opacity: 0.9;
                font-weight: 500;
            }
            
            .loading-container {
                display: flex;
                justify-content: center;
                align-items: center;
                height: 400px;
                flex-direction: column;
                gap: 20px;
            }
            
            .spinner {
                width: 60px;
                height: 60px;
                border: 4px solid rgba(28, 110, 172, 0.1);
                border-top: 4px solid #1C6EAC;
                border-radius: 50%;
                animation: spin 1s linear infinite;
            }
            
            @keyframes spin {
                0% { transform: rotate(0deg); }
                100% { transform: rotate(360deg); }
            }
            
            .loading-text {
                color: #1C6EAC;
                font-weight: 600;
                font-size: 1.1rem;
            }
            
            .footer {
                text-align: center;
                margin-top: 40px;
                color: rgba(255, 255, 255, 0.8);
                font-size: 1rem;
                font-weight: 500;
            }
            
            .icon {
                font-size: 1.2em;
                margin-right: 8px;
            }
            
            .pulse {
                animation: pulse 2s infinite;
            }
            
            @keyframes pulse {
                0% { transform: scale(1); }
                50% { transform: scale(1.05); }
                100% { transform: scale(1); }
            }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
    '''

    app.layout = html.Div([
        html.Div([
            html.Div([
                html.H1("REMFluor-MD Output", className="dashboard-title"),
            ], className="dashboard-header"),
        
        html.Div([
            html.Div([
                html.H3("🎛️ Control Panel", className="control-title", style={"fontSize": "3.0rem"}),
                
                html.Div([
                    html.Label("⏰ Time", className="control-label", style={"fontSize": "2.0rem"}),
                    html.Div([
                        dcc.Dropdown(
                            options=[{"label": f"{t:.0f} yr", "value": t} for t in time_options],
                            value=time_options[0], 
                            id="time-dropdown",
                            clearable=False
                        )
                    ], className="dropdown-container")
                ], className="control-group", id="time-selection"),
                
                html.Div([
                    html.Label("📍 X Coordinate", className="control-label", style={"fontSize": "2.0rem"}),
                    html.Div([
                        dcc.Dropdown(
                            options=[{"label": f"{x * length_display_scale:.1f} {length_unit_label}", "value": x} for x in x_options],
                            value=x_options[0], 
                            id="x-dropdown",
                            clearable=False
                        )
                    ], className="dropdown-container")
                ], className="control-group", id="x-selection"),
                
                html.Div([
                    html.Label("📍 Y Coordinate", className="control-label", style={"fontSize": "2.0rem"}),
                    html.Div([
                        dcc.Dropdown(
                            options=[{"label": f"{y * length_display_scale:.1f} {length_unit_label}", "value": y} for y in y_options],
                            value=y_options[0], 
                            id="y-dropdown",
                            clearable=False
                        )
                    ], className="dropdown-container")
                ], className="control-group", id="y-selection"),
                
                html.Div([
                    html.Label("📍 Z Coordinate", className="control-label", style={"fontSize": "2.0rem"}),
                    html.Div([
                        dcc.Dropdown(
                            options=[{"label": f"{z * length_display_scale:.1f} {length_unit_label}", "value": z} for z in z_options],
                            value=z_options[-1], 
                            id="z-dropdown",
                            clearable=False
                        )
                    ], className="dropdown-container")
                ], className="control-group", id="z-selection"),
                
                html.Div([
                    html.Label("📊 Analysis Type", className="control-label", style={"fontSize": "2.0rem"}),
                    html.Div([
                        dcc.Dropdown(
                            options=[
                                {"label": "Concentration vs Depth", "value": 1},
                                {"label": "Concentration vs Time", "value": 2},
                                {"label": "Concentration vs X Distance", "value": 3},
                                {"label": "Concentration vs Time at Source", "value": 4},
                                {"label": "Concentration vs Time at Observation Wells", "value": 5},
                                {"label": "Mass Dicharge vs Time in T-Zone", "value": 6},
                                {"label": "Mass vs Time in T-Zone", "value": 7},
                                {"label": "Mass vs Time in Low-K Zone", "value": 8},
                                {"label": "Mass vs Time at Source", "value": 9}

                            ],
                            value=1,
                            id="analysis-type-dropdown",
                            clearable=False
                        )
                    ], className="dropdown-container")
                ], className="control-group"),
                
                html.Div([
                    html.Label("📐 Axis Scale", className="control-label", style={"fontSize": "2.0rem"}),
                    dcc.RadioItems(
                        options=[
                            {"label": " Linear", "value": "linear"},
                            {"label": " Log", "value": "log"},
                        ],
                        value="linear",
                        id="log-scale-radio",
                        inline=True,
                        style={"fontSize": "1.6rem", "marginTop": "8px"}
                    )
                ], className="control-group"),
                
                # Well selection dropdown (only shown when analysis_type = 5)
                html.Div([
                    html.Label("🏥 Observation Wells", className="control-label", style={"fontSize": "2.0rem"}),
                    html.Div([
                        dcc.Dropdown(
                            options=[{"label": well, "value": well} for well in df_obs_MW['well_file'].unique() if well is not None],
                            value=[df_obs_MW['well_file'].unique()[0]] if len(df_obs_MW['well_file'].unique()) > 0 else [],
                            id="well-dropdown",
                            multi=True,
                            clearable=False
                        )
                    ], className="dropdown-container")
                ], className="control-group", id="well-selection", style={"display": "none"}),
                
                # Concentration selection dropdown (only shown when analysis_type = 5)
                html.Div([
                    html.Label("🧪 Concentration Types", className="control-label", style={"fontSize": "2.0rem"}),
                    html.Div([
                        dcc.Dropdown(
                            options = (
                                ([{"label": PFAS_names[0], "value": "C1well"}] if precursor_flag and is_valid_pfas_name(PFAS_names[0]) else [])
                                + ([{"label": PFAS_names[1], "value": "C2well"}] if is_valid_pfas_name(PFAS_names[1]) else [])
                                + ([{"label": PFAS_names[2], "value": "C3well"}] if ncomp == 2 and precursor_flag and is_valid_pfas_name(PFAS_names[2]) else [])
                                + ([{"label": PFAS_names[3], "value": "C4well"}] if ncomp == 2 and is_valid_pfas_name(PFAS_names[3]) else [])
                            ),
                            # v101: always default to C2well (PFAA-1).
                            # Previously Detailed-with-precursor users
                            # got C1well (Precursor-1) by default — if
                            # the Fortran solver wrote 0s for the
                            # precursor column (e.g. transformation
                            # rates not configured), the chart looked
                            # empty even though obs_well*.out had real
                            # PFAA data in C2well/C4well.
                            value=["C2well"] if is_valid_pfas_name(PFAS_names[1]) else (
                                ["C1well"] if precursor_flag and is_valid_pfas_name(PFAS_names[0]) else []
                            ),
                            id="concentration-dropdown",
                            multi=True,
                            clearable=False
                        )
                    ], className="dropdown-container")
                ], className="control-group", id="concentration-selection", style={"display": "none"}),
                
                # X Coordinate selection for Mass Discharge (only shown when analysis_type = 6)
                html.Div([
                    html.Label("📍 X Coordinate (Mass Discharge)", className="control-label", style={"fontSize": "2.0rem"}),
                    html.Div([
                        dcc.Dropdown(
                            options=[{"label": f"{x * length_display_scale:.1f} {length_unit_label}", "value": x} for x in x_options_massd],
                            value=x_options_massd[0] if len(x_options_massd) > 0 else None, 
                            id="x-massd-dropdown",
                            clearable=False
                        )
                    ], className="dropdown-container")
                ], className="control-group", id="x-massd-selection", style={"display": "none"}),
            ], className="control-panel"),
            
            html.Div([
                #html.H3("📈 Data Visualization", className="plot-title"),
                dcc.Loading(
                    id="loading",
                    type="default",
                    children=html.Div(id="plot-container")
                )
            ], className="plot-container")
        ], className="content-wrapper"),

        html.Div([
            html.P(f"© 2025 REMFluor-MD Analysis System | Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                className="footer")
        ])
    ], className="main-container")
    ])



    # Callback to show/hide coordinate dropdowns
    @app.callback(
        [Output("time-selection", "style"), Output("x-selection", "style"), 
        Output("y-selection", "style"), Output("z-selection", "style"),
        Output("x-massd-selection", "style")],
        Input("analysis-type-dropdown", "value")
    )
    def toggle_coordinate_dropdowns(analysis_type):
        if analysis_type in [4,5,9]:
            return {"display": "none"}, {"display": "none"}, {"display": "none"}, {"display": "none"}, {"display": "none"}
        elif analysis_type == 6:
            return {"display": "none"}, {"display": "none"}, {"display": "none"}, {"display": "none"}, {"display": "block"}
        elif analysis_type in [7, 8]:
            return {"display": "none"}, {"display": "none"}, {"display": "none"}, {"display": "none"}, {"display": "none"}
        elif analysis_type == 1:
            # Hide Z only for depth plotting
            return {"display": "block"}, {"display": "block"}, {"display": "block"}, {"display": "none"}, {"display": "none"}
        elif analysis_type == 2:
            return {"display": "none"}, {"display": "block"}, {"display": "block"}, {"display": "block"}, {"display": "none"}
        elif analysis_type == 3:
            return {"display": "block"}, {"display": "none"}, {"display": "block"}, {"display": "block"}, {"display": "none"}
        else:
            return {"display": "block"}, {"display": "block"}, {"display": "block"}, {"display": "block"}, {"display": "none"}

    # Callback to show/hide well and concentration selection dropdowns
    @app.callback(
        [Output("well-selection", "style"), Output("concentration-selection", "style")],
        Input("analysis-type-dropdown", "value")
    )
    def toggle_well_concentration_dropdowns(analysis_type):
        if analysis_type == 5:
            return {"display": "block"}, {"display": "block"}
        else:
            return {"display": "none"}, {"display": "none"}
    
    @app.callback(
        [Output("concentration-dropdown", "options"), Output("concentration-dropdown", "value")],
        Input("analysis-type-dropdown", "value")
    )
    def update_concentration_dropdown_options(analysis_type):
        # Base options for all concentration types (C3well and C4well only when ncomp == 2)
        all_options = []
        if is_valid_pfas_name(PFAS_names[0]):
            all_options.append({"label": PFAS_names[0], "value": "C1well"})
        if is_valid_pfas_name(PFAS_names[1]):
            all_options.append({"label": PFAS_names[1], "value": "C2well"})
        if ncomp == 2 and is_valid_pfas_name(PFAS_names[2]):
            all_options.append({"label": PFAS_names[2], "value": "C3well"})
        if ncomp == 2 and is_valid_pfas_name(PFAS_names[3]):
            all_options.append({"label": PFAS_names[3], "value": "C4well"})
        
        # For analysis_type == 5 and version_flag == 1, only show C2well and C4well (if ncomp == 2)
        if analysis_type == 5 and version_flag == 1:
            filtered_options = []
            if is_valid_pfas_name(PFAS_names[1]):
                filtered_options.append({"label": PFAS_names[1], "value": "C2well"})
            if ncomp == 2 and is_valid_pfas_name(PFAS_names[3]):
                filtered_options.append({"label": PFAS_names[3], "value": "C4well"})
            default_value = ["C2well"] if any(opt["value"] == "C2well" for opt in filtered_options) else ([filtered_options[0]["value"]] if filtered_options else [])
            return filtered_options, default_value
        else:
            # v101: always prefer C2well (PFAA-1) as the default — it
            # is the column most likely to have non-zero data.  Falling
            # back to C1well when precursors are configured produced
            # an empty-looking chart on Detailed runs where the user
            # hadn't actually set up transformation rates and
            # obs_well*.out's C1well column came out as all zeros.
            preferred_value = "C2well"
            default_value = [preferred_value] if any(opt["value"] == preferred_value for opt in all_options) else ([all_options[0]["value"]] if all_options else [])
            return all_options, default_value

    @app.callback(
        Output("plot-container", "children"),
        Input("time-dropdown", "value"),
        Input("x-dropdown", "value"),
        Input("y-dropdown", "value"),
        Input("z-dropdown", "value"),
        Input("analysis-type-dropdown", "value"),
        Input("well-dropdown", "value"),
        Input("concentration-dropdown", "value"),
        Input("x-massd-dropdown", "value"),
        Input("log-scale-radio", "value"),
    )
    def update_plot(time_val, x_val, y_val, z_val, analysis_type, well_vals, conc_vals, x_massd_val, log_scale):
        # Filter data based on dropdown selections
        yaxis_title = "Concentration (μg/L)"
        depth_series = None
        rmsle_stats = []  # Initialize RMSLE stats for all analysis types
        if analysis_type == 1:
            df_filtered = df_obs[(df_obs[' Time'] == time_val-StartY) & (df_obs['X'] == x_val) & (df_obs['Y'] == y_val)]
            # Compute depth from total depth and Z
            try:
                depth_series = df_filtered['Z']
                if show_in_feet and depth_series is not None:
                    depth_series = depth_series * length_display_scale
            except Exception:
                depth_series = None
            x = None  # x will be each concentration series
            xlabel = "Concentration (μg/L)"
            yaxis_title = f"Depth ({length_unit_label})"
            title = "Concentration vs. Depth"
        elif analysis_type == 2:
            df_filtered = df_obs[(df_obs['X'] == x_val) & (df_obs['Y'] == y_val) & (df_obs['Z'] == z_val)]
            x = df_filtered[' Time']+StartY
            xlabel = "Time (yr)"
            title = "Concentration vs. Time in T-Zone"
        elif analysis_type == 3:
            df_filtered = df_obs[(df_obs[' Time'] == time_val-StartY) & (df_obs['Y'] == y_val) & (df_obs['Z'] == z_val)]
            x = df_filtered['X']
            if show_in_feet and x is not None:
                x = x * length_display_scale
            xlabel = f"Distance in X Direction ({length_unit_label})"
            title = "Concentration vs. X Distance in T-Zone "
        elif analysis_type == 4:
            xlabel = "Time (yr)"
            title = "Concentration vs. Time at Source"
        elif analysis_type == 5:
            # Filter by selected wells
            if well_vals and conc_vals:
                df_filtered = df_obs_MW[df_obs_MW['well_file'].isin(well_vals)]
                x = df_filtered['Time yr']+StartY
                xlabel = "Time (yr)"
                title = "Concentration vs. Time at Observation Wells"
            else:
                df_filtered = df_obs_MW
                x = df_filtered['Time yr']+StartY
                xlabel = "Time (yr)"
                title = "Concentration vs. Time at Observation Wells"
        elif analysis_type == 6:        
            df_filtered = df_obs_MassD[(df_obs_MassD['X'] == x_massd_val)]
            x = df_filtered['Time']+StartY
            xlabel = "Time (yr)"
            yaxis_title = "Mass Discharge (kg/yr)"
            title = "Mass Discharge vs. Time in T-Zone"
        elif analysis_type == 7:
            xlabel = "Time (yr)"
            yaxis_title = "Mass (kg)"
            title = "Mass vs. Time in T-Zone"
        elif analysis_type == 8:
            xlabel = "Time (yr)"
            yaxis_title = "Mass (kg)"
            title = "Mass vs. Time in Low-K Zone"
        elif analysis_type == 9:       # Filter by selected wells
            x = SL_Time 
            xlabel = "Time (yr)"
            yaxis_title = "Cumulative Mass (kg)"
            title = "Cumulative Mass vs. Time at Source"
        else:
            df_filtered = df_obs
            x = df_filtered[' Time']
            xlabel = "Time (yr)"
            title = "Default Plot"

        # Create a more sophisticated plot
        fig = go.Figure()
        
        # Color palette for different concentration series (colorblind-friendly)
        colors = ['#1C6EAC', '#D55E00', '#009E73', '#CC79A7', '#F0E442', '#E69F00', '#56B4E9', '#0072B2']
        
        if analysis_type == 1:
            # Plot concentration (x) vs depth (y)
            conc_cols = [c for c in df_filtered.columns if c.startswith("Conc")]
            # Filter out Precursor 1 and Precursor 2 if not precursor_flag
            if not precursor_flag:
                conc_cols = [c for c in conc_cols if c not in ["Conc1", "Conc3"]]
            # Filter out Conc4 if PFAS_names[3] is None
            if ncomp == 1 and precursor_flag==True:
                conc_cols = [c for c in conc_cols if c not in ["Conc3","Conc4"]]
            try:
                if PFAS_names[3] in ('None',''):
                    conc_cols = [c for c in conc_cols if c != "Conc4"]
            except (IndexError, KeyError):
                conc_cols = [c for c in conc_cols if c != "Conc4"]
            conc_to_pfas = {
                "Conc1": PFAS_names[0],
                "Conc2": PFAS_names[1]
            }
            try:
                if PFAS_names[2] not in ('None', ''):
                    conc_to_pfas["Conc3"] = PFAS_names[2]
            except (IndexError, KeyError):
                pass
            # Only include Conc4 if PFAS_names[3] is not None
            try:
                if PFAS_names[3] not in ('None', ''):
                    conc_to_pfas["Conc4"] = PFAS_names[3]
            except (IndexError, KeyError):
                pass
            for i, col in enumerate(conc_cols):
                display_name = conc_to_pfas.get(col, col)
                fig.add_trace(go.Scatter(
                    x=df_filtered[col],
                    y=depth_series if depth_series is not None else df_filtered['Z'],
                    mode='lines+markers',
                    name=display_name,
                    line=dict(
                        color=colors[i % len(colors)],
                        width=3,
                        shape='spline'
                    ),
                    marker=dict(
                        size=8,
                        color=colors[i % len(colors)],
                        line=dict(width=2, color=colors[i % len(colors)])
                    ),
                    hovertemplate=f'<b>{display_name}</b><br>' +
                                f'Concentration: %{{x:.3f}} μg/L<br>' +
                                # v102: depth unit follows §1 feet/meters
                                f'Depth: %{{y:.2f}} {length_unit_label}<br>' +
                                '<extra></extra>'
                ))
            # Set y-axis to have 0 at the top (reverse the axis)
            fig.update_yaxes(autorange='reversed')
        elif analysis_type == 4:
            color_idx = 0
            # Filter out Precursor 1 and Precursor 2 if not precursor_flag
            # Only include PFAA2 if ncomp == 2 and PFAA2_name is not None/empty
            conc_list = [SL_Conc_PFAA1]
            pfas_indices = [1]  # PFAA1 index
            if ncomp == 2:
                try:
                    if PFAS_names[3] not in ('None', ''):
                        conc_list.append(SL_Conc_PFAA2)
                        pfas_indices.append(3)  # PFAA2 index
                except (IndexError, KeyError):
                    pass
            if version_flag != 1 and PFAS_names[2] not in ('None', ''):
                conc_list.extend([SL_Conc_Precursor1, SL_Conc_Precursor2])
                pfas_indices.extend([0, 2])  # Precursor 1 and Precursor 2 indices
            if version_flag != 1 and PFAS_names[2] in ('None', ''):
                conc_list.extend([SL_Conc_Precursor1])
                pfas_indices.extend([0])  # Precursor 1 index
            
            for conc, pfas_idx in zip(conc_list, pfas_indices):
                pfas_name = PFAS_names[pfas_idx] if pfas_idx < len(PFAS_names) else None
                # Skip if pfas_name is None or empty (e.g., PFAA2_name when G38 is blank)
                if not pfas_name or (isinstance(pfas_name, str) and not pfas_name.strip()):
                    continue
                fig.add_trace(go.Scatter(
                    x=SL_Time_org,
                    y=conc,
                    mode='lines+markers',
                    name=f'{pfas_name}',
                    line=dict[str, str | int](
                        color=colors[color_idx % len(colors)],
                        width=3,
                        shape='linear'
                    ),
                    marker=dict(
                        size=8,
                        color=colors[color_idx % len(colors)],
                        line=dict(width=2, color=colors[color_idx % len(colors)])
                    ),
                    hovertemplate=f'<b>{pfas_name}</b><br>' +
                                    f'{xlabel}: %{{x}}<br>' +
                                    'Concentration: %{y:.3f} μg/L<br>' +
                                    '<extra></extra>'
                ))
                color_idx += 1
        elif analysis_type == 5:
            # Special handling for observation wells
            if not well_vals or not conc_vals:
                well_vals = df_obs_MW['well_file'].unique().tolist() if len(df_obs_MW) > 0 else []
                conc_vals = ["C1well", "C2well", "C3well"]
                if ncomp == 2:
                    conc_vals.append("C4well")
            
            if well_vals and conc_vals:
                color_idx = 0
                # Filter out Precursor 1 and Precursor 2 if not precursor_flag
                if not precursor_flag:
                    conc_vals = [c for c in conc_vals if c not in ["C1well", "C3well"]]
                # Filter out C4well if ncomp == 1
                if precursor_flag and PFAS_names[2] in ('None', ''):    
                    conc_vals = [c for c in conc_vals if c != "C3well"]
                if ncomp == 1:
                    conc_vals = [c for c in conc_vals if c != "C4well"]
                # Create mapping from column names to PFAS names
                conc_to_pfas = {
                    "C1well": PFAS_names[0],
                    "C2well": PFAS_names[1], 
                    "C3well": PFAS_names[2],
                    'C4well': PFAS_names[3],
                }
                
                # Calculate RMSLE if model data is available
                if df_model_data is not None and len(df_model_data) > 0:
                    # Create a local copy to avoid modifying the global variable
                    model_data_copy = df_model_data.copy()
                    
                    # Convert dates in model data to decimal years (year + fraction of year)
                    # Handle different date formats
                    date_year_list = []
                    for date_val in model_data_copy['date']:
                        if pd.isna(date_val):
                            date_year_list.append(np.nan)
                        elif isinstance(date_val, (pd.Timestamp, datetime)):
                            # Calculate decimal year: year + day_of_year / days_in_year
                            # Example: 2020-01-02 = 2020 + 2/365
                            year = date_val.year
                            day_of_year = date_val.timetuple().tm_yday
                            days_in_year = 366 if pd.Timestamp(year, 12, 31).timetuple().tm_yday == 366 else 365
                            decimal_year = year + day_of_year / days_in_year
                            date_year_list.append(float(decimal_year))
                        elif isinstance(date_val, (int, float)):
                            # Could be Excel serial date or year already
                            if date_val > 1900 and date_val < 2100:
                                # Likely already a decimal year, use as is
                                date_year_list.append(float(date_val))
                            else:
                                # Might be Excel serial date, convert to decimal year
                                try:
                                    excel_date = pd.Timestamp.fromordinal(int(date_val) + 693594)
                                    year = excel_date.year
                                    day_of_year = excel_date.timetuple().tm_yday
                                    days_in_year = 366 if pd.Timestamp(year, 12, 31).timetuple().tm_yday == 366 else 365
                                    decimal_year = year + day_of_year / days_in_year
                                    date_year_list.append(float(decimal_year))
                                except:
                                    date_year_list.append(np.nan)
                        else:
                            # Try to parse as string
                            try:
                                parsed_date = pd.to_datetime(str(date_val))
                                year = parsed_date.year
                                day_of_year = parsed_date.timetuple().tm_yday
                                days_in_year = 366 if pd.Timestamp(year, 12, 31).timetuple().tm_yday == 366 else 365
                                decimal_year = year + day_of_year / days_in_year
                                date_year_list.append(float(decimal_year))
                            except:
                                date_year_list.append(np.nan)
                    
                    model_data_copy['date_year'] = date_year_list
                    model_data_copy = model_data_copy.dropna(subset=['date_year'])  # Remove rows with invalid dates
                    
                    # Get simulated time values
                    sim_times = (df_filtered['Time yr'] + StartY).unique()
                    
                    # Calculate RMSLE for each analyte
                    for conc in conc_vals:
                        pfas_name = conc_to_pfas.get(conc, conc)
                        
                        # Filter model data for this analyte (match by name)
                        analyte_data = model_data_copy[
                            model_data_copy['analyte'].str.upper().str.strip() == pfas_name.upper().strip()
                        ].copy()

                        analyte_data = analyte_data[analyte_data['well'].isin(well_vals)]
                        
                        if len(analyte_data) > 0:
                            observed_vals = []
                            predicted_vals = []
                            
                            # Match dates and wells
                            for _, row in analyte_data.iterrows():
                                obs_date = row['date_year']
                                obs_well = row['well'] if pd.notna(row['well']) else None
                                obs_conc = row['concentration']
                                
                                # Skip if concentration is invalid
                                if pd.isna(obs_conc) or obs_conc <= 0:
                                    continue
                                
                                # Find closest simulated date
                                closest_sim_date = find_closest_date(obs_date, sim_times)
                                
                                if closest_sim_date is not None:
                                    # Find matching well data
                                    if obs_well:
                                        # Try to match well name
                                        well_match = None
                                        for well in well_vals:
                                            if obs_well.upper().strip() in well.upper().strip() or well.upper().strip() in obs_well.upper().strip():
                                                well_match = well
                                                break
                                        
                                        if well_match:
                                            well_sim_data = df_filtered[
                                                (df_filtered['well_file'] == well_match) & 
                                                (np.abs((df_filtered['Time yr'] + StartY) - closest_sim_date) < 0.01)
                                            ]
                                        else:
                                            # If no well match, use all wells
                                            well_sim_data = df_filtered[
                                                np.abs((df_filtered['Time yr'] + StartY) - closest_sim_date) < 0.01
                                            ]
                                    else:
                                        # If no well specified, use all wells
                                        well_sim_data = df_filtered[
                                            np.abs((df_filtered['Time yr'] + StartY) - closest_sim_date) < 0.01
                                        ]
                                    
                                    if len(well_sim_data) > 0 and conc in well_sim_data.columns:
                                        pred_conc = well_sim_data[conc].mean()  # Average if multiple matches
                                        if pred_conc > 0:
                                            observed_vals.append(obs_conc)
                                            predicted_vals.append(pred_conc)
                            
                            # Calculate RMSLE for this analyte
                            if len(observed_vals) > 0:
                                rmsle_val = calculate_rmsle(
                                    np.array(observed_vals),
                                    np.array(predicted_vals)
                                )
                                if rmsle_val is not None:
                                    rmsle_stats.append({
                                        'analyte': pfas_name,
                                        'rmsle': rmsle_val,
                                        'n_points': len(observed_vals)
                                    })
                
                # Calculate RMSLE for version_flag == 1 using Excel PFAA1/PFAA2 (V34:V41, X34:X41)
                elif version_flag == 1 and excel_data_year is not None and (excel_PFAA1_data or (ncomp == 2 and excel_PFAA2_data)):
                    sim_times = (df_filtered['Time yr'] + StartY).unique() if len(df_filtered) > 0 else np.array([])
                    closest_time = find_closest_date(excel_data_year, sim_times) if len(sim_times) > 0 else None
                    if closest_time is not None:
                        # PFAA1 (C2well): observed from excel_PFAA1_data, predicted from df_obs_MW at closest_time
                        observed_1, predicted_1 = [], []
                        for well in well_vals:
                            obs = excel_PFAA1_data.get(well) if excel_PFAA1_data else None
                            if obs is None and excel_PFAA1_data:
                                for key in excel_PFAA1_data.keys():
                                    if (isinstance(key, tuple) and len(key) > 0 and key[0] == well) or key == well:
                                        obs = excel_PFAA1_data[key]
                                        break
                            if obs is not None and obs > 0:
                                well_data = df_filtered[(df_filtered['well_file'] == well) & (np.abs((df_filtered['Time yr'] + StartY) - closest_time) < 0.01)]
                                if len(well_data) > 0 and 'C2well' in well_data.columns:
                                    pred = well_data['C2well'].mean()
                                    if pred > 0:
                                        observed_1.append(obs)
                                        predicted_1.append(pred)
                        if len(observed_1) > 0:
                            rmsle_1 = calculate_rmsle(np.array(observed_1), np.array(predicted_1))
                            if rmsle_1 is not None:
                                pfas_1 = PFAS_names[1] if len(PFAS_names) > 1 else "PFAA1"
                                rmsle_stats.append({'analyte': pfas_1, 'rmsle': rmsle_1, 'n_points': len(observed_1)})
                        # PFAA2 (C4well) if ncomp == 2
                        if ncomp == 2 and excel_PFAA2_data:
                            observed_2, predicted_2 = [], []
                            for well in well_vals:
                                obs = excel_PFAA2_data.get(well)
                                if obs is None:
                                    for key in excel_PFAA2_data.keys():
                                        if (isinstance(key, tuple) and len(key) > 0 and key[0] == well) or key == well:
                                            obs = excel_PFAA2_data[key]
                                            break
                                if obs is not None and obs > 0:
                                    well_data = df_filtered[(df_filtered['well_file'] == well) & (np.abs((df_filtered['Time yr'] + StartY) - closest_time) < 0.01)]
                                    if len(well_data) > 0 and 'C4well' in well_data.columns:
                                        pred = well_data['C4well'].mean()
                                        if pred > 0:
                                            observed_2.append(obs)
                                            predicted_2.append(pred)
                            if len(observed_2) > 0:
                                rmsle_2 = calculate_rmsle(np.array(observed_2), np.array(predicted_2))
                                if rmsle_2 is not None:
                                    pfas_2 = PFAS_names[3] if len(PFAS_names) > 3 else "PFAA2"
                                    rmsle_stats.append({'analyte': pfas_2, 'rmsle': rmsle_2, 'n_points': len(observed_2)})
                
                # Plot the simulated data
                sim_color_map = {}
                for well in well_vals:
                    well_data = df_filtered[df_filtered['well_file'] == well]
                    for conc in conc_vals:
                        if conc in well_data.columns:
                            pfas_name = conc_to_pfas.get(conc, conc)
                            trace_color = colors[color_idx % len(colors)]
                            sim_color_map[(str(well).strip().upper(), str(pfas_name).strip().upper())] = trace_color
                            fig.add_trace(go.Scatter(
                                x=well_data['Time yr']+StartY,
                                y=well_data[conc],
                                mode='lines+markers',
                                name=f'{well} - {pfas_name}',
                                line=dict(
                                    color=trace_color,
                                    width=3,
                                    shape='spline'
                                ),
                                marker=dict(
                                    size=8,
                                    color=trace_color,
                                    line=dict(width=2, color=trace_color)
                                ),
                                hovertemplate=f'<b>{well} - {pfas_name}</b><br>' +
                                            f'{xlabel}: %{{x}}<br>' +
                                            'Concentration: %{y:.3f} μg/L<br>' +
                                            '<extra></extra>'
                            ))
                            color_idx += 1
                
                # Plot Excel data (V34-V41 for PFAA1, X34-X41 for PFAA2) if available
                # PFAA1 is at index 1, PFAA2 is at index 3 in PFAS_names
                # Only include PFAA2 if ncomp == 2
                PFAA1_name = PFAS_names[1] if len(PFAS_names) > 1 else None
                try:
                    PFAA2_name = PFAS_names[3] if ncomp == 2 and PFAS_names[3] not in ('None', '') else None
                except (IndexError, KeyError):
                    PFAA2_name = None
                
                # Plot single-point observations from §10 V34:V40 / X34:X40.
                # v98: Only fire this block when df_model_data is NOT
                # available.  Otherwise we'd duplicate every PFAA-1
                # marker (diamond from here PLUS square from the
                # df_model_data block below), and we'd miss precursor
                # observations entirely (this block only handles
                # C2well/C4well, not C1well/C3well).
                _has_model_data = (df_model_data is not None
                                   and len(df_model_data) > 0)
                if (excel_data_year is not None
                        and not _has_model_data):
                    excel_time = excel_data_year
                    # Keep legend order consistent with simulated traces: well -> concentration
                    for well in well_vals:
                        for conc in conc_vals:
                            obs_analyte_name = None
                            obs_value = None
                            obs_symbol = 'diamond'

                            if conc == "C2well" and PFAA1_name and excel_PFAA1_data:
                                obs_analyte_name = PFAA1_name
                                if well in excel_PFAA1_data:
                                    obs_value = excel_PFAA1_data[well]
                                else:
                                    for key in excel_PFAA1_data.keys():
                                        if (isinstance(key, tuple) and len(key) > 0 and key[0] == well) or key == well:
                                            obs_value = excel_PFAA1_data[key]
                                            break
                            elif conc == "C4well" and ncomp == 2 and PFAA2_name and excel_PFAA2_data:
                                obs_analyte_name = PFAA2_name
                                if well in excel_PFAA2_data:
                                    obs_value = excel_PFAA2_data[well]
                                else:
                                    for key in excel_PFAA2_data.keys():
                                        if (isinstance(key, tuple) and len(key) > 0 and key[0] == well) or key == well:
                                            obs_value = excel_PFAA2_data[key]
                                            break

                            if obs_analyte_name and obs_value is not None:
                                obs_color = sim_color_map.get(
                                    (str(well).strip().upper(), str(obs_analyte_name).strip().upper()),
                                    colors[color_idx % len(colors)]
                                )
                                fig.add_trace(go.Scatter(
                                    x=[excel_time],
                                    y=[obs_value],
                                    mode='markers',
                                    name=f'{well} - {obs_analyte_name} (obs)',
                                    marker=dict(
                                        size=12,
                                        symbol=obs_symbol,
                                        color=obs_color,
                                        line=dict(width=2, color='black')
                                    ),
                                    hovertemplate=f'<b>{well} - {obs_analyte_name} (obs)</b><br>' +
                                                f'{xlabel}: %{{x:.2f}}<br>' +
                                                'Concentration: %{y:.3f} μg/L<br>' +
                                                '<extra></extra>',
                                    legendgroup=f'{well}_{obs_analyte_name}_obs',
                                    showlegend=True
                                ))
                                color_idx += 1
                # v90: Plot Observation Points from §10 calibration file
                # for BOTH Simple and Detailed when df_model_data is loaded.
                # Previously gated on `version_flag == 2`, so Simple users
                # who imported via §10 button got RMSLE numbers but no
                # observation markers on the chart.
                if (df_model_data is not None
                        and len(df_model_data) > 0
                        and 'model_data_copy' in dir()):
                    selected_analytes = {
                        str(conc_to_pfas.get(conc, conc)).upper().strip()
                        for conc in conc_vals
                    }
                    obs_plot_data = model_data_copy.copy()
                    obs_plot_data = obs_plot_data[
                        obs_plot_data['well'].isin(well_vals)
                    ]
                    obs_plot_data = obs_plot_data[
                        obs_plot_data['analyte'].astype(str).str.upper().str.strip().isin(selected_analytes)
                    ]
                    obs_plot_data = obs_plot_data.dropna(subset=['well', 'analyte', 'date_year', 'concentration'])
                    obs_plot_data['well_key'] = obs_plot_data['well'].astype(str).str.strip().str.upper()
                    obs_plot_data['analyte_key'] = obs_plot_data['analyte'].astype(str).str.strip().str.upper()
                    grouped_obs = {
                        (wk, ak): grp.sort_values('date_year')
                        for (wk, ak), grp in obs_plot_data.groupby(['well_key', 'analyte_key'], sort=False)
                    }

                    # v98: distinct marker symbol per concentration column
                    # so PFAA-1 / PFAA-2 / Precursor-1 / Precursor-2 obs
                    # are visually distinguishable on the same well's
                    # plot (instead of all being squares).
                    SYMBOL_PER_CONC = {
                        "C1well": "circle",        # Precursor-1
                        "C2well": "diamond",       # PFAA-1
                        "C3well": "triangle-up",   # Precursor-2
                        "C4well": "square",        # PFAA-2
                    }
                    # Keep legend order consistent with simulated traces: well -> concentration
                    for well in well_vals:
                        well_key = str(well).strip().upper()
                        for conc in conc_vals:
                            pfas_name = conc_to_pfas.get(conc, conc)
                            analyte_key = str(pfas_name).strip().upper()
                            group = grouped_obs.get((well_key, analyte_key))
                            if group is None or len(group) == 0:
                                continue
                            label = f"{well} - {pfas_name} (obs)"
                            trace_color = sim_color_map.get(
                                (well_key, analyte_key),
                                colors[color_idx % len(colors)]
                            )
                            fig.add_trace(go.Scatter(
                                x=group['date_year'],
                                y=group['concentration'],
                                mode='markers',
                                name=label,
                                marker=dict(
                                    size=12,
                                    symbol=SYMBOL_PER_CONC.get(conc, "square"),
                                    color=trace_color,
                                    line=dict(width=2, color='black')
                                ),
                                hovertemplate=f"<b>{label}</b><br>" +
                                            f'{xlabel}: %{{x:.2f}}<br>' +
                                            'Concentration: %{y:.3f} μg/L<br>' +
                                            '<extra></extra>',
                                legendgroup=f'{well}_{pfas_name}_obs',
                                showlegend=True
                            ))
                            color_idx += 1

        elif analysis_type == 6:
            # Special handling for Mass Discharge
            discharge_cols = [c for c in df_filtered.columns if c.startswith("Discharge")]
            # Filter out Precursor 1 and Precursor 2 if not precursor_flag
            if not precursor_flag:
                discharge_cols = [c for c in discharge_cols if c not in ["Discharge1", "Discharge3"]]
            if precursor_flag and PFAS_names[2] in ('None', ''):
                discharge_cols = [c for c in discharge_cols if c not in ["Discharge3"]]
            try:
                if PFAS_names[3] in ('None', ''):
                    discharge_cols = [c for c in discharge_cols if c != "Discharge4"]
            except (IndexError, KeyError):
                pass
            # Create mapping from discharge column names to PFAS names
            discharge_to_pfas = {
                "Discharge1": PFAS_names[0],
                "Discharge2": PFAS_names[1], 
                "Discharge3": PFAS_names[2],
                "Discharge4": PFAS_names[3]
            }
            
            for i, col in enumerate(discharge_cols):
                pfas_name = discharge_to_pfas.get(col, col)
                fig.add_trace(go.Scatter(
                    x=x,
                    y=df_filtered[col],
                    mode='lines+markers',
                    name=pfas_name,
                    line=dict(
                        color=colors[i % len(colors)],
                        width=3,
                        shape='spline'
                    ),
                    marker=dict(
                        size=8,
                        color=colors[i % len(colors)],
                        line=dict(width=2, color=colors[i % len(colors)])
                    ),
                    hovertemplate=f'<b>{pfas_name}</b><br>' +
                                f'{xlabel}: %{{x}}<br>' +
                                'Mass Discharge: %{y:.2f} kg/yr<br>' +
                                '<extra></extra>'
                ))
        elif analysis_type == 7:
            # Special handling for Mass in T-Zone (hi columns)
            df_filtered = df_obs_D
            x = df_filtered['Time']+StartY
            xlabel = "Time (yr)"
            title = "Mass vs. Time in T-Zone"
            
            hi_cols = [c for c in df_filtered.columns if c.endswith("hi")]
            # Filter out Precursor 1 and Precursor 2 if not precursor_flag
            if not precursor_flag:
                hi_cols = [c for c in hi_cols if c not in ["C1hi", "C3hi"]]
            if precursor_flag and PFAS_names[2] in ('None', ''):
                hi_cols = [c for c in hi_cols if c not in ["C3hi"]]
            try:
                if PFAS_names[3] in ('None', ''):
                    hi_cols = [c for c in hi_cols if c != "C4hi"]
            except (IndexError, KeyError):
                pass
            # Create mapping from column names to PFAS names
            mass_to_pfas = {
                "C1hi": PFAS_names[0],
                "C2hi": PFAS_names[1], 
                "C3hi": PFAS_names[2],
                "C4hi": PFAS_names[3]
            }
            
            for i, col in enumerate(hi_cols):
                pfas_name = mass_to_pfas.get(col, col)
                fig.add_trace(go.Scatter(
                    x=x,
                    y=df_filtered[col],
                    mode='lines+markers',
                    name=pfas_name,
                    line=dict(
                        color=colors[i % len(colors)],
                        width=3,
                        shape='spline'
                    ),
                    marker=dict(
                        size=8,
                        color=colors[i % len(colors)],
                        line=dict(width=2, color=colors[i % len(colors)])
                    ),
                    hovertemplate=f'<b>{pfas_name}</b><br>' +
                                f'{xlabel}: %{{x}}<br>' +
                                'Mass in T-Zone: %{y:.2f} kg<br>' +
                                '<extra></extra>'
                ))
        elif analysis_type == 8:
            # Special handling for Mass in Low-K Zone (low columns)
            df_filtered = df_obs_D
            x = df_filtered['Time']+StartY
            xlabel = "Time (yr)"
            title = "Mass vs. Time in Low-K Zone"
            
            low_cols = [c for c in df_filtered.columns if c.endswith("low")]
            # Filter out Precursor 1 and Precursor 2 if not precursor_flag
            if not precursor_flag:
                low_cols = [c for c in low_cols if c not in ["C1low", "C3low"]]
            if precursor_flag and PFAS_names[2] in ('None', ''):
                low_cols = [c for c in low_cols if c not in ["C3low"]]
            try:
                if PFAS_names[3] in ('None', ''):
                    low_cols = [c for c in low_cols if c != "C4low"]
            except (IndexError, KeyError):
                pass
            # Create mapping from column names to PFAS names
            mass_to_pfas = {
                "C1low": PFAS_names[0],
                "C2low": PFAS_names[1], 
                "C3low": PFAS_names[2],
                "C4low": PFAS_names[3]
            }
            
            for i, col in enumerate(low_cols):
                pfas_name = mass_to_pfas.get(col, col)
                fig.add_trace(go.Scatter(
                    x=x,
                    y=df_filtered[col],
                    mode='lines+markers',
                    name=pfas_name,
                    line=dict(
                        color=colors[i % len(colors)],
                        width=3,
                        shape='spline'
                    ),
                    marker=dict(
                        size=8,
                        color=colors[i % len(colors)],
                        line=dict(width=2, color=colors[i % len(colors)])
                    ),
                    hovertemplate=f'<b>{pfas_name}</b><br>' +
                                f'{xlabel}: %{{x}}<br>' +
                                'Mass in Low-K Zone: %{y:.2f} kg<br>' +
                                '<extra></extra>'
                ))
        elif analysis_type == 9:
            color_idx = 0
            # Filter out Precursor 1 and Precursor 2 if not precursor_flag
            # Only include PFAA2 if ncomp == 2 and PFAA2_name is not None/empty
            mass_list = [SL_Mass_PFAA1]
            pfas_indices = [1]  # PFAA1 index
            if ncomp == 2:
                try:
                    if PFAS_names[3] not in ('None', ''):
                        mass_list.append(SL_Mass_PFAA2)
                        pfas_indices.append(3)  # PFAA2 index
                except (IndexError, KeyError):
                    pass
            if version_flag != 1 and ncomp == 2:
                mass_list.extend([SL_Mass_Precursor1, SL_Mass_Precursor2])
                pfas_indices.extend([0, 2])  # Precursor 1 and Precursor 2 indices
            if version_flag != 1 and ncomp == 1:
                mass_list.extend([SL_Mass_Precursor1])
                pfas_indices.extend([0])  # Precursor 1 index
            

            for mass, pfas_idx in zip(mass_list, pfas_indices):
                pfas_name = PFAS_names[pfas_idx] if pfas_idx < len(PFAS_names) else None
                # Skip if pfas_name is None or empty (e.g., PFAA2_name when G38 is blank)
                if not pfas_name or (isinstance(pfas_name, str) and not pfas_name.strip()):
                    continue
                fig.add_trace(go.Scatter(
                    x=SL_Time,
                    y=mass,
                    mode='lines+markers',
                    name=f'{pfas_name}',
                    line=dict(
                        color=colors[color_idx % len(colors)],
                        width=3,
                        shape='spline'
                    ),
                    marker=dict(
                        size=8,
                        color=colors[color_idx % len(colors)],
                        line=dict(width=2, color=colors[color_idx % len(colors)])
                    ),
                    hovertemplate=f'<b>{pfas_name}</b><br>' +
                                    f'{xlabel}: %{{x}}<br>' +
                                    'Mass: %{y:.2f} kg<br>' +
                                    '<extra></extra>'
                ))
                color_idx += 1
        else:
            # Regular handling for other analysis types
            conc_cols = [c for c in df_filtered.columns if c.startswith("Conc") or c.startswith("Discharge")]
            # Filter out Precursor 1 and Precursor 2 if not precursor_flag
            if not precursor_flag:
                conc_cols = [c for c in conc_cols if c not in ["Conc1", "Conc3", "Discharge1", "Discharge3"]]
            # Filter out Conc4 if PFAS_names[3] is None
            # Shortened using get() to avoid exceptions
            if PFAS_names[3] in ('None', '') if len(PFAS_names) > 3 else False:
                conc_cols = [c for c in conc_cols if c != "Conc4"]
            if PFAS_names[2] in ('None', '') if len(PFAS_names) > 2 else False:
                conc_cols = [c for c in conc_cols if c != "Conc3"]

            for i, col in enumerate(conc_cols):
                # Map concentration columns to PFAS display names
                if col.startswith("Conc"):
                    conc_to_pfas = {
                        "Conc1": PFAS_names[0],
                        "Conc2": PFAS_names[1],
                        #"Conc3": PFAS_names[2],
                    }
                    # Only include Conc4 if PFAS_names[3] is not None
                    if len(PFAS_names) > 3 and PFAS_names[3] not in ('None', ''):
                        conc_to_pfas["Conc4"] = PFAS_names[3]
                    if len(PFAS_names) > 2 and PFAS_names[2] not in ('None', ''):
                        conc_to_pfas["Conc3"] = PFAS_names[2]
                    display_name = conc_to_pfas.get(col, col)
                    hover_label = "Concentration: %{y:.3f} μg/L"
                else:
                    display_name = col
                    hover_label = "%{y:.2f}"
                fig.add_trace(go.Scatter(
                    x=x,
                    y=df_filtered[col],
                    mode='lines+markers',
                    name=display_name,
                    line=dict(
                        color=colors[i % len(colors)],
                        width=3,
                        shape='spline'
                    ),
                    marker=dict(
                        size=8,
                        color=colors[i % len(colors)],
                        line=dict(width=2, color=colors[i % len(colors)])
                    ),
                    hovertemplate=f'<b>{display_name}</b><br>' +
                                f'{xlabel}: %{{x}}<br>' +
                                f'{hover_label}<br>' +
                                '<extra></extra>'
                ))
        
        # Update layout with modern styling
        fig.update_layout(
        title=dict(
            text=title,
            font=dict(size=24, color='#333'),
            x=0.5,
            xanchor='center'
        ),
        xaxis=dict(
            title=dict(
                text=xlabel,
                font=dict(size=16, color='#555')
            ),
            gridcolor='rgba(28, 110, 172, 0.1)',
            showgrid=True,
            zeroline=False,
            linecolor='#e1e5e9',
            linewidth=2
        ),
        yaxis=dict(
            title=dict(
                text=yaxis_title,
                font=dict(size=16, color='#555')
            ),
            gridcolor='rgba(28, 110, 172, 0.1)',
            showgrid=True,
            zeroline=False,
            linecolor='#e1e5e9',
            linewidth=2
        ),
        legend=dict(
            orientation="h",
            yanchor="top",
            y=-0.25,
            xanchor="center",
            x=0.5,
            font=dict(size=14)
        ),
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=60, r=60, t=80, b=120),
        width=None,
        height=500,
        hovermode='x unified',
        hoverlabel=dict(
            bgcolor="white",
            font_size=14,
            font_family="Inter"
        )
        )
        
        # Apply log scale: x-axis for analysis_type 1 (Concentration vs Depth), y-axis for all others
        if log_scale == "log":
            if analysis_type == 1:
                fig.update_xaxes(type="log")
            else:
                fig.update_yaxes(type="log")
        elif analysis_type == 1:
            # Concentration on x (μg/L): show thousandths when linear
            fig.update_xaxes(tickformat=".3f")
        elif analysis_type in (2, 3, 4, 5):
            fig.update_yaxes(tickformat=".3f")
        
    #    # Add some statistics cards
    #    stats = []
    #    for col in conc_cols:
    #        if len(df_filtered[col]) > 0:
    #            max_val = df_filtered[col].max()
    #            mean_val = df_filtered[col].mean()
    #            stats.extend([
    #                html.Div([
    #                    html.Div(f"{max_val:.2f}", className="stat-value"),
    #                    html.Div(f"Max {col}", className="stat-label")
    #                ], className="stat-card"),
    #                html.Div([
    #                    html.Div(f"{mean_val:.2f}", className="stat-value"),
    #                    html.Div(f"Mean {col}", className="stat-label")
    #                ], className="stat-card")
    #            ])
        
        # Create RMSLE statistics cards if available
        rmsle_cards = []
        if analysis_type == 5 and len(rmsle_stats) > 0:
            for stat in rmsle_stats:
                rmsle_cards.append(
                    html.Div([
                        html.Div(f"{stat['rmsle']:.4f}", className="stat-value"),
                        html.Div(f"RMSLE: {stat['analyte']}", className="stat-label"),
                        html.Div(f"({stat['n_points']} points)", style={"fontSize": "0.8rem", "opacity": "0.8", "marginTop": "5px"})
                    ], className="stat-card")
                )
        
        return html.Div([
            html.Div(rmsle_cards, className="stats-container") if rmsle_cards else html.Div(),
            dcc.Graph(figure=fig, style={'height': '500px', 'width': '100%', 'max-width': '100%'})
        ])

    # Run the Dash app server
    import socket
    import time
    import threading
    
    def find_available_port(start_port=8050, max_attempts=10):
        """Find an available port starting from start_port"""
        for i in range(max_attempts):
            port = start_port + i
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            try:
                sock.bind(('127.0.0.1', port))
                sock.close()
                return port
            except OSError:
                continue
        raise RuntimeError(f"Could not find an available port after {max_attempts} attempts")
    
    def open_browser(port):
        """Open browser after a short delay to ensure server is ready"""
        time.sleep(1.5)
        webbrowser.open(f"http://127.0.0.1:{port}/")
    
    # Set up signal handlers for graceful shutdown
    def signal_handler(signum, frame):
        """Handle shutdown signals gracefully"""
        print("\nShutting down dashboard server...")
        sys.exit(0)
    
    # Register signal handlers for Ctrl+C (SIGINT) and termination (SIGTERM)
    # Note: signal.signal() only works in the main thread
    try:
        signal.signal(signal.SIGINT, signal_handler)
        if hasattr(signal, 'SIGTERM'):
            signal.signal(signal.SIGTERM, signal_handler)
    except ValueError:
        # Signal handlers can only be registered in the main thread
        # If called from a thread, skip signal registration (KeyboardInterrupt will still work)
        pass
    
    # Find an available port
    port = find_available_port(8050)
    print(f"Starting dashboard server at http://127.0.0.1:{port}/")
    print("Press Ctrl+C to stop the server and close the command window.")
    
    # Start browser opening in a separate thread
    browser_thread = threading.Thread(target=open_browser, args=(port,))
    browser_thread.daemon = True
    browser_thread.start()
    
    # Run the server (this is a blocking call)
    try:
        app.run(debug=False, host='127.0.0.1', port=port, use_reloader=False)
    except KeyboardInterrupt:
        print("\nShutting down dashboard server...")
        sys.exit(0)

if __name__ == "__main__":
    workbook_path = None
    sheet_name = None
    #if len(sys.argv) > 1:
    #    workbook_path = sys.argv[1]
    #    if len(sys.argv) > 2:
    #        sheet_name = sys.argv[2]
    try:
        main(workbook_path, sheet_name)
    except KeyboardInterrupt:
        # Handle Ctrl+C gracefully
        print("\nDashboard interrupted by user. Exiting...")
        sys.exit(0)
    except Exception as e:
        print(f"Error starting dashboard: {e}")
        import traceback
        traceback.print_exc()
        print("\nPress Ctrl+C to exit...")
        # Wait briefly to allow user to see the error, then exit
        try:
            time.sleep(2)
        except KeyboardInterrupt:
            sys.exit(0)
        sys.exit(1)
